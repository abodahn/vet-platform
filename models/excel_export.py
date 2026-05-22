"""
Excel export helper — Aleefy Platform
Requires: pip install openpyxl
"""

from io import BytesIO
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


def _border():
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def make_workbook(title: str, headers: list, rows: list,
                  sheet_name: str = "Data") -> BytesIO:
    """Build an .xlsx workbook and return it as a BytesIO stream.

    Args:
        title      - report title written in A1
        headers    - list of column header strings
        rows       - list of lists/tuples (each is one data row)
        sheet_name - worksheet tab name
    Returns BytesIO positioned at 0 ready for send_file().
    """
    if not _OPENPYXL_OK:
        raise RuntimeError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]   # Excel sheet names max 31 chars

    # ── Title row ────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font  = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor="1D4ED8")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Generated timestamp ───────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ts_cell.font  = Font(italic=True, size=9, color="6B7280")
    ts_cell.alignment = Alignment(horizontal="right")

    # ── Header row ───────────────────────────────────────────
    HDR_FILL = PatternFill("solid", fgColor="EFF6FF")
    HDR_FONT = Font(bold=True, color="1E40AF", size=10)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.border    = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[3].height = 18

    # ── Data rows ────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=4):
        row_bg = "FFFFFF" if row_idx % 2 == 0 else "F8FAFC"
        fill   = PatternFill("solid", fgColor=row_bg)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _border()
            cell.fill   = fill
            cell.alignment = Alignment(vertical="center")
            # Right-align numbers
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── Auto column widths ───────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        col_data   = [str(ws.cell(r, col_idx).value or "") for r in range(3, ws.max_row + 1)]
        max_len    = max((len(s) for s in col_data), default=10)
        max_len    = max(max_len, len(header))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Summary total row (for numeric last-column) ──────────
    if rows and isinstance(rows[0][-1], (int, float)):
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            cell_data = []
            for r in range(4, total_row):
                v = ws.cell(r, col_idx).value
                if isinstance(v, (int, float)):
                    cell_data.append(v)
            if cell_data:
                sum_cell = ws.cell(row=total_row, column=col_idx, value=sum(cell_data))
                sum_cell.font   = Font(bold=True, color="15803D")
                sum_cell.fill   = PatternFill("solid", fgColor="DCFCE7")
                sum_cell.border = _border()
                sum_cell.alignment = Alignment(horizontal="right")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
