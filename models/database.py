"""
Aleefy — Complete Platform Database
All 55 tables covering every module.  PostgreSQL backend.

Connection strategy
───────────────────
PostgreSQL connections are served from a ThreadedConnectionPool (min=2, max=20).
Every get_db() call checks out a connection; every close() / __exit__ returns it.
This eliminates the ~5-10 ms TCP handshake overhead that previously occurred on
every single HTTP request and lets the app sustain concurrent load without
exhausting PostgreSQL's connection limit.

Cache strategy
──────────────
Hot read-only data (clinic settings, service catalog prices) is served from an
in-process TTL cache (default 5 min) via _cached_query().  Cache is invalidated
explicitly when those tables are mutated.
"""

import sqlite3, hashlib, os, re, threading, time, logging
import bcrypt as _bcrypt
from datetime import datetime, date, timedelta
from typing import Optional

logger = logging.getLogger(__name__)

# Optional Flask integration — get_db() registers connections on the app context
# so they can be released automatically. Absent flask (plain scripts / seeders)
# get_db() behaves exactly as before.
try:
    from flask import g as _flask_g, has_app_context as _has_app_ctx
except ImportError:  # pragma: no cover - flask is a hard dep of the app itself
    _flask_g = None

    def _has_app_ctx():
        return False

_G_CONNS = "_db_conns"

_db_path: str = ""

# ── PostgreSQL connection pool ─────────────────────────────────
_PG_CONFIG: dict = {}
_POOL = None          # psycopg2.pool.ThreadedConnectionPool once configured
_POOL_LOCK = threading.Lock()


_TZ_WARNED = False


def _set_session_timezone(raw_conn) -> None:
    """Make this connection agree with the clinic about what day it is.

    NOW() returns the SERVER's local time and the application compares every
    such timestamp against date.today() -- the machine's local date. Managed
    PostgreSQL (Neon, RDS, most containers) runs in UTC; a clinic in Cairo is
    UTC+3. Between midnight and 03:00 the two disagree about the date, so a row
    written "today" is stamped yesterday and disappears from every today-filtered
    view. Overnight and emergency work happens in exactly that window.

    Caught by the pharmacy history test the moment this session crossed
    midnight: a dispensing done at 01:00 was missing from today's history.

    Failure here is deliberately NOT fatal. A PostgreSQL build without full
    tzdata rejects the zone name, and this module falls back to SQLite whenever
    a connection cannot be made -- so treating an unknown zone as fatal would
    quietly relocate a clinic's records into a local file. Warned once, then the
    server default stands.
    """
    global _TZ_WARNED
    tz = os.environ.get("CLINIC_TIMEZONE", "Africa/Cairo")

    # Named zone first: it is the only form that follows DST correctly.
    try:
        cur = raw_conn.cursor()
        cur.execute("SET TIME ZONE %s", (tz,))
        cur.close()
        return
    except Exception:
        try:
            raw_conn.rollback()
        except Exception:
            pass

    # A PostgreSQL built without full tzdata rejects the name. Fall back to this
    # machine's actual UTC offset, which needs no zone database and is by
    # definition the offset the date.today() comparisons are written against.
    # Fixed rather than DST-aware, so re-derive it per connection: pooled
    # connections are recycled often enough that a DST change lands within a day.
    import datetime as _dt
    off = _dt.datetime.now().astimezone().utcoffset() or _dt.timedelta(0)
    total = int(off.total_seconds())
    sign = "+" if total >= 0 else "-"
    hh, mm = divmod(abs(total) // 60, 60)
    literal = f"{sign}{hh:02d}:{mm:02d}"
    try:
        cur = raw_conn.cursor()
        cur.execute(f"SET TIME ZONE INTERVAL '{literal}' HOUR TO MINUTE")
        cur.close()
        if not _TZ_WARNED:
            _TZ_WARNED = True
            logger.info(
                "this PostgreSQL does not know the zone %r, so the session uses "
                "this machine's offset (%s) instead. Correct today; it will not "
                "follow a daylight-saving change on its own.", tz, literal)
        return
    except Exception as exc:
        try:
            raw_conn.rollback()
        except Exception:
            pass
        if not _TZ_WARNED:
            _TZ_WARNED = True
            logger.warning(
                "could not set the session timezone (%s). Timestamps will use "
                "the server's zone; if that is not the clinic's, records written "
                "after midnight can fall outside 'today'.", exc)


def rollback_quietly(conn) -> None:
    """Undo a failed statement so the connection stays usable.

    PostgreSQL aborts the ENTIRE transaction on any error: every statement after
    a failure returns "current transaction is aborted" until someone rolls back.
    SQLite does not, which is why the pattern

        try:    rows = conn.execute(...)      # optional panel
        except: rows = []                     # never mind

    is harmless on SQLite and catastrophic on PostgreSQL -- one optional panel
    whose table happens not to exist silently kills every query after it on the
    same connection, including the ones the page actually needs.

    Call this in any `except` that intends to CONTINUE using the connection.
    Safe on SQLite (rollback of a read is a no-op) and never raises, because a
    failure to clean up must not replace the error the caller is handling.
    """
    try:
        conn.rollback()
    except Exception:
        pass


def is_postgres() -> bool:
    """True when this process talks to PostgreSQL rather than SQLite.

    The SQL translator handles dialect differences inside a statement, but some
    things have no translation at all -- PRAGMA, sqlite_master, "is the database
    file writable" -- and callers need to branch instead. They were reaching
    into db._PG_CONFIG directly, or more often just assuming SQLite.
    """
    return bool(_PG_CONFIG)

# ── In-process TTL cache ───────────────────────────────────────
# { key: (value, expires_at) }
_CACHE: dict = {}
_CACHE_LOCK = threading.Lock()

def _ckey(key: str) -> str:
    """Namespace every cache key by tenant.

    This process serves many clinics, and the cache is one dict shared by all
    of them. Without this prefix the first clinic to load a page would put its
    row under 'clinic_row' and the next clinic would read it back — showing one
    clinic another clinic's name, logo and tagline. Applied here rather than at
    the call sites so no caller can forget, and so the one existing caller
    (get_clinic) needed no change at all.
    """
    slug = ""
    try:
        from models import tenancy
        slug = tenancy.current()
    except Exception:
        pass
    return f"{slug}\x00{key}" if slug else key

def _cache_get(key: str):
    key = _ckey(key)
    with _CACHE_LOCK:
        entry = _CACHE.get(key)
        if entry and time.monotonic() < entry[1]:
            return entry[0], True
        return None, False

def _cache_set(key: str, value, ttl: int = 300):
    with _CACHE_LOCK:
        _CACHE[_ckey(key)] = (value, time.monotonic() + ttl)

def cache_invalidate(key: str):
    """Call after mutating a cached table so next read is fresh."""
    with _CACHE_LOCK:
        _CACHE.pop(_ckey(key), None)


def configure_postgres(host="localhost", port=5432, dbname="vetclinic",
                       user="postgres", password="",
                       min_conn: int = 2, max_conn: int = 20):
    """Call once at startup to configure the PostgreSQL connection pool.

    Parameters
    ----------
    min_conn : int
        Minimum connections kept alive in the pool (default 2).
    max_conn : int
        Maximum simultaneous connections (default 20).  Raise if you expect
        more than ~15 concurrent Gunicorn workers.
    """
    global _PG_CONFIG, _POOL
    # Pin the session timezone to the clinic's, rather than inheriting the
    # server's.
    #
    # NOW() returns the SERVER's local time, and the application reads every one
    # of those timestamps back against date.today() -- the machine's local date.
    # Managed PostgreSQL (Neon, RDS, most containers) runs in UTC. A clinic in
    # Cairo is UTC+3, so between midnight and 03:00 the two disagree about what
    # day it is: a row written "today" is stamped yesterday and drops out of
    # every today-filtered view. Emergency and overnight work happens in exactly
    # that window.
    #
    # Caught by the pharmacy history test the moment this session crossed
    # midnight -- the same failure _fix_sql_sqlite's comment already describes
    # from the SQLite side. That fix made SQLite agree with the PG *server*;
    # this one makes the PG server agree with the *clinic*.
    #
    # ponytail: the honest fix is storing UTC and converting on display, which
    # is a schema-wide migration. TIMEZONE overrides for a clinic elsewhere.
    # Applied AFTER connecting, not as a connect option. Passing
    # `options=-c timezone=...` makes the whole connection FAIL when the server
    # lacks that zone name -- a build without full tzdata answers
    # `invalid value for parameter "TimeZone"` and refuses -- and this code
    # falls back to SQLite on a failed pool, so one unknown zone name would
    # silently move a clinic's records into a local file. Setting it per
    # connection lets an unsupported zone degrade to the server default with a
    # warning instead of taking the database down.
    _PG_CONFIG = dict(host=host, port=port, dbname=dbname,
                      user=user, password=password)
    try:
        from psycopg2.pool import ThreadedConnectionPool
        _POOL = ThreadedConnectionPool(min_conn, max_conn, **_PG_CONFIG)
        logger.info(
            "PostgreSQL pool ready — min=%d max=%d  (%s@%s/%s)",
            min_conn, max_conn, user, host, dbname,
        )
    except Exception as exc:
        _POOL = None
        # In production a missing PostgreSQL must be loud and fatal — silently
        # writing to an empty SQLite file would look like "working" while losing
        # every record. Outside production, clear _PG_CONFIG so get_db() can
        # actually reach its SQLite branch (it cannot while _PG_CONFIG is set).
        if os.environ.get("FLASK_ENV", "development").lower() == "production":
            logger.error("Could not create PG pool (%s) — will retry per request", exc)
        else:
            # Note: set_path() has not run yet at this point, so _db_path is
            # still empty — don't claim a path we don't know.
            logger.warning(
                "PostgreSQL unavailable (%s) — falling back to SQLite. "
                "This is a DEV-ONLY fallback; production fails hard instead.",
                exc,
            )
            _PG_CONFIG = {}


def set_path(path: str) -> None:
    """Set the SQLite file path.

    NOTE: this does NOT switch engines. _connect() checks the PostgreSQL pool
    first, so on a process that has called configure_postgres() this call has no
    effect on where queries go. Use use_sqlite() when you mean "send queries to
    this file".
    """
    global _db_path
    _db_path = path


def use_sqlite(path: str) -> None:
    """Actually route queries to this SQLite file, whatever was configured before.

    set_path() alone is not enough and the difference is dangerous. _connect()
    consults _POOL / _PG_CONFIG *before* _db_path, so a script that did

        db.set_path("/tmp/throwaway.db")
        wipe_everything()

    on a process connected to PostgreSQL wiped the PRODUCTION database while
    believing it was working on a scratch file. The demo seeder did exactly
    that, and its safety guard could not catch it because the guard validates
    the file path -- which was innocent -- and never asked which engine the
    process was actually pointed at.

    The pool is detached, not closed: the caller (or a test fixture) may restore
    the previous globals afterwards, and closing would invalidate a pool the
    surrounding application is still using.
    """
    global _db_path, _PG_CONFIG, _POOL
    _db_path = path
    _PG_CONFIG = {}
    _POOL = None


# ─────────────────────────────────────────────────────────────────
# PostgreSQL compatibility wrapper
# Makes psycopg2 behave like sqlite3 for all existing query code.
# ─────────────────────────────────────────────────────────────────

_FIX_CACHE: dict = {}

# Single-quoted SQL string literal, '' being the embedded-quote escape.
_SQ_STRING_RE = re.compile(r"('(?:[^']|'')*')")

# INSERT target table, used to decide whether "RETURNING id" can be appended.
_INSERT_TABLE_RE = re.compile(r'\s*INSERT\s+INTO\s+"?(\w+)"?', re.IGNORECASE)

# Tables with no `id` column, so INSERT ... RETURNING id would raise
# UndefinedColumn on PostgreSQL. tests/test_db_layer.py re-scans and fails if a
# new id-less table appears.
#
# rate_hits is created lazily in models/security.py rather than in _SCHEMA, and
# the guard test used to scan _SCHEMA alone — so it was invisible, and every
# throttled request would have failed on PostgreSQL while passing on SQLite,
# which ignores the appended RETURNING. The test now scans the whole codebase.
_TABLES_WITHOUT_ID = frozenset({
    "settings",
    # All three below were invisible to the old guard and are PRE-EXISTING, not
    # new. login_attempts is the serious one: it is written on every failed
    # login, so on PostgreSQL a wrong password raised UndefinedColumn instead of
    # being reported as a wrong password, and the lockout could never engage.
    "login_attempts",
    "petsy_usage",
    "rate_hits",
    # Always opened as plain sqlite3, never through get_db(), so it cannot hit
    # this path today. Listed so it stays safe if that ever changes.
    "tenants",
})


def _fix_sql(sql: str) -> str:
    """Translate SQLite SQL quirks to PostgreSQL."""
    if sql in _FIX_CACHE:
        return _FIX_CACHE[sql]
    s = sql
    # ? -> %s placeholders, but never inside a single-quoted string literal
    # (e.g. "SET message = 'Confirm? reply YES'" must keep its literal ?).
    parts = _SQ_STRING_RE.split(s)
    # split() keeps the captured literals at odd indices; only touch the rest.
    s = "".join(p if i % 2 else p.replace("?", "%s") for i, p in enumerate(parts))
    # SQLite datetime function -> PostgreSQL NOW()
    #
    # Both spellings, and the two-argument form is not optional. Newer code
    # writes datetime('now','localtime') — the reverse translator adds
    # 'localtime' precisely so SQLite stops disagreeing with PostgreSQL about
    # what day it is. Only the one-argument form was handled here, so the
    # two-argument form reached PostgreSQL verbatim as a call to a function it
    # does not have: every CREATE TABLE carrying such a DEFAULT failed, and so
    # did every UPDATE using it. The whole payments schema was unusable on the
    # production engine while passing on SQLite.
    #
    # NOW() is the correct target for both. PostgreSQL evaluates it in the
    # server's own timezone, which is what 'localtime' asks for.
    s = re.sub(r"datetime\(\s*'now'\s*(?:,\s*'localtime'\s*)?\)", "NOW()", s)
    # SQLite date('now') -> PostgreSQL CURRENT_DATE.
    #
    # The sibling of the bug above, and it was still live: date() is a SQLite
    # function and PostgreSQL has no such thing, so any statement using it died
    # with "function date(unknown) does not exist". Two production paths in
    # blueprints/procurement/routes.py use it -- raising a purchase order and
    # marking one received -- which means ordering stock was broken outright on
    # the production engine while every test passed on SQLite.
    #
    # `\bdate\(` cannot match inside "datetime(", which is "date" + "time(",
    # so this stays clear of the rule above regardless of order.
    # CURRENT_DATE needs no reverse rule: SQLite understands it natively.
    s = re.sub(r"\bdate\(\s*'now'\s*(?:,\s*'localtime'\s*)?\)", "CURRENT_DATE", s,
               flags=re.IGNORECASE)
    # SQLite's two-argument MAX/MIN are SCALAR functions; PostgreSQL's MAX/MIN
    # are aggregates only, and the scalar spelling is GREATEST/LEAST. Untranslated
    # this died with "function max(integer, numeric) does not exist" and took out
    # the two places that clamp a number at zero: leave balances in attendance,
    # and point-of-sale stock in petshop. Both are UPDATEs, so on PostgreSQL the
    # sale failed outright rather than writing a wrong number.
    #
    # The argument pattern deliberately excludes parens and commas, so a genuine
    # aggregate over a nested call -- MIN(SUBSTRING(x,1,10)) -- cannot match, and
    # neither can a one-argument aggregate.
    s = re.sub(r"\bMAX\s*\(\s*([^(),]+?)\s*,\s*([^(),]+?)\s*\)", r"GREATEST(\1, \2)",
               s, flags=re.IGNORECASE)
    s = re.sub(r"\bMIN\s*\(\s*([^(),]+?)\s*,\s*([^(),]+?)\s*\)", r"LEAST(\1, \2)",
               s, flags=re.IGNORECASE)
    # SQLite AUTOINCREMENT primary key -> PostgreSQL SERIAL
    s = re.sub(r'\bINTEGER\s+PRIMARY\s+KEY\s+AUTOINCREMENT\b', 'SERIAL PRIMARY KEY', s, flags=re.IGNORECASE)
    # TEXT DEFAULT (NOW()) -> TEXT DEFAULT (NOW()::TEXT), keeping the column TEXT
    # while giving PostgreSQL a default of the right type. Without the cast it
    # refuses the table outright: "column is of type text but default
    # expression is of type timestamp with time zone".
    #
    # The optional group in the middle is load-bearing. The original pattern
    # required TEXT and DEFAULT to be adjacent, so a perfectly ordinary
    # `created_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))` slipped
    # past it and failed at CREATE TABLE on PostgreSQL only. Column constraints
    # are legal between the type and its default, and this codebase writes them.
    s = re.sub(r'\bTEXT((?:\s+(?:NOT\s+NULL|NULL|UNIQUE|COLLATE\s+\w+))*'
               r'\s+DEFAULT\s+\(NOW\(\)\))',
               r"TEXT\1::TEXT", s, flags=re.IGNORECASE)
    # INSERT OR IGNORE -> INSERT ... ON CONFLICT DO NOTHING
    has_ignore = bool(re.search(r'\bINSERT\s+OR\s+IGNORE\b', s, re.IGNORECASE))
    if has_ignore:
        s = re.sub(r'\bINSERT\s+OR\s+IGNORE\b', 'INSERT', s, flags=re.IGNORECASE)
        s = s.rstrip().rstrip(';') + ' ON CONFLICT DO NOTHING'
    # INSERT OR REPLACE has NO faithful generic translation, so it is refused
    # rather than guessed at.
    #
    # This used to rewrite it to "ON CONFLICT DO NOTHING" -- described in the
    # old comment as "(simplified)", which it was not: it is the OPPOSITE
    # instruction. REPLACE means overwrite the existing row; DO NOTHING means
    # keep it. On PostgreSQL that silently turned two real updates into no-ops:
    # saving any clinic setting, and editing a staff leave balance. Both
    # reported success and changed nothing, which is the worst way for a write
    # to fail.
    #
    # DO UPDATE needs a conflict target (ON CONFLICT (cols)) that cannot be
    # inferred from the statement text, so there is nothing correct to emit.
    # Write the upsert explicitly instead -- SQLite has supported
    # `ON CONFLICT(cols) DO UPDATE SET x=excluded.x` since 3.24, so one spelling
    # works on both engines and needs no translation at all.
    if re.search(r'\bINSERT\s+OR\s+REPLACE\b', s, re.IGNORECASE):
        raise ValueError(
            "INSERT OR REPLACE cannot be translated to PostgreSQL faithfully. "
            "Write an explicit upsert instead: "
            "INSERT INTO t(...) VALUES(...) "
            "ON CONFLICT(<key cols>) DO UPDATE SET col=excluded.col. "
            "That syntax is valid on SQLite too, so it needs no translation."
        )
    _FIX_CACHE[sql] = s
    return s


# ─────────────────────────────────────────────────────────────────
# SQLite compatibility — the mirror of _fix_sql().
#
# The house convention is "write SQLite-flavoured SQL, _fix_sql() adapts it to
# PostgreSQL".  Review alone never enforced that: 158 raw %s placeholders plus a
# long tail of ::casts / EXTRACT / INTERVAL / ILIKE / NOW() shipped anyway, most
# of them hidden inside try/except blocks that blank a dashboard statistic
# instead of raising.  _fix_sql_sqlite() translates the other direction so the
# SQLite path (tests, CI, the degraded-mode fallback) survives PG-flavoured SQL.
#
# Correctness rule: anything that cannot be translated *faithfully* is left
# alone so SQLite raises, rather than silently returning a wrong number in a
# financial or clinical report.
# ─────────────────────────────────────────────────────────────────

_SQLITE_FIX_CACHE: dict = {}
# ponytail: flat cap + clear, not an LRU. _FIX_CACHE next door is unbounded;
# both are keyed by SQL text from a fixed set of call sites, so the working set
# is a few hundred entries. Swap for functools.lru_cache if SQL ever gets
# generated from unbounded user input.
_SQLITE_FIX_CACHE_MAX = 5000

# expr::type — type names we can translate exactly. Anything else is left in
# place so SQLite raises "unrecognized token: :" loudly.
_CAST_MAP = {
    "text": "TEXT", "varchar": "TEXT", "char": "TEXT",
    "int": "INTEGER", "int4": "INTEGER", "int8": "INTEGER",
    "integer": "INTEGER", "bigint": "INTEGER", "smallint": "INTEGER",
    "numeric": "REAL", "decimal": "REAL", "real": "REAL",
    "float": "REAL", "float8": "REAL", "double": "REAL",
    "date": "date()", "timestamp": "datetime()",
}
_CAST_RE = re.compile(r"::\s*(\w+)")
_EXTRACT_RE = re.compile(r"\bEXTRACT\s*\(\s*(\w+)\s+FROM\s+", re.IGNORECASE)
_EXTRACT_FMT = {"year": "%Y", "month": "%m", "day": "%d",
                "hour": "%H", "minute": "%M", "dow": "%w"}
_AGE_RE = re.compile(r"^\s*AGE\s*\((.+)\)\s*$", re.IGNORECASE | re.DOTALL)
# '<n> <unit>' interval literals whose SQLite date-modifier spelling is identical.
_INTERVAL_LIT_RE = re.compile(
    r"^'(\d+)\s+(day|days|month|months|year|years|hour|hours"
    r"|minute|minutes|second|seconds)'$", re.IGNORECASE)
_INTERVAL_RE = re.compile(r"([+-])\s*INTERVAL\s+\x00(\d+)\x00", re.IGNORECASE)


def _sqlite_mask(sql: str):
    """Replace every single-quoted literal with a \\x00N\\x00 token.

    Every rule below then operates on text that provably contains no string
    literal, which is what keeps LIKE '%foo%', 'a::b' and strftime('%Y', ...)
    out of harm's way.
    """
    lits: list = []

    def grab(m):
        lits.append(m.group(0))
        return f"\x00{len(lits) - 1}\x00"

    return _SQ_STRING_RE.sub(grab, sql), lits


def _tok(lits: list, literal: str) -> str:
    """Register a literal we are emitting, return its mask token."""
    lits.append(literal)
    return f"\x00{len(lits) - 1}\x00"


def _operand_start(s: str, end: int) -> int:
    """Index at which the expression ending at s[end-1] begins, or -1.

    Handles `col`, `t.col`, `123`, a masked literal, `fn(...)` and a bare
    parenthesised group — i.e. everything a PG cast or interval operator binds
    to in this codebase.
    """
    i = end - 1
    while i >= 0 and s[i].isspace():
        i -= 1
    if i < 0:
        return -1
    if s[i] == ")":
        depth = 0
        while i >= 0:
            if s[i] == ")":
                depth += 1
            elif s[i] == "(":
                depth -= 1
                if depth == 0:
                    break
            i -= 1
        if i < 0:
            return -1
        j = i - 1                      # pull in the function name, if any
        while j >= 0 and (s[j].isalnum() or s[j] == "_"):
            j -= 1
        return j + 1
    j = i
    while j >= 0 and (s[j].isalnum() or s[j] in '_."\x00'):
        j -= 1
    return j + 1 if j < i else -1


def _close_paren(s: str, open_idx: int) -> int:
    """Index of the ')' matching the '(' at open_idx, or -1."""
    depth = 0
    for i in range(open_idx, len(s)):
        if s[i] == "(":
            depth += 1
        elif s[i] == ")":
            depth -= 1
            if depth == 0:
                return i
    return -1


def _sqlite_extract(s: str, lits: list) -> str:
    """EXTRACT(field FROM expr) -> CAST(strftime(...) AS INTEGER)."""
    pos = 0
    while True:
        m = _EXTRACT_RE.search(s, pos)
        if not m:
            return s
        open_idx = s.index("(", m.start())
        close = _close_paren(s, open_idx)
        fmt = _EXTRACT_FMT.get(m.group(1).lower())
        if close < 0 or fmt is None:
            pos = m.end()              # untranslatable -> leave it, SQLite raises
            continue
        inner = s[m.end():close].strip()
        age = _AGE_RE.match(inner)
        if age and fmt == "%Y":
            # PG age(x) == age(current_date, x); EXTRACT(YEAR FROM ...) is the
            # whole-years component. Reproduced exactly: year difference minus
            # 1 when this year's anniversary has not happened yet.
            x = age.group(1)
            y, md = _tok(lits, "'%Y'"), _tok(lits, "'%m%d'")
            now = _tok(lits, "'now'")
            rep = (f"(strftime({y},{now}) - strftime({y},{x})"
                   f" - (strftime({md},{now}) < strftime({md},{x})))")
        elif age:
            pos = m.end()              # AGE() outside the YEAR form: not translatable
            continue
        else:
            rep = f"CAST(strftime({_tok(lits, chr(39) + fmt + chr(39))}, {inner}) AS INTEGER)"
        s = s[:m.start()] + rep + s[close + 1:]
        pos = m.start() + len(rep)


def _sqlite_interval(s: str, lits: list) -> str:
    """expr ± INTERVAL '<n> <unit>' -> date/datetime(expr, '<±n> <unit>')."""
    pos = 0
    while True:
        m = _INTERVAL_RE.search(s, pos)
        if not m:
            return s
        lit = _INTERVAL_LIT_RE.match(lits[int(m.group(2))])
        start = _operand_start(s, m.start())
        if not lit or start < 0:
            pos = m.end()              # compound/odd interval -> leave it, SQLite raises
            continue
        operand = s[start:m.start()].rstrip()
        sign = m.group(1)
        mod = _tok(lits, f"'{sign}{lit.group(1)} {lit.group(2)}'")
        # A date operand must stay a date: datetime() would append " 00:00:00"
        # and break text comparison against 'YYYY-MM-DD' columns.
        bare = operand.strip().upper()
        fn = "date" if bare == "CURRENT_DATE" or bare.startswith("DATE(") else "datetime"
        rep = f"{fn}({operand}, {mod})"
        s = s[:start] + rep + s[m.end():]
        pos = start + len(rep)


def _sqlite_casts(s: str) -> str:
    """expr::type -> CAST(expr AS ...) / date(expr) / datetime(expr)."""
    pos = 0
    while True:
        m = _CAST_RE.search(s, pos)
        if not m:
            return s
        target = _CAST_MAP.get(m.group(1).lower())
        start = _operand_start(s, m.start())
        if target is None or start < 0:
            pos = m.end()              # unknown type -> leave it, SQLite raises
            continue
        expr = s[start:m.start()].rstrip()
        rep = (f"{target[:-2]}({expr})" if target.endswith("()")
               else f"CAST({expr} AS {target})")
        s = s[:start] + rep + s[m.end():]
        pos = start + len(rep)


def _fix_sql_sqlite(sql: str) -> str:
    """Translate PostgreSQL SQL quirks to SQLite. Mirror of _fix_sql()."""
    cached = _SQLITE_FIX_CACHE.get(sql)
    if cached is not None:
        return cached
    s, lits = _sqlite_mask(sql)
    # %s -> ? placeholders. Literals are masked, so LIKE '%foo%' and
    # strftime('%Y', ...) are untouched by construction.
    s = s.replace("%s", "?")
    # ILIKE -> LIKE. SQLite's LIKE is case-insensitive for ASCII only; Arabic
    # is caseless so this is lossless here, but it is NOT a general equivalence.
    s = re.sub(r"\bILIKE\b", "LIKE", s, flags=re.IGNORECASE)
    s = _sqlite_extract(s, lits)
    s = _sqlite_interval(s, lits)
    s = _sqlite_casts(s)
    # SQLite's datetime('now') is UTC; PostgreSQL's NOW() is the SERVER's local
    # time. That divergence is a real bug, not a cosmetic one, because the whole
    # application reads back against local dates — 147 datetime('now') writes
    # against 164 date.today() comparisons. Wherever local and UTC differ (Cairo
    # between midnight and 03:00, and permanently for clinics far enough east),
    # a row written "today" carries yesterday's date and vanishes from today's
    # view. It hid a dispensing from the pharmacy history and re-sent every
    # WhatsApp reminder on every run.
    #
    # Appending 'localtime' makes SQLite agree with PostgreSQL rather than
    # introducing a third behaviour. Applied here, at the one choke point every
    # statement passes through, instead of at 147 call sites.
    #
    # ponytail: the honest fix is storing UTC and converting on display; that is
    # a schema-wide migration. This makes the two engines consistent today.
    s = re.sub(r"\bdatetime\(\x00(\d+)\x00\)",
               lambda m: (f"datetime({_tok(lits, chr(39) + 'now' + chr(39))},"
                          f"{_tok(lits, chr(39) + 'localtime' + chr(39))})")
               if lits[int(m.group(1))].strip("'\"") == "now"
               else m.group(0), s)
    s = s.replace("NOW()",
                  f"datetime({_tok(lits, chr(39) + 'now' + chr(39))},"
                  f"{_tok(lits, chr(39) + 'localtime' + chr(39))})")

    # CURRENT_DATE / CURRENT_TIMESTAMP / CURRENT_TIME are UTC in SQLite and
    # LOCAL in PostgreSQL — the same divergence as datetime('now') above, and
    # fixing only that one left this half broken: a row stamped with local time
    # then compared against CURRENT_DATE missed by a day. Caught by
    # test_cast_end_to_end rather than by reading, which is the point of it.
    _now = _tok(lits, chr(39) + "now" + chr(39))
    _loc = _tok(lits, chr(39) + "localtime" + chr(39))
    for kw, fn in (("CURRENT_TIMESTAMP", "datetime"),
                   ("CURRENT_DATE", "date"),
                   ("CURRENT_TIME", "time")):
        call = f"{fn}({_now},{_loc})"
        # A column DEFAULT must be a constant OR a parenthesised expression.
        # `DEFAULT CURRENT_DATE` is legal because the keyword is special-cased
        # by the parser; `DEFAULT date(...)` without parentheses is a syntax
        # error — which is why the schema already writes DEFAULT (datetime(..)).
        # Handle that form first, then any remaining bare keyword.
        s = re.sub(rf"\bDEFAULT\s+{kw}\b", f"DEFAULT ({call})", s,
                   flags=re.IGNORECASE)
        s = re.sub(rf"\b{kw}\b", call, s)

    out = re.sub(r"\x00(\d+)\x00", lambda m: lits[int(m.group(1))], s)
    if len(_SQLITE_FIX_CACHE) >= _SQLITE_FIX_CACHE_MAX:
        _SQLITE_FIX_CACHE.clear()
    _SQLITE_FIX_CACHE[sql] = out
    return out


class _SQLiteCursor(sqlite3.Cursor):
    """sqlite3.Cursor that runs every statement through _fix_sql_sqlite().

    Subclassing rather than wrapping keeps lastrowid, rowcount, description,
    row_factory, fetch*/iteration and close() as the genuine sqlite3 article.
    """

    def execute(self, sql, params=()):
        return super().execute(_fix_sql_sqlite(sql), params)

    def executemany(self, sql, seq_of_params):
        return super().executemany(_fix_sql_sqlite(sql), seq_of_params)


class _SQLiteConn(sqlite3.Connection):
    """sqlite3.Connection whose cursors translate PG-flavoured SQL.

    `with conn:`, executescript(), commit/rollback/close and the interpreter's
    own reference handling are all inherited unchanged.
    """

    def cursor(self, factory=_SQLiteCursor):
        return super().cursor(factory)

    def execute(self, sql, params=(), _protect=False):
        # _protect is a PostgreSQL savepoint concern (_PGCursor.execute); a
        # failed statement never poisons a SQLite transaction, so it is a no-op
        # here — accepted so shared call sites like hr/routes.py:63 work on both.
        return self.cursor().execute(sql, params)

    def executemany(self, sql, seq_of_params):
        return self.cursor().executemany(sql, seq_of_params)


class _PGCursor:
    """Wraps psycopg2 DictCursor to behave like sqlite3.Cursor.

    Key design: savepoint management uses a *separate* admin cursor so that
    SAVEPOINT/RELEASE statements never overwrite the main cursor's result set.
    """

    def __init__(self, raw_cur, raw_conn):
        self._cur = raw_cur
        self._raw_conn = raw_conn
        self.lastrowid = None
        self.rowcount = 0
        self._sp_seq = 0

    @property
    def description(self):
        """Column metadata for the last query.

        Part of the DB-API that sqlite3.Cursor provides, so anything reading
        column names off a cursor -- the usual "does this table have column X
        yet" check -- worked on SQLite and raised AttributeError on PostgreSQL.
        This class exists to be sqlite3-shaped; leaving a standard attribute off
        it means the difference surfaces as a crash somewhere far away.
        """
        return self._cur.description

    def _new_sp(self) -> str:
        self._sp_seq += 1
        return f"pgsp{abs(id(self)) % 999999}_{self._sp_seq}"

    def _admin(self):
        """Fresh plain cursor for savepoint management (no DictCursor needed)."""
        return self._raw_conn.cursor()

    @staticmethod
    def _clean_params(params):
        """Strip NUL bytes from string parameters — psycopg2 rejects them."""
        if not params:
            return params
        cleaned = []
        for p in params:
            cleaned.append(p.replace('\x00', '') if isinstance(p, str) else p)
        return type(params)(cleaned) if isinstance(params, tuple) else cleaned

    @staticmethod
    def _returning_id_target(fixed: str):
        """Return the INSERT target table if 'RETURNING id' should be appended."""
        m = _INSERT_TABLE_RE.match(fixed)
        if not m or 'RETURNING' in fixed.upper():
            return None
        table = m.group(1).lower()
        return None if table in _TABLES_WITHOUT_ID else table

    def _run(self, fixed, params):
        """Bare execute — no savepoint, one round-trip. Errors propagate."""
        if self._returning_id_target(fixed):
            # params must be None (not ()) when there are none, otherwise
            # psycopg2 interpolates and chokes on any literal % in the SQL.
            self._cur.execute(
                fixed.rstrip().rstrip(';') + ' RETURNING id', params or None
            )
            row = self._cur.fetchone()
            # ON CONFLICT DO NOTHING that hit a conflict returns no row.
            self.lastrowid = row[0] if row else None
        else:
            self._cur.execute(fixed, params or None)
            self.lastrowid = None
        self.rowcount = self._cur.rowcount

    def execute(self, sql, params=(), _protect=False):
        """Run one statement.

        _protect=True wraps the statement in a SAVEPOINT so a failure does not
        poison the surrounding transaction. Only executescript() needs that (it
        runs idempotent DDL that is expected to fail on re-runs); every ordinary
        query runs bare, which is 1 round-trip instead of 3-5.
        """
        fixed = _fix_sql(sql)
        params = self._clean_params(params)
        if not _protect:
            self._run(fixed, params)
            return self

        adm = self._admin()
        sp = self._new_sp()
        adm.execute(f'SAVEPOINT {sp}')
        try:
            self._run(fixed, params)
            adm.execute(f'RELEASE SAVEPOINT {sp}')
        except Exception:
            try:
                adm.execute(f'ROLLBACK TO SAVEPOINT {sp}')
            except Exception:
                logger.warning("ROLLBACK TO SAVEPOINT %s failed", sp, exc_info=True)
            adm.close()
            raise
        adm.close()
        return self

    def executemany(self, sql, params_list):
        fixed = _fix_sql(sql)
        self._cur.executemany(fixed, params_list)
        self.rowcount = self._cur.rowcount
        return self

    def fetchone(self):
        return self._cur.fetchone()

    def fetchall(self):
        return self._cur.fetchall()

    def __iter__(self):
        return iter(self._cur)

    def close(self):
        try:
            self._cur.close()
        except Exception:
            pass


# psycopg2's "no transaction open" status. Imported lazily below rather than at
# module scope, because this module must import cleanly with no psycopg2 at all.
try:                                            # pragma: no cover
    from psycopg2.extensions import TRANSACTION_STATUS_IDLE as _PG_TXN_IDLE
except Exception:                               # pragma: no cover
    _PG_TXN_IDLE = 0


class _PGConn:
    """Wraps a psycopg2 connection to behave like sqlite3.Connection.

    When constructed with a *pool* reference, close() returns the underlying
    connection back to the pool (after a safety rollback) instead of closing
    it — keeping the TCP socket alive for the next request.
    """

    def __init__(self, raw_conn, pool=None):
        import psycopg2.extras
        self._conn = raw_conn
        self._pool = pool
        # A connection can come back from the pool mid-transaction if its
        # previous holder used it after `with conn:` (which commits AND closes).
        # Assigning autocommit then raises "set_session cannot be used inside a
        # transaction" HERE — in the next, innocent caller — which sends anyone
        # debugging it to entirely the wrong place. Roll back first so one
        # misbehaving call site cannot poison the connection for everyone after
        # it. The caller's own bug still needs fixing; this stops it spreading.
        if raw_conn.get_transaction_status() != _PG_TXN_IDLE:
            logger.warning("pooled connection came back mid-transaction — "
                           "rolling back before reuse")
            try:
                raw_conn.rollback()
            except Exception:
                logger.warning("rollback of a dirty pooled connection failed",
                               exc_info=True)
        self._conn.autocommit = False
        self._dict_factory = psycopg2.extras.DictCursor
        self._closed = False
        _set_session_timezone(self._conn)

    def cursor(self):
        return _PGCursor(
            self._conn.cursor(cursor_factory=self._dict_factory),
            self._conn
        )

    def execute(self, sql, params=(), _protect=False):
        cur = self.cursor()
        cur.execute(sql, params, _protect=_protect)
        return cur

    def executemany(self, sql, params_list):
        cur = self.cursor()
        cur.executemany(sql, params_list)
        return cur

    def executescript(self, script):
        """Execute multi-statement DDL script against PostgreSQL.

        Splits on semicolons, translates SQLite syntax via _fix_sql, and runs
        each statement. Failures on individual statements are silently ignored
        so that IF NOT EXISTS and idempotent DDL work correctly on re-runs.
        """
        import re
        # Strip SQL line comments so they don't interfere with splitting
        cleaned = re.sub(r'--[^\n]*', '', script)
        stmts = cleaned.split(';')
        for stmt in stmts:
            stmt = stmt.strip()
            if not stmt:
                continue
            try:
                self.execute(stmt, _protect=True)
            except Exception:
                pass  # IF NOT EXISTS handles duplicates; savepoints keep tx alive

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        """Return connection to the pool (if pooled) or close it.

        Idempotent: many call sites do `with conn:` (which closes) *and* then
        conn.close(). A second putconn() would corrupt the pool.
        """
        if self._closed:
            return
        self._closed = True
        try:
            if self._pool is not None:
                # Must be in a clean state before returning to pool.
                # A rollback is safe even if nothing is open.
                try:
                    self._conn.rollback()
                except Exception:
                    logger.warning("rollback before putconn failed", exc_info=True)
                self._pool.putconn(self._conn)
            else:
                self._conn.close()
        except Exception:
            logger.warning("closing DB connection failed", exc_info=True)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, *args):
        if exc_type:
            self.rollback()
        else:
            self.commit()
        self.close()


def get_db():
    """Check out a connection from the pool (or open a fresh one if pooling
    is unavailable) and return it wrapped as a sqlite3-compatible object.

    Always pair with conn.close() or use as a context manager (with conn:).
    When a Flask app context is active the connection is also registered for
    automatic release by close_context_connections() — see that function for the
    teardown hook app.py must install. Outside an app context (scripts, seeders)
    this is a no-op and the caller stays responsible for close().
    """
    return _track_conn(_connect())


def _track_conn(conn):
    if _flask_g is not None and _has_app_ctx():
        conns = _flask_g.get(_G_CONNS)
        if conns is None:
            conns = []
            setattr(_flask_g, _G_CONNS, conns)
        conns.append(conn)
    return conn


def close_context_connections(exc=None):
    """Release every connection get_db() handed out during this app context.

    Register once in app.py:  app.teardown_appcontext(database.close_context_connections)
    Safe to call twice — _PGConn.close()/sqlite3.close() are both idempotent, and
    the list is popped off `g` here.
    """
    if _flask_g is None or not _has_app_ctx():
        return
    for conn in _flask_g.pop(_G_CONNS, None) or ():
        try:
            conn.close()
        except Exception:
            logger.warning("teardown: failed to release DB connection", exc_info=True)


def current_target() -> dict:
    """Which database get_db() will hand back for the CURRENT request.

    Empty dict = the configured default (legacy single-clinic mode). Callers
    that cache per-database state must key it on this, not on a bare bool —
    see the comment on _ensure_schema_once().
    """
    try:
        from models import tenancy
        return tenancy.target()
    except ImportError:
        return {}


def _schema_target() -> str:
    """Identity of the database a lazy schema check applies to."""
    return repr((_db_path, _PG_CONFIG, current_target()))


def _ensure_schema_once(flag_holder: dict, key: str) -> bool:
    """Does this module still need to build its tables in THIS database?

    A plain `if _ready: return` boolean is correct for exactly one database and
    wrong the moment the process serves a second clinic: the flag latches on
    whichever tenant happened to load first, and every clinic provisioned after
    it is left without those tables — surfacing much later as "no such table"
    from a route that works perfectly for the first clinic.

    Checks only. The caller records success with _schema_done() AFTER its DDL,
    which is what the boolean flags this replaces did — recording up front
    would mark a failed CREATE as done and never retry it.
    """
    return flag_holder.get(key) != _schema_target()


def _schema_done(flag_holder: dict, key: str) -> None:
    """Record that the DDL succeeded against the current database."""
    flag_holder[key] = _schema_target()


def _connect():
    # Multi-tenant: route to the clinic that owns this request. An empty dict
    # means no tenant was resolved, so everything below behaves exactly as it
    # did before tenancy existed.
    tgt = current_target()
    dsn = tgt.get("pg_dsn")
    if dsn:
        return _PGConn(_tenant_pool(dsn).getconn(), pool=_tenant_pool(dsn))

    path = tgt.get("db_path")
    if path:
        return _sqlite_connect(path)

    if _POOL is not None:
        # Fast path: get from pool (no TCP handshake)
        raw = _POOL.getconn()
        return _PGConn(raw, pool=_POOL)

    if _PG_CONFIG:
        # Pool not ready yet (race at startup) — open a direct connection
        import psycopg2
        raw = psycopg2.connect(**_PG_CONFIG)
        return _PGConn(raw, pool=None)

    return _sqlite_connect(_db_path)


# One pool per tenant database. Pools are expensive to build and cheap to keep,
# and a dict keyed by DSN means a clinic's pool is created on its first request
# and reused forever after.
_TENANT_POOLS: dict = {}


def _tenant_pool(dsn: str):
    with _POOL_LOCK:
        pool = _TENANT_POOLS.get(dsn)
        if pool is None:
            from psycopg2.pool import ThreadedConnectionPool
            # Smaller than the single-tenant pool on purpose: N clinics each
            # holding 20 connections would exhaust PostgreSQL's max_connections
            # long before any one of them was busy.
            # ponytail: fixed 2/8 per tenant. Make it configurable when a
            # clinic's traffic actually justifies its own sizing.
            pool = ThreadedConnectionPool(1, 8, dsn=dsn)
            _TENANT_POOLS[dsn] = pool
            logger.info("PostgreSQL pool ready for tenant database")
        return pool


def _sqlite_connect(path: str):
    conn = sqlite3.connect(path, check_same_thread=False, factory=_SQLiteConn)
    conn.row_factory = sqlite3.Row
    # Performance PRAGMAs — applied once per connection
    conn.execute("PRAGMA foreign_keys  = ON")
    conn.execute("PRAGMA journal_mode  = WAL")       # concurrent reads during writes
    conn.execute("PRAGMA synchronous   = NORMAL")    # safe with WAL, 3-5x faster than FULL
    conn.execute("PRAGMA cache_size    = -20000")    # 20 MB page cache (was 2 MB)
    conn.execute("PRAGMA temp_store    = MEMORY")    # temp tables in RAM, not disk
    conn.execute("PRAGMA mmap_size     = 134217728") # 128 MB memory-mapped I/O
    conn.execute("PRAGMA busy_timeout  = 5000")      # wait 5 s instead of failing instantly
    return conn


# ════════════════════════════════════════════════════════════════
# SCHEMA — ALL 55 TABLES
# ════════════════════════════════════════════════════════════════

_SCHEMA = """
-- ── CORE ──────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS clinic (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL DEFAULT 'Aleefy',
    name_ar     TEXT DEFAULT 'اليفي',
    phone       TEXT, email TEXT, address TEXT, address_ar TEXT,
    website     TEXT, tax_number TEXT, license_number TEXT,
    doctor_name TEXT,
    tagline     TEXT,
    logo_data   TEXT,
    -- Instapay: the clinic's own payment handle and QR. Stored in the clinic
    -- ROW, not on disk, for the same reason as the logo -- models/backup.py
    -- copies the database and nothing else, so anything under uploads/ survives
    -- a backup and vanishes on restore.
    instapay_handle TEXT,
    instapay_qr     TEXT,
    -- The ipn.eg payment link. Not a duplicate of the QR: a client reading the
    -- invoice on their OWN phone cannot scan their own screen, and a link sent
    -- by WhatsApp is tappable where an image is not.
    instapay_link   TEXT,
    currency    TEXT DEFAULT 'EGP',
    timezone    TEXT DEFAULT 'Africa/Cairo',
    created_at  TEXT DEFAULT (datetime('now')),
    updated_at  TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS branches (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    clinic_id   INTEGER DEFAULT 1,
    name        TEXT NOT NULL,
    name_ar     TEXT,
    phone       TEXT, address TEXT,
    manager_id  INTEGER,
    is_active   INTEGER DEFAULT 1,
    created_at  TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS departments (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    branch_id   INTEGER DEFAULT 1,
    name        TEXT NOT NULL,
    name_ar     TEXT,
    head_id     INTEGER,
    is_active   INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS users (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    username         TEXT UNIQUE NOT NULL,
    password_hash    TEXT NOT NULL,
    full_name        TEXT,
    full_name_ar     TEXT,
    email            TEXT,
    phone            TEXT,
    role             TEXT NOT NULL DEFAULT 'staff',
    department_id    INTEGER,
    branch_id        INTEGER DEFAULT 1,
    is_active        INTEGER DEFAULT 1,
    theme_preference TEXT DEFAULT 'medical',
    language         TEXT DEFAULT 'en',
    last_login_at    TEXT,
    created_at       TEXT DEFAULT (datetime('now')),
    updated_at       TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS roles (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    name             TEXT UNIQUE NOT NULL,
    display_name     TEXT,
    display_name_ar  TEXT,
    permissions_json TEXT DEFAULT '[]',
    color            TEXT DEFAULT '#1a3a6b',
    created_at       TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS audit_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp   TEXT DEFAULT (datetime('now')),
    user_id     INTEGER,
    username    TEXT, role TEXT, action TEXT, module TEXT,
    entity_type TEXT, entity_id TEXT, details TEXT,
    ip TEXT, user_agent TEXT
);

CREATE TABLE IF NOT EXISTS settings (
    key        TEXT PRIMARY KEY,
    value      TEXT,
    category   TEXT DEFAULT 'general',
    updated_at TEXT DEFAULT (datetime('now')),
    updated_by TEXT
);

CREATE TABLE IF NOT EXISTS user_sessions (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    session_token TEXT UNIQUE,
    user_id       INTEGER,
    username      TEXT, role TEXT, ip TEXT, user_agent TEXT,
    created_at    TEXT DEFAULT (datetime('now')),
    last_seen_at  TEXT DEFAULT (datetime('now')),
    ended_at      TEXT
);

-- ── CRM ───────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS owners (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    full_name         TEXT NOT NULL,
    full_name_ar      TEXT,
    phone             TEXT,
    whatsapp_phone    TEXT,
    email             TEXT,
    address           TEXT,
    address_ar        TEXT,
    preferred_contact TEXT DEFAULT 'WhatsApp',
    preferred_doctor  TEXT,
    preferred_branch  INTEGER DEFAULT 1,
    vip_flag          INTEGER DEFAULT 0,
    outstanding_balance REAL DEFAULT 0.0,
    marketing_consent INTEGER DEFAULT 1,
    notes             TEXT,
    created_by        TEXT,
    created_at        TEXT DEFAULT (datetime('now')),
    updated_at        TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS owner_phones (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    owner_id    INTEGER NOT NULL,
    phone       TEXT NOT NULL,
    label       TEXT DEFAULT 'Mobile',
    is_whatsapp INTEGER DEFAULT 0,
    is_primary  INTEGER DEFAULT 0,
    FOREIGN KEY (owner_id) REFERENCES owners(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS pets (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    owner_id            INTEGER NOT NULL,
    pet_name            TEXT NOT NULL,
    species             TEXT,
    breed               TEXT,
    sex                 TEXT DEFAULT 'Unknown',
    dob                 TEXT,
    weight_kg           REAL,
    color               TEXT,
    microchip_id        TEXT,
    neutered            INTEGER DEFAULT 0,
    allergies           TEXT,
    chronic_conditions  TEXT,
    diet_notes          TEXT,
    insurance_number    TEXT,
    notes               TEXT,
    is_active           INTEGER DEFAULT 1,
    created_at          TEXT DEFAULT (datetime('now')),
    updated_at          TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS pet_attachments (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    pet_id      INTEGER NOT NULL,
    filename    TEXT, filetype TEXT, filedata TEXT,
    caption     TEXT,
    uploaded_by TEXT,
    uploaded_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (pet_id) REFERENCES pets(id) ON DELETE CASCADE
);

-- ── APPOINTMENTS ──────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS appointments (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    owner_id         INTEGER NOT NULL,
    pet_id           INTEGER NOT NULL,
    branch_id        INTEGER DEFAULT 1,
    doctor_id        INTEGER,
    doctor_name      TEXT,
    room             TEXT,
    appointment_type TEXT DEFAULT 'Consultation',
    priority         TEXT DEFAULT 'Normal',
    status           TEXT DEFAULT 'Scheduled',
    channel          TEXT DEFAULT 'Walk-in',
    appt_date        TEXT NOT NULL,
    appt_start       TEXT NOT NULL,
    appt_end         TEXT,
    duration_min     INTEGER DEFAULT 30,
    reason           TEXT,
    symptoms         TEXT,
    notes            TEXT,
    confirmed        INTEGER DEFAULT 0,
    reminder_sent    INTEGER DEFAULT 0,
    checked_in_at    TEXT,
    checked_out_at   TEXT,
    created_by       TEXT,
    created_at       TEXT DEFAULT (datetime('now')),
    updated_at       TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id),
    FOREIGN KEY (pet_id)   REFERENCES pets(id)
);

-- ── MEDICAL RECORDS ───────────────────────────────────────────
CREATE TABLE IF NOT EXISTS visits (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    appointment_id  INTEGER,
    owner_id        INTEGER NOT NULL,
    pet_id          INTEGER NOT NULL,
    doctor_id       INTEGER,
    doctor_name     TEXT,
    branch_id       INTEGER DEFAULT 1,
    room            TEXT,
    visit_date      TEXT NOT NULL,
    visit_type      TEXT DEFAULT 'Consultation',
    status          TEXT DEFAULT 'Open',
    chief_complaint TEXT,
    symptoms        TEXT,
    weight_kg       REAL,
    temp_c          REAL,
    heart_rate      INTEGER,
    respiratory_rate INTEGER,
    notes           TEXT,
    created_by      TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    updated_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id),
    FOREIGN KEY (pet_id)   REFERENCES pets(id)
);

CREATE TABLE IF NOT EXISTS diagnoses (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    visit_id    INTEGER NOT NULL,
    pet_id      INTEGER NOT NULL,
    diagnosis   TEXT NOT NULL,
    diagnosis_code TEXT,
    severity    TEXT DEFAULT 'Moderate',
    is_chronic  INTEGER DEFAULT 0,
    notes       TEXT,
    created_by  TEXT,
    created_at  TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (visit_id) REFERENCES visits(id) ON DELETE CASCADE,
    FOREIGN KEY (pet_id)   REFERENCES pets(id)
);

CREATE TABLE IF NOT EXISTS treatment_plans (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    visit_id    INTEGER NOT NULL,
    pet_id      INTEGER NOT NULL,
    plan_text   TEXT NOT NULL,
    goals       TEXT,
    duration    TEXT,
    followup_in INTEGER,
    followup_unit TEXT DEFAULT 'days',
    created_by  TEXT,
    created_at  TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (visit_id) REFERENCES visits(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS prescriptions (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    visit_id       INTEGER NOT NULL,
    pet_id         INTEGER NOT NULL,
    owner_id       INTEGER NOT NULL,
    prescribed_by  TEXT,
    status         TEXT DEFAULT 'Active',
    notes          TEXT,
    dispensed_at   TEXT,
    created_at     TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (visit_id) REFERENCES visits(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS prescription_items (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    prescription_id INTEGER NOT NULL,
    item_id         INTEGER,
    medication_name TEXT NOT NULL,
    dosage          TEXT,
    frequency       TEXT,
    duration        TEXT,
    route           TEXT DEFAULT 'Oral',
    quantity        REAL DEFAULT 1,
    unit            TEXT DEFAULT 'tablet',
    instructions    TEXT,
    dispensed       INTEGER DEFAULT 0,
    FOREIGN KEY (prescription_id) REFERENCES prescriptions(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS lab_requests (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    visit_id    INTEGER NOT NULL,
    pet_id      INTEGER NOT NULL,
    test_name   TEXT NOT NULL,
    test_code   TEXT,
    priority    TEXT DEFAULT 'Routine',
    status      TEXT DEFAULT 'Pending',
    sample_type TEXT,
    collected_at TEXT,
    notes       TEXT,
    requested_by TEXT,
    created_at  TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (visit_id) REFERENCES visits(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS lab_results (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    lab_request_id  INTEGER NOT NULL,
    pet_id          INTEGER NOT NULL,
    result_text     TEXT,
    result_value    REAL,
    unit            TEXT,
    reference_range TEXT,
    is_abnormal     INTEGER DEFAULT 0,
    reviewed_by     TEXT,
    reviewed_at     TEXT,
    report_data     TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (lab_request_id) REFERENCES lab_requests(id)
);

CREATE TABLE IF NOT EXISTS vaccinations (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    pet_id          INTEGER NOT NULL,
    visit_id        INTEGER,
    vaccine_name    TEXT NOT NULL,
    vaccine_brand   TEXT,
    batch_number    TEXT,
    dose_number     INTEGER DEFAULT 1,
    administered_by TEXT,
    administered_at TEXT NOT NULL,
    next_due_at     TEXT,
    site            TEXT DEFAULT 'Subcutaneous',
    notes           TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (pet_id) REFERENCES pets(id)
);

CREATE TABLE IF NOT EXISTS surgeries (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    pet_id          INTEGER NOT NULL,
    visit_id        INTEGER,
    procedure_name  TEXT NOT NULL,
    surgeon         TEXT,
    anesthetist     TEXT,
    surgery_date    TEXT NOT NULL,
    duration_min    INTEGER,
    anesthesia_type TEXT,
    pre_op_notes    TEXT,
    intra_op_notes  TEXT,
    post_op_notes   TEXT,
    outcome         TEXT DEFAULT 'Successful',
    followup_date   TEXT,
    consent_given   INTEGER DEFAULT 0,
    created_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (pet_id) REFERENCES pets(id)
);

CREATE TABLE IF NOT EXISTS followups (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    visit_id     INTEGER,
    pet_id       INTEGER NOT NULL,
    owner_id     INTEGER NOT NULL,
    due_date     TEXT NOT NULL,
    reason       TEXT,
    status       TEXT DEFAULT 'Pending',
    reminder_sent INTEGER DEFAULT 0,
    completed_at TEXT,
    notes        TEXT,
    created_at   TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (pet_id)   REFERENCES pets(id),
    FOREIGN KEY (owner_id) REFERENCES owners(id)
);

-- ── INVENTORY ─────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS item_categories (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT UNIQUE NOT NULL,
    name_ar     TEXT,
    parent_id   INTEGER,
    description TEXT,
    is_active   INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS items (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    category_id     INTEGER,
    sku             TEXT UNIQUE,
    barcode         TEXT,
    name            TEXT NOT NULL,
    name_ar         TEXT,
    description     TEXT,
    unit            TEXT DEFAULT 'unit',
    cost_price      REAL DEFAULT 0.0,
    sell_price      REAL DEFAULT 0.0,
    reorder_level   REAL DEFAULT 10.0,
    max_stock       REAL DEFAULT 1000.0,
    is_medication   INTEGER DEFAULT 0,
    is_controlled   INTEGER DEFAULT 0,
    requires_rx     INTEGER DEFAULT 0,
    supplier_id     INTEGER,
    storage_notes   TEXT,
    is_active       INTEGER DEFAULT 1,
    created_at      TEXT DEFAULT (datetime('now')),
    updated_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (category_id) REFERENCES item_categories(id)
);

CREATE TABLE IF NOT EXISTS warehouses (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    branch_id   INTEGER DEFAULT 1,
    name        TEXT NOT NULL,
    name_ar     TEXT,
    description TEXT,
    is_active   INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS batches (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id        INTEGER NOT NULL,
    warehouse_id   INTEGER DEFAULT 1,
    batch_number   TEXT,
    lot_number     TEXT,
    manufacture_date TEXT,
    expiry_date    TEXT,
    quantity       REAL DEFAULT 0.0,
    unit_cost      REAL DEFAULT 0.0,
    received_at    TEXT DEFAULT (datetime('now','localtime')),
    received_by    TEXT,
    notes          TEXT,
    FOREIGN KEY (item_id)      REFERENCES items(id),
    FOREIGN KEY (warehouse_id) REFERENCES warehouses(id)
);

CREATE TABLE IF NOT EXISTS stock_movements (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id         INTEGER NOT NULL,
    batch_id        INTEGER,
    warehouse_id    INTEGER DEFAULT 1,
    movement_type   TEXT NOT NULL,  -- in/out/adjustment/transfer/expired/damaged
    quantity        REAL NOT NULL,
    unit_cost       REAL DEFAULT 0.0,
    reference_type  TEXT,           -- visit/purchase/adjustment/etc.
    reference_id    INTEGER,
    notes           TEXT,
    created_by      TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (item_id) REFERENCES items(id)
);

CREATE TABLE IF NOT EXISTS reorder_rules (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id         INTEGER NOT NULL UNIQUE,
    reorder_point   REAL DEFAULT 10.0,
    reorder_qty     REAL DEFAULT 50.0,
    preferred_supplier_id INTEGER,
    auto_suggest    INTEGER DEFAULT 1,
    FOREIGN KEY (item_id) REFERENCES items(id)
);

-- ── PHARMACY ──────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS dosage_templates (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id     INTEGER NOT NULL,
    species     TEXT DEFAULT 'All',
    dosage      TEXT NOT NULL,
    frequency   TEXT,
    route       TEXT DEFAULT 'Oral',
    notes       TEXT,
    FOREIGN KEY (item_id) REFERENCES items(id)
);

CREATE TABLE IF NOT EXISTS dispensing_log (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    prescription_item_id INTEGER,
    item_id             INTEGER NOT NULL,
    batch_id            INTEGER,
    visit_id            INTEGER,
    pet_id              INTEGER,
    quantity            REAL NOT NULL,
    dispensed_by        TEXT,
    dispensed_at        TEXT DEFAULT (datetime('now')),
    notes               TEXT,
    FOREIGN KEY (item_id) REFERENCES items(id)
);

-- ── FINANCE ───────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS invoices (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_number  TEXT UNIQUE NOT NULL,
    owner_id        INTEGER NOT NULL,
    pet_id          INTEGER,
    visit_id        INTEGER,
    branch_id       INTEGER DEFAULT 1,
    doctor_name     TEXT,
    issue_date      TEXT NOT NULL,
    due_date        TEXT,
    status          TEXT DEFAULT 'Unpaid',   -- Unpaid/Paid/Partial/Cancelled
    subtotal        REAL DEFAULT 0.0,
    discount_type   TEXT DEFAULT 'value',
    discount_value  REAL DEFAULT 0.0,
    discount_amount REAL DEFAULT 0.0,
    tax_rate        REAL DEFAULT 0.0,
    tax_amount      REAL DEFAULT 0.0,
    total           REAL DEFAULT 0.0,
    paid_amount     REAL DEFAULT 0.0,
    due_amount      REAL DEFAULT 0.0,
    notes           TEXT,
    created_by      TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    updated_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id)
);

CREATE TABLE IF NOT EXISTS invoice_lines (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id   INTEGER NOT NULL,
    line_type    TEXT DEFAULT 'service',  -- service/product/medication
    item_id      INTEGER,
    description  TEXT NOT NULL,
    quantity     REAL DEFAULT 1.0,
    unit_price   REAL DEFAULT 0.0,
    discount     REAL DEFAULT 0.0,
    total        REAL DEFAULT 0.0,
    FOREIGN KEY (invoice_id) REFERENCES invoices(id) ON DELETE CASCADE
);

-- Estimates (quotes) live in their OWN tables rather than as an invoices row
-- with status='Estimate'. That shortcut was tempting and wrong: 27 queries in
-- this codebase sum invoice money, and at least two filter only on
-- status!='Cancelled'. An estimate stored as an invoice would have been booked
-- as revenue by every one of those, silently. Separate tables mean a forgotten
-- WHERE cannot inflate the books -- the same reasoning as database-per-tenant.
CREATE TABLE IF NOT EXISTS estimates (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    estimate_number TEXT UNIQUE NOT NULL,
    owner_id        INTEGER NOT NULL,
    pet_id          INTEGER,
    visit_id        INTEGER,
    branch_id       INTEGER DEFAULT 1,
    doctor_name     TEXT,
    issue_date      TEXT NOT NULL,
    valid_until     TEXT,
    -- Draft/Sent/Approved/Declined/Expired/Converted
    status          TEXT DEFAULT 'Draft',
    subtotal        REAL DEFAULT 0.0,
    discount_type   TEXT DEFAULT 'value',
    discount_value  REAL DEFAULT 0.0,
    discount_amount REAL DEFAULT 0.0,
    tax_rate        REAL DEFAULT 0.0,
    tax_amount      REAL DEFAULT 0.0,
    total           REAL DEFAULT 0.0,
    notes           TEXT,
    decided_at      TEXT,
    decided_by      TEXT,
    invoice_id      INTEGER,
    created_by      TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    updated_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id)
);

CREATE TABLE IF NOT EXISTS estimate_lines (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    estimate_id  INTEGER NOT NULL,
    line_type    TEXT DEFAULT 'service',
    item_id      INTEGER,
    description  TEXT NOT NULL,
    quantity     REAL DEFAULT 1.0,
    unit_price   REAL DEFAULT 0.0,
    discount     REAL DEFAULT 0.0,
    total        REAL DEFAULT 0.0,
    FOREIGN KEY (estimate_id) REFERENCES estimates(id) ON DELETE CASCADE
);

-- Client money held before there is an invoice to put it against: boarding and
-- surgery deposits. It cannot live in `payments` because invoice_id is NOT NULL
-- there, and it must not be invented as a column on `owners` because a single
-- mutable balance loses the history of how it got that way. Append-only signed
-- rows: +ve took money in, -ve gave it back or spent it. The balance is always
-- SUM(amount), so it can be recomputed from the ledger and cannot drift.
CREATE TABLE IF NOT EXISTS owner_credits (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    owner_id   INTEGER NOT NULL,
    amount     REAL NOT NULL,
    kind       TEXT NOT NULL,          -- deposit / applied / refund
    invoice_id INTEGER,                -- set when kind='applied'
    method     TEXT DEFAULT 'Cash',
    reference  TEXT,
    note       TEXT,
    created_by TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id)
);

CREATE TABLE IF NOT EXISTS payments (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id     INTEGER NOT NULL,
    owner_id       INTEGER NOT NULL,
    amount         REAL NOT NULL,
    method         TEXT DEFAULT 'Cash',   -- Cash/Card/Transfer/Insurance
    channel        TEXT DEFAULT 'Cash',   -- Cash/Visa/Instapay
    reference      TEXT,
    notes          TEXT,
    received_by    TEXT,
    received_at    TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY (invoice_id) REFERENCES invoices(id),
    FOREIGN KEY (owner_id)   REFERENCES owners(id)
);

-- An attempt to collect money, whatever the method. One row per attempt,
-- including the ones that fail: an invoice that shows "paid" with no record of
-- who took the money, when, or by what means cannot be reconciled or disputed,
-- and a failed card attempt that leaves no trace is indistinguishable from one
-- that never happened.
--
-- Separate from `payments` on purpose. `payments` is the LEDGER — money that
-- actually arrived. This is the ATTEMPT log, which includes failures,
-- cancellations and duplicates. Collapsing them would mean either recording
-- money that never arrived, or losing the evidence of what went wrong.
CREATE TABLE IF NOT EXISTS payment_intents (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id      INTEGER NOT NULL,
    owner_id        INTEGER NOT NULL,
    gateway         TEXT NOT NULL DEFAULT 'cash',
    amount          NUMERIC(12,2) NOT NULL,
    currency        TEXT NOT NULL DEFAULT 'EGP',
    -- pending -> succeeded | failed | cancelled ; succeeded -> refunded
    status          TEXT NOT NULL DEFAULT 'pending',
    -- The client's own key for this attempt. UNIQUE, so a double-submitted
    -- form or a retried request returns the FIRST attempt instead of charging
    -- twice. This constraint is the duplicate protection; nothing else is.
    idempotency_key TEXT NOT NULL UNIQUE,
    gateway_ref     TEXT,
    failure_reason  TEXT,
    refunded_amount NUMERIC(12,2) NOT NULL DEFAULT 0,
    created_by      TEXT,
    created_at      TEXT NOT NULL DEFAULT (datetime('now','localtime')),
    updated_at      TEXT,
    FOREIGN KEY (invoice_id) REFERENCES invoices(id),
    FOREIGN KEY (owner_id)   REFERENCES owners(id)
);
CREATE INDEX IF NOT EXISTS idx_pay_intent_invoice ON payment_intents(invoice_id);
CREATE INDEX IF NOT EXISTS idx_pay_intent_status  ON payment_intents(status);

-- Append-only history of everything that happened to an intent. Never updated,
-- never deleted: this is what a dispute, a chargeback or an audit is answered
-- with, and a row that can be edited answers nothing.
CREATE TABLE IF NOT EXISTS payment_events (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    intent_id   INTEGER NOT NULL,
    event       TEXT NOT NULL,
    detail      TEXT,
    actor       TEXT,
    created_at  TEXT NOT NULL DEFAULT (datetime('now','localtime')),
    FOREIGN KEY (intent_id) REFERENCES payment_intents(id)
);
CREATE INDEX IF NOT EXISTS idx_pay_event_intent ON payment_events(intent_id);

CREATE TABLE IF NOT EXISTS expenses (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    branch_id    INTEGER DEFAULT 1,
    category     TEXT,
    description  TEXT NOT NULL,
    amount       REAL NOT NULL,
    vendor       TEXT,
    receipt_ref  TEXT,
    expense_date TEXT NOT NULL,
    notes        TEXT,
    created_by   TEXT,
    created_at   TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS daily_closings (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    branch_id       INTEGER DEFAULT 1,
    closing_date    TEXT NOT NULL,
    cash_sales      REAL DEFAULT 0.0,
    card_sales      REAL DEFAULT 0.0,
    transfer_sales  REAL DEFAULT 0.0,
    total_sales     REAL DEFAULT 0.0,
    total_expenses  REAL DEFAULT 0.0,
    net_revenue     REAL DEFAULT 0.0,
    opening_cash    REAL DEFAULT 0.0,
    closing_cash    REAL DEFAULT 0.0,
    notes           TEXT,
    closed_by       TEXT,
    created_at      TEXT DEFAULT (datetime('now'))
);

-- ── PROCUREMENT ───────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS suppliers (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    name         TEXT NOT NULL,
    name_ar      TEXT,
    contact_name TEXT,
    phone        TEXT,
    email        TEXT,
    address      TEXT,
    tax_number   TEXT,
    payment_terms TEXT DEFAULT 'Net 30',
    notes        TEXT,
    is_active    INTEGER DEFAULT 1,
    created_at   TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS purchase_orders (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    po_number    TEXT UNIQUE NOT NULL,
    supplier_id  INTEGER NOT NULL,
    branch_id    INTEGER DEFAULT 1,
    status       TEXT DEFAULT 'Draft',  -- Draft/Sent/Received/Cancelled
    order_date   TEXT NOT NULL,
    expected_date TEXT,
    received_date TEXT,
    subtotal     REAL DEFAULT 0.0,
    tax_amount   REAL DEFAULT 0.0,
    total        REAL DEFAULT 0.0,
    notes        TEXT,
    created_by   TEXT,
    created_at   TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (supplier_id) REFERENCES suppliers(id)
);

CREATE TABLE IF NOT EXISTS po_lines (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    po_id        INTEGER NOT NULL,
    item_id      INTEGER NOT NULL,
    quantity     REAL NOT NULL,
    unit_cost    REAL DEFAULT 0.0,
    total        REAL DEFAULT 0.0,
    received_qty REAL DEFAULT 0.0,
    FOREIGN KEY (po_id)    REFERENCES purchase_orders(id) ON DELETE CASCADE,
    FOREIGN KEY (item_id)  REFERENCES items(id)
);

-- ── COMMUNICATIONS ────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS reminders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    owner_id        INTEGER NOT NULL,
    pet_id          INTEGER,
    appointment_id  INTEGER,
    reminder_type   TEXT NOT NULL,  -- appointment/followup/vaccine/medication/custom
    message         TEXT,
    channel         TEXT DEFAULT 'WhatsApp',
    scheduled_for   TEXT NOT NULL,
    status          TEXT DEFAULT 'Pending',  -- Pending/Sent/Failed/Cancelled
    sent_at         TEXT,
    api_response    TEXT,
    retry_count     INTEGER DEFAULT 0,
    created_by      TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id)
);

CREATE TABLE IF NOT EXISTS whatsapp_templates (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    name          TEXT UNIQUE NOT NULL,
    scenario      TEXT,         -- appointment/followup/vaccine/invoice/custom
    language      TEXT DEFAULT 'en',
    template_text TEXT NOT NULL,
    variables_json TEXT DEFAULT '[]',
    is_active     INTEGER DEFAULT 1,
    is_default    INTEGER DEFAULT 0,
    created_at    TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS whatsapp_log (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    reminder_id  INTEGER,
    owner_id     INTEGER,
    pet_id       INTEGER,
    phone        TEXT,
    message      TEXT,
    template_name TEXT,
    status       TEXT DEFAULT 'Pending',
    http_status  INTEGER,
    response     TEXT,
    error        TEXT,
    sent_at      TEXT DEFAULT (datetime('now'))
);

-- ── GROOMING ──────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS grooming_services (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL,
    name_ar     TEXT,
    duration_min INTEGER DEFAULT 60,
    price       REAL DEFAULT 0.0,
    species     TEXT DEFAULT 'All',
    is_active   INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS grooming_bookings (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    pet_id        INTEGER NOT NULL,
    owner_id      INTEGER NOT NULL,
    service_id    INTEGER,
    groomer_name  TEXT,
    booking_date  TEXT NOT NULL,
    status        TEXT DEFAULT 'Scheduled',
    notes         TEXT,
    before_photo  TEXT,
    after_photo   TEXT,
    invoice_id    INTEGER,
    created_at    TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (pet_id)   REFERENCES pets(id),
    FOREIGN KEY (owner_id) REFERENCES owners(id)
);

-- ── BOARDING ──────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS boarding_rooms (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL,
    room_type   TEXT DEFAULT 'Standard',   -- Standard/Premium/ICU
    capacity    INTEGER DEFAULT 1,
    price_per_night REAL DEFAULT 0.0,
    is_active   INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS boarding_bookings (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    pet_id        INTEGER NOT NULL,
    owner_id      INTEGER NOT NULL,
    room_id       INTEGER,
    check_in      TEXT NOT NULL,
    check_out     TEXT,
    actual_checkout TEXT,
    status        TEXT DEFAULT 'Booked',
    feeding_instructions TEXT,
    medication_instructions TEXT,
    vet_notes     TEXT,
    invoice_id    INTEGER,
    created_at    TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (pet_id)   REFERENCES pets(id),
    FOREIGN KEY (owner_id) REFERENCES owners(id)
);

-- ── SYSTEM ────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS app_logs (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp   TEXT DEFAULT (datetime('now')),
    severity    TEXT DEFAULT 'INFO',
    module      TEXT,
    message     TEXT,
    details     TEXT,
    username    TEXT,
    ip          TEXT
);

CREATE TABLE IF NOT EXISTS diagnostic_runs (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp      TEXT DEFAULT (datetime('now')),
    run_by         TEXT,
    overall_status TEXT,
    passed         INTEGER DEFAULT 0,
    warnings       INTEGER DEFAULT 0,
    failed         INTEGER DEFAULT 0,
    summary        TEXT,
    details_json   TEXT
);

CREATE TABLE IF NOT EXISTS ai_conversations (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id     INTEGER,
    username    TEXT,
    role        TEXT,
    module      TEXT,
    context_type TEXT,   -- visit/pet/inventory/finance/etc.
    context_id  INTEGER,
    prompt      TEXT,
    response    TEXT,
    model_used  TEXT,
    tokens_used INTEGER,
    action_taken TEXT,
    created_at  TEXT DEFAULT (datetime('now'))
);

-- ── ATTENDANCE & LEAVE MANAGEMENT ────────────────────────────
CREATE TABLE IF NOT EXISTS shifts (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    name          TEXT NOT NULL,
    name_ar       TEXT,
    start_time    TEXT NOT NULL DEFAULT '08:00',
    end_time      TEXT NOT NULL DEFAULT '17:00',
    break_minutes INTEGER DEFAULT 60,
    -- Sun=0 … Sat=6, and Sun-Thu is the Egyptian week. This defaulted to
    -- '1,2,3,4,5' (Mon-Fri), so any shift inserted without days silently got
    -- the American week — and since nothing read the column until now, nobody
    -- could see it. Fixing the seed alone was not enough: the default is what
    -- every other INSERT falls back to.
    days_of_week  TEXT DEFAULT '0,1,2,3,4',
    color         TEXT DEFAULT '#3b82f6',
    is_active     INTEGER DEFAULT 1,
    created_at    TEXT DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS staff_shifts (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id    INTEGER NOT NULL,
    shift_id   INTEGER NOT NULL,
    effective_from TEXT NOT NULL,
    effective_to   TEXT,
    FOREIGN KEY (user_id)  REFERENCES users(id),
    FOREIGN KEY (shift_id) REFERENCES shifts(id)
);

CREATE TABLE IF NOT EXISTS attendance_records (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id         INTEGER NOT NULL,
    username        TEXT,
    full_name       TEXT,
    work_date       TEXT NOT NULL,
    check_in        TEXT,
    check_out       TEXT,
    break_minutes   INTEGER DEFAULT 0,
    hours_worked    REAL DEFAULT 0,
    status          TEXT DEFAULT 'Present',
    notes           TEXT,
    recorded_by     TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    updated_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (user_id) REFERENCES users(id)
);

CREATE TABLE IF NOT EXISTS leave_types (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT UNIQUE NOT NULL,
    name_ar         TEXT,
    days_per_year   REAL DEFAULT 21,
    is_paid         INTEGER DEFAULT 1,
    requires_approval INTEGER DEFAULT 1,
    min_notice_days INTEGER DEFAULT 1,
    max_consecutive INTEGER DEFAULT 30,
    color           TEXT DEFAULT '#6366f1',
    is_active       INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS leave_balances (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id         INTEGER NOT NULL,
    leave_type_id   INTEGER NOT NULL,
    year            INTEGER NOT NULL,
    allocated       REAL DEFAULT 0,
    used            REAL DEFAULT 0,
    pending         REAL DEFAULT 0,
    remaining       REAL DEFAULT 0,
    UNIQUE(user_id, leave_type_id, year),
    FOREIGN KEY (user_id)       REFERENCES users(id),
    FOREIGN KEY (leave_type_id) REFERENCES leave_types(id)
);

CREATE TABLE IF NOT EXISTS leave_requests (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id         INTEGER NOT NULL,
    username        TEXT,
    full_name       TEXT,
    leave_type_id   INTEGER NOT NULL,
    leave_type_name TEXT,
    start_date      TEXT NOT NULL,
    end_date        TEXT NOT NULL,
    days_requested  REAL NOT NULL,
    reason          TEXT,
    status          TEXT DEFAULT 'Pending',
    approved_by     TEXT,
    approved_at     TEXT,
    rejection_reason TEXT,
    attachment_name TEXT,
    created_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (user_id)       REFERENCES users(id),
    FOREIGN KEY (leave_type_id) REFERENCES leave_types(id)
);

CREATE TABLE IF NOT EXISTS public_holidays (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL,
    name_ar     TEXT,
    holiday_date TEXT NOT NULL UNIQUE,
    is_recurring INTEGER DEFAULT 0,
    created_at  TEXT DEFAULT (datetime('now'))
);

CREATE INDEX IF NOT EXISTS idx_attendance_user ON attendance_records(user_id);
CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance_records(work_date);
CREATE INDEX IF NOT EXISTS idx_leave_user      ON leave_requests(user_id);
CREATE INDEX IF NOT EXISTS idx_leave_dates     ON leave_requests(start_date, end_date);

-- ── INDEXES ───────────────────────────────────────────────────
CREATE INDEX IF NOT EXISTS idx_pets_owner         ON pets(owner_id);
CREATE INDEX IF NOT EXISTS idx_appts_date         ON appointments(appt_date);
CREATE INDEX IF NOT EXISTS idx_appts_pet          ON appointments(pet_id);
CREATE INDEX IF NOT EXISTS idx_visits_pet         ON visits(pet_id);
CREATE INDEX IF NOT EXISTS idx_diagnoses_visit    ON diagnoses(visit_id);
CREATE INDEX IF NOT EXISTS idx_prescriptions_visit ON prescriptions(visit_id);
CREATE INDEX IF NOT EXISTS idx_stock_item         ON stock_movements(item_id);
CREATE INDEX IF NOT EXISTS idx_stock_date         ON stock_movements(created_at);
CREATE INDEX IF NOT EXISTS idx_invoices_owner     ON invoices(owner_id);
CREATE INDEX IF NOT EXISTS idx_invoices_date      ON invoices(issue_date);
CREATE INDEX IF NOT EXISTS idx_payments_invoice   ON payments(invoice_id);
CREATE INDEX IF NOT EXISTS idx_reminders_date     ON reminders(scheduled_for);
CREATE INDEX IF NOT EXISTS idx_batches_expiry     ON batches(expiry_date);
CREATE INDEX IF NOT EXISTS idx_owners_phone       ON owners(phone);
CREATE INDEX IF NOT EXISTS idx_owners_name        ON owners(full_name);
-- hot FK joins (detail rows fetched per parent record)
CREATE INDEX IF NOT EXISTS idx_appts_owner        ON appointments(owner_id);
CREATE INDEX IF NOT EXISTS idx_visits_owner       ON visits(owner_id);
CREATE INDEX IF NOT EXISTS idx_treatment_visit    ON treatment_plans(visit_id);
CREATE INDEX IF NOT EXISTS idx_rx_items_rx        ON prescription_items(prescription_id);
CREATE INDEX IF NOT EXISTS idx_labreq_visit       ON lab_requests(visit_id);
CREATE INDEX IF NOT EXISTS idx_labres_request     ON lab_results(lab_request_id);
CREATE INDEX IF NOT EXISTS idx_vaccinations_pet   ON vaccinations(pet_id);
CREATE INDEX IF NOT EXISTS idx_surgeries_pet      ON surgeries(pet_id);
CREATE INDEX IF NOT EXISTS idx_followups_pet      ON followups(pet_id);
CREATE INDEX IF NOT EXISTS idx_invlines_invoice   ON invoice_lines(invoice_id);
CREATE INDEX IF NOT EXISTS idx_payments_owner     ON payments(owner_id);
CREATE INDEX IF NOT EXISTS idx_po_lines_po        ON po_lines(po_id);
-- date-range report / due-list filters
CREATE INDEX IF NOT EXISTS idx_visits_date        ON visits(visit_date);
CREATE INDEX IF NOT EXISTS idx_payments_date      ON payments(received_at);
CREATE INDEX IF NOT EXISTS idx_expenses_date      ON expenses(expense_date);
CREATE INDEX IF NOT EXISTS idx_followups_due      ON followups(due_date);
CREATE INDEX IF NOT EXISTS idx_vaccinations_due   ON vaccinations(next_due_at);

-- ── NOTIFICATIONS ─────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS notifications (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    recipient_id INTEGER NOT NULL,
    recipient_role TEXT,
    title        TEXT NOT NULL,
    body         TEXT,
    icon         TEXT DEFAULT '🔔',
    link         TEXT,
    module       TEXT,
    entity_type  TEXT,
    entity_id    INTEGER,
    is_read      INTEGER DEFAULT 0,
    created_at   TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (recipient_id) REFERENCES users(id)
);
CREATE INDEX IF NOT EXISTS idx_notif_recipient ON notifications(recipient_id, is_read);

-- ── SERVICE / PRICE CATALOG ───────────────────────────────────
CREATE TABLE IF NOT EXISTS service_catalog (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    code         TEXT UNIQUE,
    name         TEXT NOT NULL,
    name_ar      TEXT,
    category     TEXT DEFAULT 'Consultation',
    description  TEXT,
    standard_price REAL DEFAULT 0,
    tax_rate     REAL DEFAULT 0,
    duration_min INTEGER DEFAULT 0,
    species      TEXT DEFAULT 'All',
    is_active    INTEGER DEFAULT 1,
    sort_order   INTEGER DEFAULT 0,
    created_at   TEXT DEFAULT (datetime('now')),
    updated_at   TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_svc_category ON service_catalog(category, is_active);

-- ── REMINDER RUNS (deduplication) ─────────────────────────────
CREATE TABLE IF NOT EXISTS reminder_runs (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    run_type     TEXT NOT NULL,
    entity_id    INTEGER,
    entity_type  TEXT,
    status       TEXT DEFAULT 'sent',
    run_at       TEXT DEFAULT (datetime('now')),
    UNIQUE(run_type, entity_id, entity_type)
);

-- ── FILE ATTACHMENTS ──────────────────────────────────────────
CREATE TABLE IF NOT EXISTS attachments (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    entity_type  TEXT NOT NULL,
    entity_id    INTEGER NOT NULL,
    filename     TEXT NOT NULL,
    original_name TEXT,
    mime_type    TEXT,
    size_bytes   INTEGER DEFAULT 0,
    category     TEXT DEFAULT 'general',
    caption      TEXT,
    uploaded_by  TEXT,
    uploaded_at  TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_attach_entity ON attachments(entity_type, entity_id);

-- ── BUDGET TARGETS ───────────────────────────────────────────
CREATE TABLE IF NOT EXISTS budget_targets (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    category    TEXT NOT NULL UNIQUE,
    monthly_egp REAL NOT NULL DEFAULT 0,
    updated_by  TEXT,
    updated_at  TEXT DEFAULT (datetime('now'))
);

-- ── LOYALTY POINTS ────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS loyalty_points (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    owner_id    INTEGER NOT NULL,
    points      INTEGER NOT NULL,
    reason      TEXT,
    ref_type    TEXT DEFAULT 'manual',
    ref_id      INTEGER,
    created_by  TEXT,
    created_at  TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_loyalty_owner ON loyalty_points(owner_id);

-- ── INPATIENT / HOSPITALISATION ───────────────────────────────
CREATE TABLE IF NOT EXISTS inpatient_stays (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    pet_id          INTEGER NOT NULL,
    owner_id        INTEGER NOT NULL,
    visit_id        INTEGER,
    ward            TEXT DEFAULT 'General',
    cage_number     TEXT,
    admitted_by     INTEGER NOT NULL,
    reason          TEXT NOT NULL,
    diagnosis       TEXT,
    treatment_plan  TEXT,
    status          TEXT NOT NULL DEFAULT 'Admitted',
    admitted_at     TEXT DEFAULT (datetime('now')),
    expected_discharge DATE,
    discharged_at   TEXT,
    discharge_notes TEXT,
    daily_rate      NUMERIC(10,2) DEFAULT 0,
    created_at      TEXT DEFAULT (datetime('now')),
    updated_at      TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (pet_id)   REFERENCES pets(id),
    FOREIGN KEY (owner_id) REFERENCES owners(id),
    FOREIGN KEY (admitted_by) REFERENCES users(id)
);
CREATE INDEX IF NOT EXISTS idx_inpatient_pet    ON inpatient_stays(pet_id);
CREATE INDEX IF NOT EXISTS idx_inpatient_status ON inpatient_stays(status);

CREATE TABLE IF NOT EXISTS inpatient_rounds (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    stay_id     INTEGER NOT NULL,
    recorded_by INTEGER NOT NULL,
    round_time  TEXT DEFAULT (datetime('now')),
    temp_c      REAL,
    heart_rate  INTEGER,
    resp_rate   INTEGER,
    weight_kg   REAL,
    pain_score  INTEGER,
    food_intake TEXT,
    fluid_input REAL,
    fluid_output REAL,
    observations TEXT,
    treatment_given TEXT,
    created_at  TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (stay_id) REFERENCES inpatient_stays(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS inpatient_meds (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    stay_id     INTEGER NOT NULL,
    given_by    INTEGER,
    medication  TEXT NOT NULL,
    dose        TEXT,
    route       TEXT DEFAULT 'PO',
    given_at    TEXT DEFAULT (datetime('now')),
    notes       TEXT,
    FOREIGN KEY (stay_id) REFERENCES inpatient_stays(id) ON DELETE CASCADE
);

-- ── PRODUCTION LOGGING TABLES ─────────────────────────────────

CREATE TABLE IF NOT EXISTS backend_logs (
    id                        INTEGER PRIMARY KEY AUTOINCREMENT,
    correlation_id            TEXT,
    request_id                TEXT,
    user_id                   INTEGER,
    username                  TEXT,
    level                     TEXT DEFAULT 'INFO',
    module_name               TEXT,
    action_name               TEXT,
    http_method               TEXT,
    endpoint                  TEXT,
    status_code               INTEGER,
    duration_ms               INTEGER,
    ip_address                TEXT,
    user_agent                TEXT,
    request_payload_summary   TEXT,
    response_payload_summary  TEXT,
    error_message             TEXT,
    stack_trace               TEXT,
    metadata                  TEXT DEFAULT '{}',
    created_at                TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_blog_level    ON backend_logs(level);
CREATE INDEX IF NOT EXISTS idx_blog_created  ON backend_logs(created_at);
CREATE INDEX IF NOT EXISTS idx_blog_endpoint ON backend_logs(endpoint);
CREATE INDEX IF NOT EXISTS idx_blog_user     ON backend_logs(user_id);

CREATE TABLE IF NOT EXISTS frontend_logs (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    correlation_id    TEXT,
    session_id        TEXT,
    user_id           INTEGER,
    username          TEXT,
    level             TEXT DEFAULT 'INFO',
    page_url          TEXT,
    route_name        TEXT,
    component_name    TEXT,
    event_name        TEXT,
    message           TEXT,
    browser_name      TEXT,
    browser_version   TEXT,
    device_type       TEXT,
    os_name           TEXT,
    network_status    TEXT DEFAULT 'online',
    api_endpoint      TEXT,
    api_status_code   INTEGER,
    error_stack       TEXT,
    metadata          TEXT DEFAULT '{}',
    created_at        TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_flog_level   ON frontend_logs(level);
CREATE INDEX IF NOT EXISTS idx_flog_created ON frontend_logs(created_at);
CREATE INDEX IF NOT EXISTS idx_flog_user    ON frontend_logs(user_id);

CREATE TABLE IF NOT EXISTS audit_logs (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    correlation_id TEXT,
    user_id        INTEGER,
    username       TEXT,
    action_type    TEXT NOT NULL,
    entity_name    TEXT,
    entity_id      TEXT,
    old_value      TEXT,
    new_value      TEXT,
    ip_address     TEXT,
    user_agent     TEXT,
    created_at     TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_alog_action  ON audit_logs(action_type);
CREATE INDEX IF NOT EXISTS idx_alog_entity  ON audit_logs(entity_name, entity_id);
CREATE INDEX IF NOT EXISTS idx_alog_user    ON audit_logs(user_id);
CREATE INDEX IF NOT EXISTS idx_alog_created ON audit_logs(created_at);

-- ── OFFLINE SYNC TABLES ───────────────────────────────────────

CREATE TABLE IF NOT EXISTS sync_queue (
    id                 TEXT PRIMARY KEY,
    local_uuid         TEXT NOT NULL,
    server_uuid        TEXT,
    device_id          TEXT NOT NULL,
    user_id            INTEGER,
    entity_name        TEXT NOT NULL,
    operation_type     TEXT NOT NULL,
    payload            TEXT NOT NULL,
    status             TEXT DEFAULT 'PENDING',
    retry_count        INTEGER DEFAULT 0,
    last_error         TEXT,
    priority           INTEGER DEFAULT 5,
    created_offline_at TEXT,
    last_attempt_at    TEXT,
    synced_at          TEXT,
    created_at         TEXT DEFAULT (datetime('now')),
    updated_at         TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_sq_status    ON sync_queue(status);
CREATE INDEX IF NOT EXISTS idx_sq_device    ON sync_queue(device_id);
CREATE INDEX IF NOT EXISTS idx_sq_entity    ON sync_queue(entity_name);
CREATE INDEX IF NOT EXISTS idx_sq_priority  ON sync_queue(priority, created_offline_at);

CREATE TABLE IF NOT EXISTS sync_conflicts (
    id                TEXT PRIMARY KEY,
    sync_queue_id     TEXT NOT NULL,
    entity_name       TEXT,
    local_payload     TEXT,
    server_payload    TEXT,
    conflict_type     TEXT,
    resolution_status TEXT DEFAULT 'PENDING',
    resolved_by       TEXT,
    resolved_at       TEXT,
    created_at        TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_sc_status ON sync_conflicts(resolution_status);

CREATE TABLE IF NOT EXISTS devices (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    device_id     TEXT UNIQUE NOT NULL,
    device_name   TEXT,
    branch_id     INTEGER,
    user_id       INTEGER,
    platform      TEXT,
    app_version   TEXT,
    last_online_at TEXT,
    last_sync_at  TEXT,
    is_active     INTEGER DEFAULT 1,
    created_at    TEXT DEFAULT (datetime('now')),
    updated_at    TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_dev_device ON devices(device_id);
CREATE INDEX IF NOT EXISTS idx_dev_user   ON devices(user_id);

-- Tasks: "call this owner about the lab result", "chase the unpaid invoice",
-- "recheck the dressing on Thursday". The clinic owner asked for a Tasks icon
-- on the exam screen alongside History and Invoices; nothing of the sort
-- existed anywhere in the platform.
--
-- owner_id and pet_id are nullable on purpose: plenty of clinic work ("order
-- more suture") belongs to nobody in particular, and a task that could only
-- exist against a client would push people back to paper for the rest.
CREATE TABLE IF NOT EXISTS tasks (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    title         TEXT NOT NULL,
    details       TEXT,
    owner_id      INTEGER,
    pet_id        INTEGER,
    visit_id      INTEGER,
    assigned_to   TEXT,
    priority      TEXT DEFAULT 'Normal',   -- Low/Normal/High
    status        TEXT DEFAULT 'Open',     -- Open/Done
    due_date      TEXT,
    done_at       TEXT,
    done_by       TEXT,
    created_by    TEXT,
    created_at    TEXT DEFAULT (datetime('now')),
    updated_at    TEXT DEFAULT (datetime('now')),
    FOREIGN KEY (owner_id) REFERENCES owners(id)
);
CREATE INDEX IF NOT EXISTS idx_tasks_owner  ON tasks(owner_id);
CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status, due_date);
CREATE INDEX IF NOT EXISTS idx_tasks_assign ON tasks(assigned_to, status);
"""

# ── Seed data ──────────────────────────────────────────────────
_SEED_ROLES = [
    ("super_admin",    "Super Administrator",   "مدير النظام الأعلى",    "#dc2626"),
    ("clinic_owner",   "Clinic Owner",          "صاحب العيادة",           "#7c3aed"),
    ("branch_manager", "Branch Manager",        "مدير الفرع",             "#1d4ed8"),
    ("doctor",         "Doctor / Veterinarian", "طبيب بيطري",             "#0891b2"),
    ("nurse",          "Nurse / Technician",    "ممرض / تقني",            "#0d9488"),
    ("reception",      "Receptionist",          "موظف استقبال",           "#ca8a04"),
    ("inventory_mgr",  "Inventory Manager",     "مدير المخزون",           "#b45309"),
    ("pharmacist",     "Pharmacist",            "صيدلاني",                "#7c3aed"),
    ("finance",        "Finance User",          "موظف مالية",             "#166534"),
    ("hr",             "HR Officer",            "موظف الموارد البشرية",    "#7e22ce"),
    ("groomer",        "Groomer",               "موظف تجميل",             "#be185d"),
    ("boarding_staff", "Boarding Staff",        "موظف الإيواء",           "#6b7280"),
    ("support_admin",  "Support Admin",         "مدير الدعم الفني",       "#374151"),
    ("auditor",        "Read-only Auditor",     "مدقق للقراءة فقط",       "#6b7280"),
]

_SEED_CATEGORIES = [
    ("Medications", "أدوية"), ("Vaccines", "تطعيمات"),
    ("Consumables", "مستهلكات"), ("Surgical Materials", "مواد جراحية"),
    ("Lab Materials", "مواد مخبرية"), ("Grooming Products", "منتجات تجميل"),
    ("Pet Food", "غذاء حيوانات"), ("Pet Accessories", "إكسسوارات"),
    ("Cleaning", "مواد تنظيف"), ("Office Supplies", "مستلزمات مكتبية"),
]

_SEED_WA_TEMPLATES = [
    ("appointment_reminder", "appointment", "en",
     "Dear {owner_name}, this is a reminder for {pet_name}'s appointment at {clinic_name} on {date} at {time}. Please confirm by replying YES. Thank you!"),
    ("appointment_confirmation", "appointment", "en",
     "Your appointment for {pet_name} at {clinic_name} on {date} at {time} is confirmed. See you soon!"),
    ("followup_reminder", "followup", "en",
     "Dear {owner_name}, it's time for {pet_name}'s follow-up visit at {clinic_name}. Please call us to schedule at your convenience."),
    ("vaccine_due", "vaccine", "en",
     "Dear {owner_name}, {pet_name} is due for {vaccine_name} vaccination. Please contact {clinic_name} to schedule. Stay ahead of preventive care!"),
    ("invoice_sent", "invoice", "en",
     "Dear {owner_name}, your invoice #{invoice_number} for {amount} EGP is ready. Please contact us for payment details. Thank you!"),
    ("appointment_reminder_ar", "appointment", "ar",
     "عزيزي {owner_name}، تذكير بموعد {pet_name} في {clinic_name} يوم {date} الساعة {time}. يرجى التأكيد بالرد بـ نعم. شكراً!"),
]


def _try_stmt(conn, sql, params=()):
    """Run a statement that is *expected* to fail once already applied (idempotent
    migrations / probes for tables that may not exist yet).

    On PostgreSQL this uses the SAVEPOINT path so a failure does not abort the
    surrounding transaction — ordinary execute() no longer takes savepoints.
    Returns the cursor, or None if the statement failed.
    """
    try:
        if isinstance(conn, _PGConn):
            return conn.execute(sql, params, _protect=True)
        return conn.execute(sql, params)
    except Exception as exc:
        logger.debug("idempotent statement skipped: %s (%s)", sql.strip()[:120], exc)
        return None


def _run_pg_migrations(conn) -> None:
    """Create any tables/columns that were added after initial PostgreSQL setup.
    Safe to run on every startup — all statements use IF NOT EXISTS / try-except.
    """
    # Budget targets table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS budget_targets (
            id          SERIAL PRIMARY KEY,
            category    TEXT NOT NULL UNIQUE,
            monthly_egp REAL NOT NULL DEFAULT 0,
            updated_by  TEXT,
            updated_at  TIMESTAMP DEFAULT NOW()
        )
    """)
    # Loyalty points table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS loyalty_points (
            id          SERIAL PRIMARY KEY,
            owner_id    INTEGER NOT NULL,
            points      INTEGER NOT NULL,
            reason      TEXT,
            ref_type    TEXT DEFAULT 'manual',
            ref_id      INTEGER,
            created_by  TEXT,
            created_at  TIMESTAMP DEFAULT NOW(),
            FOREIGN KEY (owner_id) REFERENCES owners(id) ON DELETE CASCADE
        )
    """)
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_loyalty_owner ON loyalty_points(owner_id)"
    )
    # Tasks. _SCHEMA declares this with INTEGER PRIMARY KEY AUTOINCREMENT, which
    # is SQLite-only, so without this the Tasks tab would work on the dev box and
    # 500 on every real clinic.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            id            SERIAL PRIMARY KEY,
            title         TEXT NOT NULL,
            details       TEXT,
            owner_id      INTEGER,
            pet_id        INTEGER,
            visit_id      INTEGER,
            assigned_to   TEXT,
            priority      TEXT DEFAULT 'Normal',
            status        TEXT DEFAULT 'Open',
            due_date      TEXT,
            done_at       TEXT,
            done_by       TEXT,
            created_by    TEXT,
            created_at    TIMESTAMP DEFAULT NOW(),
            updated_at    TIMESTAMP DEFAULT NOW()
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tasks_owner  ON tasks(owner_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status, due_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tasks_assign ON tasks(assigned_to, status)")
    # loyalty_balance column on owners
    _try_stmt(conn, "ALTER TABLE owners ADD COLUMN loyalty_balance INTEGER DEFAULT 0")


def init_db(admin_user: str = "admin", admin_pass: str = "admin1234") -> None:
    _dir = os.path.dirname(_db_path)
    if _dir:
        os.makedirs(_dir, exist_ok=True)
    conn = get_db()
    with conn:
        conn.executescript(_SCHEMA)
        # PostgreSQL-mode migrations: create tables that were added after initial
        # schema.
        #
        # Gated on the CONNECTION, not on _PG_CONFIG. Those are different things
        # under multi-tenancy: _connect() routes to the current tenant first, so
        # on a PostgreSQL deployment provisioning a SQLite-backed clinic gives a
        # _SQLiteConn while _PG_CONFIG is still set. The old check then fed
        # `SERIAL PRIMARY KEY` and `TIMESTAMP DEFAULT NOW()` to SQLite, which
        # died with `near "(": syntax error` -- so provisioning a SQLite clinic
        # rolled back and failed on exactly the deployments that have more than
        # one clinic.
        if isinstance(conn, _PGConn):
            _run_pg_migrations(conn)
        # SOAP columns migration (safe: ADD COLUMN is idempotent via try/except)
        for _col, _type in [
            ("soap_subjective", "TEXT"),
            ("soap_objective",  "TEXT"),
            ("soap_assessment", "TEXT"),
            ("soap_plan",       "TEXT"),
        ]:
            _try_stmt(conn, f"ALTER TABLE visits ADD COLUMN {_col} {_type}")
        # Loyalty balance column on owners
        _try_stmt(conn, "ALTER TABLE owners ADD COLUMN loyalty_balance INTEGER DEFAULT 0")
        # Instapay handle and QR — added after clinics were already running, so
        # existing databases need them backfilled here, not only in _SCHEMA.
        for _col in ("instapay_handle", "instapay_qr", "instapay_link"):
            _try_stmt(conn, f"ALTER TABLE clinic ADD COLUMN {_col} TEXT")
        # Pet insurance columns
        for _col, _type in [
            ("insurance_provider", "TEXT"),
            ("policy_number",      "TEXT"),
            ("policy_expiry",      "TEXT"),
        ]:
            _try_stmt(conn, f"ALTER TABLE pets ADD COLUMN {_col} {_type}")
        # Imaging studies table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS imaging_studies (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                pet_id      INTEGER NOT NULL,
                owner_id    INTEGER,
                visit_id    INTEGER,
                study_type  TEXT NOT NULL,
                body_region TEXT,
                file_path   TEXT,
                notes       TEXT,
                ai_analysis TEXT,
                created_by  TEXT,
                created_at  TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (pet_id) REFERENCES pets(id) ON DELETE CASCADE
            )
        """)
        # Seed default budget targets (idempotent — only if table is empty).
        # Table may not exist yet in this transaction; _try_stmt keeps the probe
        # from aborting the surrounding transaction on PostgreSQL.
        _bt = _try_stmt(conn, "SELECT COUNT(*) FROM budget_targets")
        if _bt is not None and _bt.fetchone()[0] == 0:
            for _cat, _amt in [
                ("Medicines/Supplies", 50000),
                ("Staff Salaries",     120000),
                ("Utilities",          15000),
                ("Equipment",          25000),
                ("Marketing",          10000),
                ("Miscellaneous",      8000),
            ]:
                conn.execute(
                    "INSERT INTO budget_targets (category, monthly_egp) VALUES (?,?)",
                    (_cat, _amt)
                )
        # clinic
        if conn.execute("SELECT COUNT(*) FROM clinic").fetchone()[0] == 0:
            conn.execute(
                # Seed BLANK, not the vendor's own brand. A fresh clinic that
                # has not opened Settings yet would otherwise print "Aleefy" on
                # the invoices it hands its own customers. Blank lets the
                # neutral fallbacks in pdf_generator and base.html take over.
                # Only the INSERT is changed — the column DEFAULTs are left
                # alone so the Alembic baseline does not drift.
                "INSERT INTO clinic (name, name_ar, doctor_name) VALUES (?,?,?)",
                ("", "", ""),
            )
        # branches
        if conn.execute("SELECT COUNT(*) FROM branches").fetchone()[0] == 0:
            conn.execute("INSERT INTO branches (name, name_ar) VALUES (?,?)",
                         ("Main Branch","الفرع الرئيسي"))
        # roles
        for (rn, rd, rda, rc) in _SEED_ROLES:
            conn.execute(
                "INSERT OR IGNORE INTO roles (name,display_name,display_name_ar,color) VALUES (?,?,?,?)",
                (rn, rd, rda, rc))
        # Roles shipped with no grants at all, which the access check reads as
        # "fall open" — so every role reached every module. Fill only the empty
        # ones, so an administrator's own choices are never overwritten.
        _seeded = seed_default_permissions(conn)
        if _seeded:
            logger.info("seeded default permissions for %d role(s)", _seeded)
        # admin user
        if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
            conn.execute(
                "INSERT INTO users (username,password_hash,full_name,role,is_active) VALUES (?,?,?,?,1)",
                (admin_user, _hash(admin_pass), "Platform Administrator", "super_admin"))
        # item categories
        for (cn, cna) in _SEED_CATEGORIES:
            conn.execute("INSERT OR IGNORE INTO item_categories (name,name_ar) VALUES (?,?)", (cn, cna))
        # default warehouse
        if conn.execute("SELECT COUNT(*) FROM warehouses").fetchone()[0] == 0:
            conn.execute("INSERT INTO warehouses (name,name_ar) VALUES (?,?)",
                         ("Main Pharmacy","الصيدلية الرئيسية"))
        # whatsapp templates
        for (tn, sc, lg, txt) in _SEED_WA_TEMPLATES:
            conn.execute(
                "INSERT OR IGNORE INTO whatsapp_templates (name,scenario,language,template_text) VALUES (?,?,?,?)",
                (tn, sc, lg, txt))
        # shifts
        if conn.execute("SELECT COUNT(*) FROM shifts").fetchone()[0] == 0:
            for (sn, st, et, bk, days) in [
                # Sun=0 … Sat=6, and the Egyptian week: Sunday to Thursday, with
                # Friday and Saturday off. These used to seed a Monday-to-Friday
                # week with a "Weekend Morning" shift on Sat+Sun, which is the
                # American week — so every clinic started life expecting staff in
                # on Friday and treating Sunday as a day off.
                ("Morning Shift",   "08:00", "16:00", 60, "0,1,2,3,4"),
                ("Evening Shift",   "14:00", "22:00", 60, "0,1,2,3,4"),
                ("Night Shift",     "22:00", "06:00", 60, "0,1,2,3,4,5,6"),
                ("Weekend Morning", "09:00", "15:00", 30, "5,6"),
            ]:
                conn.execute(
                    "INSERT INTO shifts(name,start_time,end_time,break_minutes,days_of_week) VALUES(?,?,?,?,?)",
                    (sn, st, et, bk, days))
        # leave types
        if conn.execute("SELECT COUNT(*) FROM leave_types").fetchone()[0] == 0:
            for (ln, la, days, paid) in [
                ("Annual Leave",    "إجازة سنوية",    21, 1),
                ("Sick Leave",      "إجازة مرضية",    14, 1),
                ("Emergency Leave", "إجازة طارئة",     3, 1),
                ("Maternity Leave", "إجازة أمومة",    90, 1),
                ("Unpaid Leave",    "إجازة بدون راتب",30, 0),
                ("Study Leave",     "إجازة دراسية",    5, 1),
            ]:
                conn.execute(
                    "INSERT INTO leave_types(name,name_ar,days_per_year,is_paid) VALUES(?,?,?,?)",
                    (ln, la, days, paid))
        # grooming services
        if conn.execute("SELECT COUNT(*) FROM grooming_services").fetchone()[0] == 0:
            for (n, p, d) in [("Basic Bath","200",60),("Full Grooming","350",90),("Nail Trim","80",20),("Ear Cleaning","100",15)]:
                conn.execute("INSERT INTO grooming_services (name,price,duration_min) VALUES (?,?,?)",(n,p,d))
        # boarding rooms
        if conn.execute("SELECT COUNT(*) FROM boarding_rooms").fetchone()[0] == 0:
            for (n, rt, p) in [("Room A1","Standard",150),("Room A2","Standard",150),("Suite B1","Premium",300),("ICU 1","ICU",500)]:
                conn.execute("INSERT INTO boarding_rooms (name,room_type,price_per_night) VALUES (?,?,?)",(n,rt,p))
        # service catalog
        if conn.execute("SELECT COUNT(*) FROM service_catalog").fetchone()[0] == 0:
            _seed_services(conn)
    conn.close()

    # Retail tables live in the petshop blueprint and are created lazily by its
    # routes, so a freshly provisioned database had no ps_orders until somebody
    # opened a petshop page. Every petshop route calls ensure_petshop_tables()
    # first, so the module itself was fine — but anything CROSS-module reached
    # them without that guarantee, and a report joining retail sales to the
    # accounts would fail on a new deployment until a specific page was visited.
    # A fresh install should have a complete schema.
    #
    # Imported here rather than at module scope: blueprints import this module,
    # so a top-level import is circular. Guarded because a blueprint problem
    # must never stop the core schema being built.
    try:
        from blueprints.petshop.routes import ensure_petshop_tables
        ensure_petshop_tables()
    except Exception:
        logger.warning("could not create the retail tables during init_db",
                       exc_info=True)


def _seed_services(conn) -> None:
    """Seed default service price catalog."""
    services = [
        # code, name, name_ar, category, price, duration_min
        ("CONS-GEN", "General Consultation",   "استشارة عامة",         "Consultation", 150, 20),
        ("CONS-EMG", "Emergency Consultation", "استشارة طارئة",        "Consultation", 300, 30),
        ("CONS-FOL", "Follow-up Consultation", "زيارة متابعة",         "Consultation", 80,  15),
        ("VAC-RAB",  "Rabies Vaccine",         "تطعيم الكلب الأسود",  "Vaccination",  120, 10),
        ("VAC-DHPP", "DHPP Combo Vaccine",     "تطعيم رباعي",          "Vaccination",  150, 10),
        ("VAC-FVR",  "Feline FVRCP Vaccine",   "تطعيم القطط الرباعي", "Vaccination",  130, 10),
        ("LAB-CBC",  "CBC Blood Count",        "صورة دم كاملة",        "Laboratory",   200, 30),
        ("LAB-BIO",  "Biochemistry Panel",     "تحاليل كيميائية",      "Laboratory",   350, 45),
        ("LAB-URI",  "Urinalysis",             "تحليل بول",            "Laboratory",   150, 20),
        ("LAB-XRY",  "X-Ray (1 view)",         "أشعة سينية",           "Laboratory",   300, 20),
        ("LAB-ULT",  "Ultrasound",             "سونار",                "Laboratory",   400, 30),
        ("SRG-SPN",  "Spay/Neuter",           "تعقيم",                "Surgery",      800, 90),
        ("SRG-DEN",  "Dental Cleaning",        "تنظيف الأسنان",        "Surgery",      500, 60),
        ("SRG-MAS",  "Mass Removal",           "استئصال ورم",          "Surgery",      1200,120),
        ("GRM-BTH",  "Basic Bath",             "استحمام بسيط",         "Grooming",     200, 60),
        ("GRM-FUL",  "Full Grooming",          "تجميل كامل",           "Grooming",     350, 90),
        ("GRM-NAL",  "Nail Trim",              "قص أظافر",             "Grooming",     80,  20),
        ("BRD-STD",  "Boarding (Standard)",    "إيواء عادي",           "Boarding",     150, 0),
        ("BRD-PRM",  "Boarding (Premium Suite)","إيواء مميز",          "Boarding",     300, 0),
        ("HOSP-DAY", "Day Hospitalization",    "إقامة نهارية",         "Hospitalization",200,480),
        ("MED-ADM",  "IV Fluid Administration","تعطية سوائل",          "Treatment",    150, 30),
        ("MED-INJ",  "Injection",              "حقنة",                 "Treatment",    50,  5),
        ("MED-WND",  "Wound Dressing",         "تضميد جرح",            "Treatment",    100, 20),
    ]
    for (code, name, name_ar, cat, price, dur) in services:
        conn.execute(
            "INSERT OR IGNORE INTO service_catalog(code,name,name_ar,category,standard_price,duration_min) VALUES(?,?,?,?,?,?)",
            (code, name, name_ar, cat, price, dur))


# ── AUTH ───────────────────────────────────────────────────────
_SALT = "pah_platform_2026"
_BCRYPT_PREFIX = b"$2b$"


def _hash_sha256(pw: str) -> str:
    """Legacy SHA-256 hash (kept for migration detection only)."""
    return hashlib.sha256(f"{_SALT}{pw}".encode()).hexdigest()


def _hash(pw: str) -> str:
    """New primary hash: bcrypt. Used for all new passwords."""
    return _bcrypt.hashpw(pw.encode(), _bcrypt.gensalt(rounds=12)).decode()


def _hash_password(pw: str) -> str:
    """Alias for _hash — public API used by HR/reset routes."""
    return _hash(pw)


# A real bcrypt hash of a value nobody will ever type, computed once at import.
# Its only job is to cost the same as a genuine check, so an unknown username
# and a wrong password take the same time to refuse.
_DUMMY_HASH = _bcrypt.hashpw(
    b"aleefy-constant-time-placeholder", _bcrypt.gensalt(rounds=12))


def _verify_dummy(password: str) -> None:
    """Spend a bcrypt verification on nothing, so timing reveals nothing."""
    try:
        _bcrypt.checkpw((password or "").encode(), _DUMMY_HASH)
    except Exception:
        pass


def _verify_and_migrate(row, password: str, conn) -> bool:
    """
    Verify password against bcrypt (preferred) or SHA-256 (legacy).
    On SHA-256 match, transparently rehash with bcrypt and save.
    Returns True if password matches.
    """
    stored = row["password_hash"]
    # Try bcrypt first (new hashes start with $2b$)
    if stored and stored.startswith("$2b$"):
        try:
            return _bcrypt.checkpw(password.encode(), stored.encode())
        except Exception:
            return False
    # Legacy SHA-256 check
    if stored == _hash_sha256(password):
        # Rehash with bcrypt transparently
        new_hash = _hash(password)
        try:
            conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                         (new_hash, row["id"]))
            conn.commit()
        except Exception:
            # The sign-in itself has already succeeded and must not be undone
            # by a failed upgrade - but a rehash that silently never lands
            # means legacy SHA-256 hashes live for ever with nobody informed.
            logger.exception("could not upgrade a legacy password hash")
        return True
    return False


def verify_credentials(username: str, password: str) -> Optional[dict]:
    conn = get_db()
    row = conn.execute(
        "SELECT * FROM users WHERE username=? AND is_active=1", (username,)).fetchone()
    if not row:
        conn.close()
        # Burn the same time an existing user would. Returning immediately made
        # the response ~0.26s for an unknown username and ~0.59s for a real one
        # (bcrypt cost 12 runs only when there is a hash to check), which is a
        # reliable oracle: an anonymous stranger can script the login form and
        # read the clock to recover every staff username without one correct
        # password. That list then feeds a targeted lockout.
        _verify_dummy(password)
        return None
    ok = _verify_and_migrate(row, password, conn)
    conn.close()
    return dict(row) if ok else None

def touch_last_login(user_id: int) -> None:
    from datetime import datetime
    now = datetime.utcnow().isoformat(timespec='seconds')
    conn = get_db()
    with conn:
        conn.execute("UPDATE users SET last_login_at=? WHERE id=?", (now, user_id))
    conn.close()

def get_user(username: str) -> Optional[dict]:
    conn = get_db()
    row = conn.execute("SELECT * FROM users WHERE username=? AND is_active=1",(username,)).fetchone()
    conn.close()
    return dict(row) if row else None

def get_user_by_id(user_id: int) -> Optional[dict]:
    """A user by id, ACTIVE OR NOT — the caller decides what to do about it.

    get_user() filters on is_active=1, which is right for a login. It is wrong
    for the shared-desk switcher, which has to tell "no such user" apart from
    "this account was deactivated while it sat signed in on a reception PC" —
    the second must drop the account off that machine and say so, not fail
    silently as though it never existed.
    """
    conn = get_db()
    row = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def update_user_theme(username: str, theme: str) -> None:
    conn = get_db()
    with conn:
        conn.execute("UPDATE users SET theme_preference=? WHERE username=?", (theme, username))
    conn.close()

# ── CLINIC ─────────────────────────────────────────────────────
def get_clinic() -> dict:
    """Return clinic row — cached 5 min so context_processor pays zero DB cost."""
    cached, hit = _cache_get("clinic_row")
    if hit:
        return cached
    conn = get_db()
    row = conn.execute("SELECT * FROM clinic LIMIT 1").fetchone()
    conn.close()
    result = dict(row) if row else {}
    _cache_set("clinic_row", result, ttl=300)
    return result

def update_clinic(data: dict, updated_by: str = "system") -> None:
    """Update clinic settings and invalidate the cache.

    Had zero callers and was broken for both engines when found: it built
    `%s` placeholders by hand and used `NOW()`, neither of which survives the
    SQLite path, while `_fix_sql` only translates in the other direction. The
    settings blueprint writes its own portable UPDATE instead. Repaired rather
    than deleted so the next person who reaches for the obvious function name
    gets working code instead of a landmine.
    """
    if not data:
        return
    # ? is the portable placeholder — _fix_sql rewrites it to %s for PostgreSQL.
    # Column names are interpolated, so they must never come from user input;
    # restrict them to the real columns of the table.
    allowed = {
        "name", "name_ar", "phone", "email", "address", "address_ar", "website",
        "tax_number", "license_number", "doctor_name", "tagline", "logo_data",
        "currency", "timezone",
    }
    fields = {k: v for k, v in data.items() if k in allowed}
    if not fields:
        logger.warning("update_clinic called with no recognised columns: %s",
                       sorted(data))
        return
    sets = ", ".join(f"{k}=?" for k in fields)
    vals = list(fields.values())
    conn = get_db()
    try:
        with conn:
            conn.execute(
                f"UPDATE clinic SET {sets}, updated_at=datetime('now') WHERE id=1",
                vals,
            )
    finally:
        conn.close()
    cache_invalidate("clinic_row")

# ── SETTINGS ───────────────────────────────────────────────────
def get_setting(key: str, default: str = "") -> str:
    cache_key = f"setting:{key}"
    cached, hit = _cache_get(cache_key)
    if hit:
        return cached
    conn = get_db()
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    result = (row[0] or default) if row else default
    _cache_set(cache_key, result, ttl=300)
    return result

def set_setting(key: str, value: str, category: str = "general", updated_by: str = "system") -> None:
    conn = get_db()
    with conn:
        conn.execute(
            # Explicit upsert. _fix_sql renders "INSERT OR REPLACE" as
            # "ON CONFLICT DO NOTHING", so on PostgreSQL every setting that
            # already had a value silently kept its OLD one -- set_setting()
            # returned normally and changed nothing. This spelling works
            # unchanged on both engines.
            "INSERT INTO settings(key,value,category,updated_at,updated_by) "
            "VALUES(?,?,?,datetime('now'),?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value, "
            "category=excluded.category, updated_at=excluded.updated_at, "
            "updated_by=excluded.updated_by",
            (key, value, category, updated_by))
    conn.close()
    cache_invalidate(f"setting:{key}")

# ── AUDIT ──────────────────────────────────────────────────────
def log_audit(username="", role="", action="", module="",
              entity_type="", entity_id="", details="", ip="", user_agent=""):
    """Record one audit row. Never raises — an audit failure must not roll back
    the operation being audited — but never fails silently either: an audit log
    everyone believes in that quietly stopped recording is worse than none.

    The INSERT runs on its own connection, so a failure here cannot abort the
    caller's PostgreSQL transaction.
    """
    try:
        conn = get_db()
        try:
            with conn:
                conn.execute(
                    "INSERT INTO audit_log(username,role,action,module,entity_type,entity_id,details,ip,user_agent) VALUES(?,?,?,?,?,?,?,?,?)",
                    (username,role,action,module,entity_type,entity_id,details,ip,user_agent))
        finally:
            # `with conn:` commits/rolls back but does not close on sqlite3.
            conn.close()
    except Exception:
        logger.error("audit log write FAILED — action=%r module=%r user=%r "
                     "entity=%r/%r was not recorded",
                     action, module, username, entity_type, entity_id,
                     exc_info=True)

def get_audit_log(limit: int = 200) -> list:
    conn = get_db()
    rows = conn.execute("SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ── HR ─────────────────────────────────────────────────────────
def list_users() -> list:
    conn = get_db()
    rows = conn.execute("SELECT * FROM users ORDER BY full_name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def create_user(data: dict) -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO users(username,password_hash,full_name,email,phone,role,is_active) VALUES(?,?,?,?,?,?,?)",
            (data["username"], _hash(data.get("password", "changeme")),
             data.get("full_name", ""), data.get("email", ""), data.get("phone", ""),
             data.get("role", "staff"), 1))
        uid = cur.lastrowid
    conn.close()
    return uid

def _create_user_safe(data: dict) -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO users(username,password_hash,full_name,email,phone,role,is_active) VALUES(?,?,?,?,?,?,?)",
            (data["username"], _hash(data.get("password","changeme")),
             data.get("full_name",""), data.get("email",""), data.get("phone",""),
             data.get("role","staff"), 1))
        uid = cur.lastrowid
    conn.close()
    return uid

def toggle_user_active(user_id: int, active: int) -> None:
    conn = get_db()
    with conn:
        conn.execute("UPDATE users SET is_active=? WHERE id=?", (active, user_id))
    conn.close()

def update_user_role(user_id: int, role: str) -> None:
    conn = get_db()
    with conn:
        conn.execute("UPDATE users SET role=? WHERE id=?", (role, user_id))
    conn.close()

# ── CRM — OWNERS ───────────────────────────────────────────────
def list_owners(search: str = "", limit: int = 100, offset: int = 0) -> list:
    """Return owners with pet_count in a single aggregated query (no N+1)."""
    conn = get_db()
    base = (
        "SELECT o.*, COALESCE(pc.cnt, 0) AS pet_count"
        " FROM owners o"
        " LEFT JOIN (SELECT owner_id, COUNT(*) AS cnt FROM pets GROUP BY owner_id) pc"
        "   ON pc.owner_id = o.id"
    )
    if search:
        q = f"%{search}%"
        rows = conn.execute(
            base + " WHERE o.full_name LIKE ? OR o.phone LIKE ?"
                   " OR o.whatsapp_phone LIKE ? OR o.email LIKE ?"
                   " ORDER BY o.full_name LIMIT ? OFFSET ?",
            (q, q, q, q, limit, offset)).fetchall()
    else:
        rows = conn.execute(
            base + " ORDER BY o.created_at DESC LIMIT ? OFFSET ?",
            (limit, offset)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def count_owners(search: str = "") -> int:
    conn = get_db()
    if search:
        q = f"%{search}%"
        n = conn.execute(
            "SELECT COUNT(*) FROM owners WHERE full_name LIKE ? OR phone LIKE ? OR email LIKE ?",
            (q,q,q)).fetchone()[0]
    else:
        n = conn.execute("SELECT COUNT(*) FROM owners").fetchone()[0]
    conn.close()
    return n

def get_owner(owner_id: int) -> Optional[dict]:
    conn = get_db()
    row = conn.execute("SELECT * FROM owners WHERE id=?", (owner_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

class DuplicatePhone(ValueError):
    """Raised when a mobile number already belongs to another client.

    Carries the existing client so the caller can offer to OPEN them rather
    than just refusing — the person at the desk almost always meant that one.
    """

    def __init__(self, message, owner_id=None, owner_name=""):
        super().__init__(message)
        self.owner_id = owner_id
        self.owner_name = owner_name


def normalise_phone(phone: str) -> str:
    """A comparable form of an Egyptian mobile.

    The same number is typed a dozen ways at a front desk — "0100 123 4567",
    "+201001234567", "0100-123-4567", "٠١٠٠١٢٣٤٥٦٧" — and every variant used to
    create a NEW client, splitting one family's pets across several records.
    Only digits are kept, Arabic-Indic digits are folded to ASCII, and the
    Egyptian country code is dropped so 0100… and +20100… compare equal.
    """
    if not phone:
        return ""
    s = str(phone).translate(str.maketrans(
        "٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789"))
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits.startswith("0020"):
        digits = digits[4:]
    elif digits.startswith("20") and len(digits) > 10:
        digits = digits[2:]
    return digits.lstrip("0")


def owner_by_phone(phone: str, exclude_id=None):
    """The client already holding this number, comparing normalised forms."""
    key = normalise_phone(phone)
    if not key:
        return None
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT id, full_name, phone, whatsapp_phone FROM owners").fetchall()
    finally:
        conn.close()
    for r in rows:
        if exclude_id and r["id"] == exclude_id:
            continue
        if key in (normalise_phone(r["phone"]), normalise_phone(r["whatsapp_phone"])):
            return dict(r)
    return None


def assert_phone_is_free(phone: str, exclude_id=None) -> None:
    """Refuse a mobile that already belongs to somebody else.

    Enforced HERE rather than with a UNIQUE index, because a real database
    already contains duplicates — this one has 15 groups — and a constraint
    added over them fails at startup and takes the clinic down. The rule
    applies to every new write; the existing pairs are reported by
    scripts/find_duplicate_owners.py so they can be merged deliberately.
    """
    if not normalise_phone(phone):
        return
    existing = owner_by_phone(phone, exclude_id=exclude_id)
    if existing:
        raise DuplicatePhone(
            "%s already uses this mobile number." % (existing["full_name"] or "Another client"),
            owner_id=existing["id"], owner_name=existing["full_name"] or "")


def create_owner(data: dict) -> int:
    assert_phone_is_free(data.get("phone") or data.get("whatsapp_phone"))
    conn = get_db()
    with conn:
        cur = conn.execute(
            """INSERT INTO owners(full_name,phone,whatsapp_phone,email,address,
               preferred_contact,preferred_doctor,vip_flag,notes,marketing_consent,created_by)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("full_name",""), data.get("phone",""), data.get("whatsapp_phone",""),
             data.get("email",""), data.get("address",""), data.get("preferred_contact","WhatsApp"),
             data.get("preferred_doctor",""), int(data.get("vip_flag",0)),
             data.get("notes",""), int(data.get("marketing_consent",1)), data.get("created_by","")))
        oid = cur.lastrowid
    conn.close()
    return oid

def update_owner(owner_id: int, data: dict) -> None:
    assert_phone_is_free(data.get("phone") or data.get("whatsapp_phone"),
                         exclude_id=owner_id)
    conn = get_db()
    with conn:
        conn.execute(
            """UPDATE owners SET full_name=?,phone=?,whatsapp_phone=?,email=?,address=?,
               preferred_contact=?,preferred_doctor=?,vip_flag=?,notes=?,marketing_consent=?,
               updated_at=datetime('now') WHERE id=?""",
            (data.get("full_name",""), data.get("phone",""), data.get("whatsapp_phone",""),
             data.get("email",""), data.get("address",""), data.get("preferred_contact","WhatsApp"),
             data.get("preferred_doctor",""), int(data.get("vip_flag",0)),
             data.get("notes",""), int(data.get("marketing_consent",1)), owner_id))
    conn.close()

def delete_owner(owner_id: int) -> None:
    conn = get_db()
    with conn:
        conn.execute("DELETE FROM owners WHERE id=?", (owner_id,))
    conn.close()

def get_owner_balance(owner_id: int) -> float:
    conn = get_db()
    row = conn.execute(
        "SELECT COALESCE(SUM(due_amount),0) FROM invoices WHERE owner_id=? AND status!='Cancelled'",
        (owner_id,)).fetchone()
    conn.close()
    return float(row[0]) if row else 0.0

# ── CRM — PETS ─────────────────────────────────────────────────
def list_pets(owner_id: Optional[int] = None, search: str = "") -> list:
    conn = get_db()
    if owner_id:
        rows = conn.execute(
            "SELECT p.*, o.full_name owner_name FROM pets p JOIN owners o ON o.id=p.owner_id"
            " WHERE p.owner_id=? ORDER BY p.pet_name", (owner_id,)).fetchall()
    elif search:
        q = f"%{search}%"
        rows = conn.execute(
            "SELECT p.*, o.full_name owner_name FROM pets p JOIN owners o ON o.id=p.owner_id"
            " WHERE p.pet_name LIKE ? OR p.microchip_id LIKE ?"
            " ORDER BY p.pet_name LIMIT 100", (q,q)).fetchall()
    else:
        rows = conn.execute(
            "SELECT p.*, o.full_name owner_name FROM pets p JOIN owners o ON o.id=p.owner_id"
            " ORDER BY p.created_at DESC LIMIT 100").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_pet(pet_id: int) -> Optional[dict]:
    conn = get_db()
    row = conn.execute(
        "SELECT p.*, o.full_name owner_name, o.phone owner_phone, o.whatsapp_phone"
        " FROM pets p JOIN owners o ON o.id=p.owner_id WHERE p.id=?", (pet_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def create_pet(data: dict) -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            """INSERT INTO pets(owner_id,pet_name,species,breed,sex,dob,weight_kg,
               color,microchip_id,neutered,allergies,chronic_conditions,notes)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data["owner_id"], data.get("pet_name",""), data.get("species",""),
             data.get("breed",""), data.get("sex","Unknown"), data.get("dob",""),
             data.get("weight_kg") or None, data.get("color",""), data.get("microchip_id",""),
             int(data.get("neutered",0)), data.get("allergies",""),
             data.get("chronic_conditions",""), data.get("notes","")))
        pid = cur.lastrowid
    conn.close()
    return pid

def update_pet(pet_id: int, data: dict) -> None:
    conn = get_db()
    with conn:
        conn.execute(
            """UPDATE pets SET pet_name=?,species=?,breed=?,sex=?,dob=?,weight_kg=?,
               color=?,microchip_id=?,neutered=?,allergies=?,chronic_conditions=?,
               notes=?,updated_at=datetime('now') WHERE id=?""",
            (data.get("pet_name",""), data.get("species",""), data.get("breed",""),
             data.get("sex","Unknown"), data.get("dob",""), data.get("weight_kg") or None,
             data.get("color",""), data.get("microchip_id",""),
             int(data.get("neutered",0)), data.get("allergies",""),
             data.get("chronic_conditions",""), data.get("notes",""), pet_id))
    conn.close()

def get_pet_timeline(pet_id: int) -> list:
    """Return all clinical events for a pet, sorted newest first."""
    conn = get_db()
    events = []
    # Visits
    for r in conn.execute("SELECT id, visit_date dt, visit_type etype, chief_complaint summary, status FROM visits WHERE pet_id=? ORDER BY visit_date DESC", (pet_id,)).fetchall():
        events.append({"dt": r["dt"], "type": "visit", "icon": "🩺", "title": f"Visit — {r['etype']}", "summary": r["summary"] or "", "id": r["id"], "status": r["status"]})
    # Vaccinations
    for r in conn.execute("SELECT id, administered_at dt, vaccine_name vname FROM vaccinations WHERE pet_id=? ORDER BY administered_at DESC", (pet_id,)).fetchall():
        events.append({"dt": r["dt"], "type": "vaccine", "icon": "💉", "title": f"Vaccine — {r['vname']}", "summary": "", "id": r["id"]})
    # Surgeries
    for r in conn.execute("SELECT id, surgery_date dt, procedure_name pname, outcome FROM surgeries WHERE pet_id=? ORDER BY surgery_date DESC", (pet_id,)).fetchall():
        events.append({"dt": r["dt"], "type": "surgery", "icon": "🔧", "title": f"Surgery — {r['pname']}", "summary": r["outcome"] or "", "id": r["id"]})
    # Grooming
    for r in conn.execute("SELECT id, booking_date dt, status FROM grooming_bookings WHERE pet_id=? ORDER BY booking_date DESC", (pet_id,)).fetchall():
        events.append({"dt": r["dt"], "type": "grooming", "icon": "✂️", "title": "Grooming", "summary": r["status"], "id": r["id"]})
    # Invoices (linked to pet via invoice table)
    for r in conn.execute(
        "SELECT id, issue_date dt, invoice_number inv_no, total, status FROM invoices WHERE pet_id=? ORDER BY issue_date DESC",
        (pet_id,)
    ).fetchall():
        events.append({
            "dt": r["dt"], "type": "invoice", "icon": "🧾",
            "title": f"Invoice {r['inv_no']} — {r['status']}",
            "summary": f"Total: {r['total']:.2f}" if r["total"] else "",
            "id": r["id"],
        })
    # Lab requests (linked through visits)
    for r in conn.execute(
        """SELECT lr.id, lr.created_at dt, lr.test_name, lr.status
           FROM lab_requests lr
           JOIN visits v ON v.id = lr.visit_id
           WHERE v.pet_id=? ORDER BY lr.created_at DESC""",
        (pet_id,)
    ).fetchall():
        events.append({
            "dt": r["dt"], "type": "lab", "icon": "🔬",
            "title": f"Lab — {r['test_name']}",
            "summary": r["status"] or "",
            "id": r["id"],
        })
    conn.close()
    events.sort(key=lambda x: x["dt"] or "", reverse=True)
    return events

# ── APPOINTMENTS ───────────────────────────────────────────────
def list_appointments(date_from: str = "", date_to: str = "",
                      status: str = "", doctor: str = "",
                      limit: int = 100) -> list:
    conn = get_db()
    q = """SELECT a.*, o.full_name owner_name, o.phone owner_phone,
                  p.pet_name, p.species
           FROM appointments a
           JOIN owners o ON o.id=a.owner_id
           JOIN pets   p ON p.id=a.pet_id
           WHERE 1=1"""
    params: list = []
    if date_from: q += " AND a.appt_date >= ?"; params.append(date_from)
    if date_to:   q += " AND a.appt_date <= ?"; params.append(date_to)
    if status:    q += " AND a.status = ?";      params.append(status)
    if doctor:    q += " AND a.doctor_name LIKE ?"; params.append(f"%{doctor}%")
    q += " ORDER BY a.appt_date, a.appt_start LIMIT ?"; params.append(limit)
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_today_appointments() -> list:
    today = date.today().isoformat()
    return list_appointments(date_from=today, date_to=today, limit=200)

def get_appointment(appt_id: int) -> Optional[dict]:
    conn = get_db()
    row = conn.execute(
        "SELECT a.*, o.full_name owner_name, p.pet_name, p.species FROM appointments a"
        " JOIN owners o ON o.id=a.owner_id JOIN pets p ON p.id=a.pet_id WHERE a.id=?",
        (appt_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def create_appointment(data: dict) -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            """INSERT INTO appointments(owner_id,pet_id,doctor_name,room,appointment_type,
               priority,status,channel,appt_date,appt_start,appt_end,duration_min,
               reason,symptoms,notes,created_by)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data["owner_id"], data["pet_id"], data.get("doctor_name",""),
             data.get("room",""), data.get("appointment_type","Consultation"),
             data.get("priority","Normal"), data.get("status","Scheduled"),
             data.get("channel","Walk-in"), data["appt_date"], data.get("appt_start","09:00"),
             data.get("appt_end",""), data.get("duration_min",30),
             data.get("reason",""), data.get("symptoms",""), data.get("notes",""),
             data.get("created_by","")))
        aid = cur.lastrowid
    conn.close()
    return aid

def update_appointment_status(appt_id: int, status: str, username: str = "") -> None:
    conn = get_db()
    with conn:
        extra = ""
        if status == "Checked-in":
            extra = ", checked_in_at=datetime('now')"
        elif status in ("Completed","No-Show","Cancelled"):
            extra = ", checked_out_at=datetime('now')"
        conn.execute(f"UPDATE appointments SET status=?,updated_at=datetime('now'){extra} WHERE id=?",
                     (status, appt_id))
    conn.close()

def get_appointment_stats_today() -> dict:
    """Return today's appointment counts in a single aggregated query."""
    conn = get_db()
    today = date.today().isoformat()
    rows = conn.execute(
        "SELECT status, COUNT(*) n FROM appointments WHERE appt_date=? GROUP BY status",
        (today,)
    ).fetchall()
    conn.close()
    by_status = {r["status"]: r["n"] for r in rows}
    total = sum(by_status.values())
    return {"total": total, "by_status": by_status, "date": today}

# ── VISITS ─────────────────────────────────────────────────────
def list_visits(pet_id: Optional[int] = None, limit: int = 50) -> list:
    conn = get_db()
    if pet_id:
        rows = conn.execute(
            "SELECT v.*, p.pet_name, o.full_name owner_name FROM visits v"
            " JOIN pets p ON p.id=v.pet_id JOIN owners o ON o.id=v.owner_id"
            " WHERE v.pet_id=? ORDER BY v.visit_date DESC LIMIT ?", (pet_id, limit)).fetchall()
    else:
        rows = conn.execute(
            "SELECT v.*, p.pet_name, o.full_name owner_name FROM visits v"
            " JOIN pets p ON p.id=v.pet_id JOIN owners o ON o.id=v.owner_id"
            " ORDER BY v.visit_date DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_visit(visit_id: int) -> Optional[dict]:
    conn = get_db()
    row = conn.execute(
        "SELECT v.*, p.pet_name, p.species, p.breed, p.weight_kg pet_weight,"
        " o.full_name owner_name, o.phone owner_phone FROM visits v"
        " JOIN pets p ON p.id=v.pet_id JOIN owners o ON o.id=v.owner_id WHERE v.id=?",
        (visit_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def get_visit_diagnoses(visit_id: int) -> list:
    conn = get_db()
    rows = conn.execute("SELECT * FROM diagnoses WHERE visit_id=?", (visit_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_visit_prescriptions(visit_id: int) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT p.*, (SELECT json_group_array(json_object('name',pi.medication_name,'dosage',pi.dosage,'freq',pi.frequency,'qty',pi.quantity,'unit',pi.unit,'instructions',pi.instructions))"
        " FROM prescription_items pi WHERE pi.prescription_id=p.id) items_json"
        " FROM prescriptions p WHERE p.visit_id=?", (visit_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def create_visit(data: dict) -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            """INSERT INTO visits(appointment_id,owner_id,pet_id,doctor_name,room,
               visit_date,visit_type,status,chief_complaint,symptoms,weight_kg,
               temp_c,heart_rate,notes,created_by)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("appointment_id"), data["owner_id"], data["pet_id"],
             data.get("doctor_name",""), data.get("room",""),
             data.get("visit_date", date.today().isoformat()),
             data.get("visit_type","Consultation"), data.get("status","Open"),
             data.get("chief_complaint",""), data.get("symptoms",""),
             data.get("weight_kg") or None, data.get("temp_c") or None,
             data.get("heart_rate") or None, data.get("notes",""),
             data.get("created_by","")))
        vid = cur.lastrowid
    conn.close()
    return vid

def add_diagnosis(visit_id: int, pet_id: int, diagnosis: str,
                  severity: str = "Moderate", notes: str = "", created_by: str = "") -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO diagnoses(visit_id,pet_id,diagnosis,severity,notes,created_by) VALUES(?,?,?,?,?,?)",
            (visit_id, pet_id, diagnosis, severity, notes, created_by))
        did = cur.lastrowid
    conn.close()
    return did

def complete_visit(visit_id: int) -> None:
    conn = get_db()
    with conn:
        conn.execute("UPDATE visits SET status='Completed',updated_at=datetime('now') WHERE id=?",
                     (visit_id,))
    conn.close()

# ── INVENTORY ──────────────────────────────────────────────────
def list_items(search: str = "", category_id: Optional[int] = None,
               low_stock_only: bool = False, limit: int = 100) -> list:
    conn = get_db()
    q = """SELECT i.*, ic.name category_name,
                  COALESCE((SELECT SUM(b.quantity) FROM batches b WHERE b.item_id=i.id),0) stock_qty
           FROM items i LEFT JOIN item_categories ic ON ic.id=i.category_id
           WHERE i.is_active=1"""
    params: list = []
    if search:
        q += " AND (i.name LIKE ? OR i.sku LIKE ? OR i.barcode LIKE ?)"; s=f"%{search}%"; params+=[s,s,s]
    if category_id:
        q += " AND i.category_id=?"; params.append(category_id)
    q += " ORDER BY i.name LIMIT ?"; params.append(limit)
    rows = conn.execute(q, params).fetchall()
    result = [dict(r) for r in rows]
    if low_stock_only:
        result = [r for r in result if r["stock_qty"] <= r["reorder_level"]]
    conn.close()
    return result

def get_item(item_id: int) -> Optional[dict]:
    conn = get_db()
    row = conn.execute(
        "SELECT i.*, ic.name category_name,"
        " COALESCE((SELECT SUM(b.quantity) FROM batches b WHERE b.item_id=i.id),0) stock_qty"
        " FROM items i LEFT JOIN item_categories ic ON ic.id=i.category_id WHERE i.id=?",
        (item_id,)).fetchone()
    conn.close()
    return dict(row) if row else None

def create_item(data: dict) -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            """INSERT INTO items(category_id,sku,barcode,name,unit,cost_price,sell_price,
               reorder_level,is_medication,is_controlled,requires_rx,supplier_id,storage_notes)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (data.get("category_id"), data.get("sku",""), data.get("barcode",""),
             data["name"], data.get("unit","unit"),
             float(data.get("cost_price",0)), float(data.get("sell_price",0)),
             float(data.get("reorder_level",10)),
             int(data.get("is_medication",0)), int(data.get("is_controlled",0)),
             int(data.get("requires_rx",0)), data.get("supplier_id") or None,
             data.get("storage_notes","")))
        iid = cur.lastrowid
    conn.close()
    return iid

def add_stock_batch(item_id: int, warehouse_id: int, batch_number: str,
                    expiry_date: str, quantity: float, unit_cost: float,
                    received_by: str = "") -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO batches(item_id,warehouse_id,batch_number,expiry_date,quantity,unit_cost,received_by) VALUES(?,?,?,?,?,?,?)",
            (item_id, warehouse_id, batch_number, expiry_date, quantity, unit_cost, received_by))
        bid = cur.lastrowid
        conn.execute(
            "INSERT INTO stock_movements(item_id,batch_id,warehouse_id,movement_type,quantity,unit_cost,reference_type,created_by) VALUES(?,?,?,?,?,?,?,?)",
            (item_id, bid, warehouse_id, "in", quantity, unit_cost, "receiving", received_by))
    conn.close()
    return bid

def deduct_stock(item_id: int, quantity: float, reference_type: str = "dispensing",
                 reference_id: Optional[int] = None, by: str = "") -> bool:
    """Deduct stock using FEFO (First Expiry First Out). Returns True if sufficient stock."""
    conn = get_db()
    available = conn.execute(
        "SELECT COALESCE(SUM(quantity),0) FROM batches WHERE item_id=? AND quantity>0", (item_id,)).fetchone()[0]
    if float(available or 0) < quantity:
        conn.close()
        return False
    remaining = quantity
    batches = conn.execute(
        "SELECT * FROM batches WHERE item_id=? AND quantity>0 ORDER BY expiry_date ASC NULLS LAST",
        (item_id,)).fetchall()
    with conn:
        for b in batches:
            if remaining <= 0:
                break
            use = min(float(b["quantity"]), remaining)
            conn.execute("UPDATE batches SET quantity=quantity-? WHERE id=?", (use, b["id"]))
            conn.execute(
                "INSERT INTO stock_movements(item_id,batch_id,warehouse_id,movement_type,quantity,reference_type,reference_id,created_by) VALUES(?,?,?,?,?,?,?,?)",
                (item_id, b["id"], b["warehouse_id"], "out", use, reference_type, reference_id, by))
            remaining -= use
    conn.close()
    return True

def get_low_stock_items() -> list:
    return list_items(low_stock_only=True, limit=500)

def get_expiry_alerts(days: int = 30) -> list:
    conn = get_db()
    threshold = (date.today() + timedelta(days=days)).isoformat()
    rows = conn.execute(
        "SELECT b.*, i.name item_name, i.unit FROM batches b JOIN items i ON i.id=b.item_id"
        " WHERE b.expiry_date <= ? AND b.quantity > 0 ORDER BY b.expiry_date", (threshold,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def list_stock_movements(item_id: Optional[int] = None, limit: int = 100) -> list:
    conn = get_db()
    if item_id:
        rows = conn.execute(
            "SELECT sm.*, i.name item_name FROM stock_movements sm JOIN items i ON i.id=sm.item_id"
            " WHERE sm.item_id=? ORDER BY sm.created_at DESC LIMIT ?", (item_id, limit)).fetchall()
    else:
        rows = conn.execute(
            "SELECT sm.*, i.name item_name FROM stock_movements sm JOIN items i ON i.id=sm.item_id"
            " ORDER BY sm.created_at DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def list_categories() -> list:
    conn = get_db()
    rows = conn.execute("SELECT * FROM item_categories ORDER BY name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def list_warehouses() -> list:
    conn = get_db()
    rows = conn.execute("SELECT * FROM warehouses WHERE is_active=1").fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ── FINANCE ────────────────────────────────────────────────────
def _next_invoice_number() -> str:
    conn = get_db()
    n = conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0]
    conn.close()
    return f"INV-{date.today().year}-{(n+1):05d}"

def create_invoice(data: dict, lines: list) -> int:
    inv_no = _next_invoice_number()
    # Round at EVERY money step, not just the last few. subtotal and disc_amt
    # were the two left unrounded here, so a stored invoice could have a
    # subtotal that is not representable to 2dp while its total was — meaning
    # the header did not equal the sum of its own lines, and the discrepancy
    # was invisible on screen because the template formats to 2dp anyway.
    # ponytail: real fix is NUMERIC(12,2) end to end — see docs/MONEY_PRECISION.md
    subtotal = round(sum(round(float(l.get("total", 0)), 2) for l in lines), 2)
    disc_type = data.get("discount_type","value")
    disc_val  = float(data.get("discount_value",0))
    disc_amt  = round(disc_val, 2) if disc_type == "value" else round(subtotal * disc_val / 100, 2)
    # A discount may not exceed what is being discounted, and may not be
    # negative. Unclamped, "discount 5000" on a 560 invoice stored total
    # -4440.00 and it landed in Outstanding as money the CLINIC owed. Clamped
    # here rather than in each caller so every screen that bills is covered.
    disc_amt  = max(0.0, min(disc_amt, subtotal))
    tax_rate  = float(data.get("tax_rate",0))
    tax_amt   = round((subtotal - disc_amt) * tax_rate / 100, 2)
    total     = round(subtotal - disc_amt + tax_amt, 2)
    conn = get_db()
    with conn:
        cur = conn.execute(
            """INSERT INTO invoices(invoice_number,owner_id,pet_id,visit_id,doctor_name,issue_date,
               status,subtotal,discount_type,discount_value,discount_amount,tax_rate,tax_amount,
               total,paid_amount,due_amount,notes,created_by)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (inv_no, data["owner_id"], data.get("pet_id"), data.get("visit_id"),
             data.get("doctor_name",""), data.get("issue_date", date.today().isoformat()),
             "Unpaid", subtotal, disc_type, disc_val, disc_amt, tax_rate, tax_amt,
             total, 0.0, total, data.get("notes",""), data.get("created_by","")))
        inv_id = cur.lastrowid
        for line in lines:
            lt = float(line.get("total", float(line.get("quantity",1)) * float(line.get("unit_price",0))))
            conn.execute(
                "INSERT INTO invoice_lines(invoice_id,line_type,item_id,description,quantity,unit_price,discount,total) VALUES(?,?,?,?,?,?,?,?)",
                (inv_id, line.get("line_type","service"), line.get("item_id"),
                 line.get("description",""), float(line.get("quantity",1)),
                 float(line.get("unit_price",0)), float(line.get("discount",0)), lt))
    conn.close()
    return inv_id

def get_invoice(inv_id: int) -> Optional[dict]:
    conn = get_db()
    row = conn.execute(
        "SELECT i.*, o.full_name owner_name, o.phone owner_phone, o.whatsapp_phone,"
        " p.pet_name FROM invoices i JOIN owners o ON o.id=i.owner_id"
        " LEFT JOIN pets p ON p.id=i.pet_id WHERE i.id=?", (inv_id,)).fetchone()
    if not row:
        conn.close()
        return None
    inv = dict(row)
    inv["lines"] = [dict(r) for r in conn.execute(
        "SELECT * FROM invoice_lines WHERE invoice_id=?", (inv_id,)).fetchall()]
    inv["payments"] = []  # vet payments table not in shared DB; paid_amount tracked on invoice row
    conn.close()
    return inv

def list_invoices(owner_id: Optional[int] = None, status: str = "",
                  date_from: str = "", date_to: str = "", limit: int = 100) -> list:
    conn = get_db()
    q = "SELECT i.*, o.full_name owner_name, p.pet_name FROM invoices i JOIN owners o ON o.id=i.owner_id LEFT JOIN pets p ON p.id=i.pet_id WHERE 1=1"
    params: list = []
    if owner_id:  q += " AND i.owner_id=?";    params.append(owner_id)
    if status:    q += " AND i.status=?";       params.append(status)
    if date_from: q += " AND i.issue_date>=?";  params.append(date_from)
    if date_to:   q += " AND i.issue_date<=?";  params.append(date_to)
    q += " ORDER BY i.created_at DESC LIMIT ?"; params.append(limit)
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ─── Estimates (quotes) ───────────────────────────────────────────────────────

def _next_estimate_number() -> str:
    """Deliberately MAX(id)+1, not COUNT(*).

    _next_invoice_number() uses COUNT(*), which repeats a number as soon as any
    row is deleted -- and invoice_number is UNIQUE, so the next insert raises.
    Not fixing that here (it would renumber a live ledger), but not copying it
    either.
    """
    conn = get_db()
    n = conn.execute("SELECT COALESCE(MAX(id),0) FROM estimates").fetchone()[0]
    conn.close()
    return f"EST-{date.today().year}-{(n+1):05d}"


def _money(lines: list, data: dict) -> tuple:
    """(subtotal, disc_amt, tax_amt, total) -- same arithmetic as create_invoice.

    Shared so an approved estimate cannot total differently from the invoice it
    becomes. A quote the client signed and a bill that says something else is
    the one bug this feature absolutely must not have.
    """
    subtotal  = round(sum(round(float(l.get("total", 0)), 2) for l in lines), 2)
    disc_type = data.get("discount_type", "value")
    disc_val  = float(data.get("discount_value", 0) or 0)
    disc_amt  = round(disc_val, 2) if disc_type == "value" else round(subtotal * disc_val / 100, 2)
    tax_rate  = float(data.get("tax_rate", 0) or 0)
    tax_amt   = round((subtotal - disc_amt) * tax_rate / 100, 2)
    return subtotal, disc_amt, tax_amt, round(subtotal - disc_amt + tax_amt, 2)


def create_estimate(data: dict, lines: list) -> int:
    est_no = _next_estimate_number()
    subtotal, disc_amt, tax_amt, total = _money(lines, data)
    conn = get_db()
    with conn:
        cur = conn.execute(
            """INSERT INTO estimates(estimate_number,owner_id,pet_id,visit_id,doctor_name,
               issue_date,valid_until,status,subtotal,discount_type,discount_value,
               discount_amount,tax_rate,tax_amount,total,notes,created_by)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (est_no, data["owner_id"], data.get("pet_id"), data.get("visit_id"),
             data.get("doctor_name", ""), data.get("issue_date", date.today().isoformat()),
             data.get("valid_until"), data.get("status", "Draft"), subtotal,
             data.get("discount_type", "value"), float(data.get("discount_value", 0) or 0),
             disc_amt, float(data.get("tax_rate", 0) or 0), tax_amt, total,
             data.get("notes", ""), data.get("created_by", "")))
        est_id = cur.lastrowid
        for line in lines:
            lt = float(line.get("total", float(line.get("quantity", 1)) * float(line.get("unit_price", 0))))
            conn.execute(
                "INSERT INTO estimate_lines(estimate_id,line_type,item_id,description,"
                "quantity,unit_price,discount,total) VALUES(?,?,?,?,?,?,?,?)",
                (est_id, line.get("line_type", "service"), line.get("item_id"),
                 line.get("description", ""), float(line.get("quantity", 1)),
                 float(line.get("unit_price", 0)), float(line.get("discount", 0)), lt))
    conn.close()
    return est_id


def get_estimate(est_id: int) -> Optional[dict]:
    conn = get_db()
    row = conn.execute(
        "SELECT e.*, o.full_name owner_name, o.phone owner_phone, o.whatsapp_phone,"
        " p.pet_name FROM estimates e JOIN owners o ON o.id=e.owner_id"
        " LEFT JOIN pets p ON p.id=e.pet_id WHERE e.id=?", (est_id,)).fetchone()
    if not row:
        conn.close()
        return None
    est = dict(row)
    est["lines"] = [dict(r) for r in conn.execute(
        "SELECT * FROM estimate_lines WHERE estimate_id=? ORDER BY id", (est_id,)).fetchall()]
    conn.close()
    return est


def list_estimates(owner_id: Optional[int] = None, status: str = "",
                   limit: int = 100) -> list:
    conn = get_db()
    q = ("SELECT e.*, o.full_name owner_name, p.pet_name FROM estimates e"
         " JOIN owners o ON o.id=e.owner_id LEFT JOIN pets p ON p.id=e.pet_id WHERE 1=1")
    params = []
    if owner_id: q += " AND e.owner_id=?"; params.append(owner_id)
    if status:   q += " AND e.status=?";   params.append(status)
    q += " ORDER BY e.created_at DESC, e.id DESC LIMIT ?"; params.append(limit)
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def decide_estimate(est_id: int, decision: str, decided_by: str = "") -> None:
    """Record the client's answer. 'Approved' or 'Declined'."""
    conn = get_db()
    with conn:
        conn.execute(
            "UPDATE estimates SET status=?, decided_at=datetime('now'), decided_by=?,"
            " updated_at=datetime('now') WHERE id=?", (decision, decided_by, est_id))
    conn.close()


def convert_estimate(est_id: int, created_by: str = "") -> int:
    """Turn an approved estimate into a real invoice. Returns the invoice id.

    Guarded against double-conversion: two clicks on 'Convert' would otherwise
    bill the client twice for one quote. The guard is a re-read inside the same
    call rather than a UNIQUE constraint because invoice_id is nullable for
    every estimate that never converts.
    """
    est = get_estimate(est_id)
    if not est:
        raise ValueError("estimate not found")
    if est.get("invoice_id"):
        return int(est["invoice_id"])
    if est.get("status") != "Approved":
        raise ValueError("only an approved estimate can be converted")

    inv_id = create_invoice({
        "owner_id":       est["owner_id"],
        "pet_id":         est.get("pet_id"),
        "visit_id":       est.get("visit_id"),
        "doctor_name":    est.get("doctor_name", ""),
        "issue_date":     date.today().isoformat(),
        "discount_type":  est.get("discount_type", "value"),
        "discount_value": est.get("discount_value", 0),
        "tax_rate":       est.get("tax_rate", 0),
        "notes":          f"From estimate {est['estimate_number']}. {est.get('notes','') or ''}".strip(),
        "created_by":     created_by,
    }, est["lines"])

    conn = get_db()
    with conn:
        conn.execute(
            "UPDATE estimates SET status='Converted', invoice_id=?,"
            " updated_at=datetime('now') WHERE id=?", (inv_id, est_id))
    conn.close()
    return inv_id


# ─── Client deposits / account credit ─────────────────────────────────────────

def owner_credit_balance(owner_id: int) -> float:
    """Always derived from the ledger, never stored. See the table comment."""
    conn = get_db()
    v = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM owner_credits WHERE owner_id=?",
        (owner_id,)).fetchone()[0]
    conn.close()
    return round(float(v or 0), 2)


def list_owner_credits(owner_id: int, limit: int = 100) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM owner_credits WHERE owner_id=? ORDER BY id DESC LIMIT ?",
        (owner_id, limit)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_deposit(owner_id: int, amount: float, method: str = "Cash",
                reference: str = "", note: str = "", created_by: str = "") -> int:
    """Take money from a client before there is anything to bill it against."""
    amount = round(float(amount), 2)
    if amount <= 0:
        raise ValueError("a deposit must be a positive amount")
    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO owner_credits(owner_id,amount,kind,method,reference,note,created_by)"
            " VALUES(?,?,'deposit',?,?,?,?)",
            (owner_id, amount, method, reference, note, created_by))
        cid = cur.lastrowid
    conn.close()
    return cid


def apply_credit(owner_id: int, invoice_id: int, amount: float,
                 created_by: str = "") -> None:
    """Spend held credit against an invoice.

    Two guards, both of which protect real money:

      - never more than the client actually has on account, or the clinic would
        be crediting an invoice with money nobody ever paid;
      - never more than the invoice still owes, or the excess disappears -- the
        invoice cannot go below zero, so the credit would be consumed and not
        show up anywhere.

    The invoice side goes through add_payment() rather than touching
    paid_amount, so it lands in the same ledger as every other payment and is
    refundable and reconcilable by the same code.
    """
    amount = round(float(amount), 2)
    if amount <= 0:
        raise ValueError("the amount to apply must be positive")

    balance = owner_credit_balance(owner_id)
    if amount > balance:
        raise ValueError(f"only {balance:.2f} is available on account")

    inv = get_invoice(invoice_id)
    if not inv:
        raise ValueError("invoice not found")
    due = round(float(inv.get("due_amount") or 0), 2)
    if amount > due:
        raise ValueError(f"this invoice only owes {due:.2f}")

    if (inv.get("status") or "") == "Cancelled":
        raise ValueError("this invoice is cancelled — credit cannot be applied to it")

    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO owner_credits(owner_id,amount,kind,invoice_id,method,note,created_by)"
            " VALUES(?,?,'applied',?,'Credit',?,?)",
            (owner_id, -amount, invoice_id,
             f"Applied to invoice {inv.get('invoice_number','')}", created_by))
        credit_row_id = cur.lastrowid
    conn.close()

    # The deduction above is already committed. If the payment then fails —
    # add_payment() validates again and can raise — the client's money has been
    # taken off their account and put nowhere: the deposit is destroyed and the
    # screen shows a 500. Put it back before re-raising.
    try:
        add_payment(invoice_id, owner_id, amount, method="Credit",
                    reference="account credit", received_by=created_by)
    except Exception:
        conn = get_db()
        try:
            with conn:
                conn.execute("DELETE FROM owner_credits WHERE id=?", (credit_row_id,))
        except Exception:
            logger.error("COULD NOT REVERSE credit row %s for owner %s after a "
                         "failed payment on invoice %s — %.2f is missing from "
                         "their account", credit_row_id, owner_id, invoice_id, amount)
        finally:
            conn.close()
        raise


def refund_credit(owner_id: int, amount: float, note: str = "",
                  created_by: str = "") -> None:
    """Give unspent credit back to the client."""
    amount = round(float(amount), 2)
    if amount <= 0:
        raise ValueError("the refund must be a positive amount")
    balance = owner_credit_balance(owner_id)
    if amount > balance:
        raise ValueError(f"only {balance:.2f} is available to refund")
    conn = get_db()
    with conn:
        conn.execute(
            "INSERT INTO owner_credits(owner_id,amount,kind,method,note,created_by)"
            " VALUES(?,?,'refund','Cash',?,?)",
            (owner_id, -amount, note or "Refunded to client", created_by))
    conn.close()


def add_payment(invoice_id: int, owner_id: int, amount: float,
                method: str = "Cash", reference: str = "", received_by: str = "",
                idempotency_key: str = "") -> None:
    """Record a payment against an invoice.

    Now a thin wrapper over models.payments, which is the only path that writes
    money. It used to be the whole implementation, and it:

      - accepted `method`, `reference` and `received_by` and DISCARDED all three,
        so an invoice could be "Paid" with no record of who took the money, when
        or how, and nothing to reconcile the till against;
      - never wrote to the `payments` table at all — the docstring blamed a
        table belonging to another product, which had long since stopped being
        true — so there was no ledger, no refund and no reversal;
      - INCREMENTED paid_amount, which drifts away from reality the first time
        anything is refunded or replayed. It is now derived by summing the
        ledger, so it cannot.

    Keeping the signature means the two existing call sites did not change.
    """
    from models import payments
    intent = payments.create_intent(
        invoice_id, owner_id, amount,
        gateway=payments.gateway_for_method(method),
        idempotency_key=idempotency_key or "",
        created_by=received_by,
        reference=reference)
    payments.capture(intent["id"], actor=received_by)

def get_finance_summary(date_from: str = "", date_to: str = "") -> dict:
    conn = get_db()
    today = date.today().isoformat()
    df = date_from or today
    dt = date_to   or today
    # TWO different questions, and conflating them was the bug.
    #
    #   revenue   — accrual: what was INVOICED in this window and has been
    #               paid. A closed month's P&L must not move afterwards, and
    #               every figure on a historical report has to be derivable
    #               from the rows in that window.
    #   collected — cash: what actually ARRIVED at the till in this window,
    #               whenever the invoice was raised.
    #
    # The dashboard's "Today's Revenue" was showing `revenue`, so 120 EGP taken
    # today against last week's invoice appeared as 0 and the till never
    # reconciled with the screen. That screen wants `collected`; the P&L wants
    # `revenue`. Returning both means neither has to lie.
    revenue = conn.execute(
        "SELECT COALESCE(SUM(paid_amount),0) FROM invoices"
        " WHERE issue_date BETWEEN ? AND ? AND status IN ('Paid','Partial')",
        (df, dt)).fetchone()[0]
    try:
        collected = conn.execute(
            "SELECT COALESCE(SUM(amount),0) FROM payments"
            " WHERE DATE(received_at) BETWEEN ? AND ?", (df, dt)).fetchone()[0]
    except Exception:
        # An install predating the payments ledger has nothing better to offer.
        collected = revenue
    invoiced = conn.execute(
        "SELECT COALESCE(SUM(total),0) FROM invoices WHERE issue_date BETWEEN ? AND ? AND status!='Cancelled'",
        (df, dt)).fetchone()[0]
    outstanding = conn.execute(
        "SELECT COALESCE(SUM(due_amount),0) FROM invoices WHERE status IN ('Unpaid','Partial')").fetchone()[0]
    expenses = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE expense_date BETWEEN ? AND ?",
        (df, dt)).fetchone()[0]
    inv_count = conn.execute(
        "SELECT COUNT(*) FROM invoices WHERE issue_date BETWEEN ? AND ? AND status!='Cancelled'",
        (df, dt)).fetchone()[0]
    conn.close()
    return {
        "revenue": float(revenue or 0),
        # Money that actually arrived in this window. Use this for anything a
        # human would reconcile against the cash drawer.
        "collected": float(collected or 0),
        "invoiced": float(invoiced or 0),
        "outstanding": float(outstanding or 0),
        "expenses": float(expenses or 0),
        "net": float(revenue or 0) - float(expenses or 0),
        "invoice_count": int(inv_count or 0),
        "date_from": df, "date_to": dt,
    }

# ── REPORTS ────────────────────────────────────────────────────
def get_dashboard_stats() -> dict:
    conn = get_db()
    today = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()
    stats = {
        "owners_total":    conn.execute("SELECT COUNT(*) FROM owners").fetchone()[0],
        "pets_total":      conn.execute("SELECT COUNT(*) FROM pets").fetchone()[0],
        "visits_today":    conn.execute("SELECT COUNT(*) FROM visits WHERE visit_date=?", (today,)).fetchone()[0],
        "appts_today":     conn.execute("SELECT COUNT(*) FROM appointments WHERE appt_date=?", (today,)).fetchone()[0],
        "revenue_today":   float(conn.execute("SELECT COALESCE(SUM(paid_amount),0) FROM invoices WHERE issue_date=? AND status IN ('Paid','Partial')", (today,)).fetchone()[0] or 0),
        "revenue_month":   float(conn.execute("SELECT COALESCE(SUM(paid_amount),0) FROM invoices WHERE issue_date >= ? AND status IN ('Paid','Partial')", (month_start,)).fetchone()[0] or 0),
        "invoices_unpaid": conn.execute("SELECT COUNT(*) FROM invoices WHERE status IN ('Unpaid','Partial')").fetchone()[0],
        "outstanding":     float(conn.execute("SELECT COALESCE(SUM(due_amount),0) FROM invoices WHERE status IN ('Unpaid','Partial')").fetchone()[0] or 0),
        "low_stock_count": conn.execute("SELECT COUNT(*) FROM items i WHERE (SELECT COALESCE(SUM(b.quantity),0) FROM batches b WHERE b.item_id=i.id) <= i.reorder_level AND i.is_active=1").fetchone()[0],
        "expiry_soon":     conn.execute("SELECT COUNT(*) FROM batches WHERE expiry_date <= ? AND quantity>0", ((date.today()+timedelta(days=30)).isoformat(),)).fetchone()[0],
        "pending_reminders": conn.execute("SELECT COUNT(*) FROM reminders WHERE status='Pending'").fetchone()[0],
        "vip_owners":      conn.execute("SELECT COUNT(*) FROM owners WHERE vip_flag=1").fetchone()[0],
    }
    conn.close()
    return stats

def get_revenue_by_day(days: int = 30) -> list:
    conn = get_db()
    since = (date.today() - timedelta(days=days)).isoformat()
    # Accrual, matching get_finance_summary()['revenue'] and the P&L this chart
    # sits beside. For money-at-the-till use get_cash_by_day().
    rows = conn.execute(
        "SELECT issue_date d, COALESCE(SUM(paid_amount),0) revenue FROM invoices"
        " WHERE issue_date >= ? AND status IN ('Paid','Partial')"
        " GROUP BY issue_date ORDER BY d", (since,)).fetchall()
    conn.close()
    return [{"date": r["d"], "revenue": float(r["revenue"])} for r in rows]

def get_cash_by_day(days: int = 30) -> list:
    """Money that ARRIVED on each day, from the payments ledger.

    The cash-basis twin of get_revenue_by_day(). A spike here is a day the
    clinic actually took money, which is the question anyone reconciling a
    till is asking.
    """
    conn = get_db()
    since = (date.today() - timedelta(days=days)).isoformat()
    try:
        rows = conn.execute(
            "SELECT DATE(received_at) d, COALESCE(SUM(amount),0) revenue"
            " FROM payments WHERE DATE(received_at) >= ?"
            " GROUP BY DATE(received_at) ORDER BY d", (since,)).fetchall()
    except Exception:
        rows = []
    conn.close()
    return [{"date": r["d"], "revenue": float(r["revenue"])} for r in rows]


def get_top_services(limit: int = 10) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT description, COUNT(*) count, SUM(total) revenue FROM invoice_lines"
        " WHERE line_type='service' GROUP BY description ORDER BY revenue DESC LIMIT ?",
        (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ── SUPPLIERS ──────────────────────────────────────────────────
def list_suppliers() -> list:
    conn = get_db()
    rows = conn.execute("SELECT * FROM suppliers WHERE is_active=1 ORDER BY name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def create_supplier(data: dict) -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO suppliers(name,contact_name,phone,email,address,payment_terms,notes) VALUES(?,?,?,?,?,?,?)",
            (data["name"], data.get("contact_name",""), data.get("phone",""),
             data.get("email",""), data.get("address",""),
             data.get("payment_terms","Net 30"), data.get("notes","")))
        sid = cur.lastrowid
    conn.close()
    return sid

# ── REMINDERS / WHATSAPP ───────────────────────────────────────
def list_reminders(status: str = "", limit: int = 100) -> list:
    conn = get_db()
    q = "SELECT r.*, o.full_name owner_name, o.whatsapp_phone, p.pet_name FROM reminders r JOIN owners o ON o.id=r.owner_id LEFT JOIN pets p ON p.id=r.pet_id WHERE 1=1"
    params: list = []
    if status: q += " AND r.status=?"; params.append(status)
    q += " ORDER BY r.scheduled_for DESC LIMIT ?"; params.append(limit)
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def create_reminder(data: dict) -> int:
    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO reminders(owner_id,pet_id,appointment_id,reminder_type,message,channel,scheduled_for,created_by) VALUES(?,?,?,?,?,?,?,?)",
            (data["owner_id"], data.get("pet_id"), data.get("appointment_id"),
             data.get("reminder_type","appointment"), data.get("message",""),
             data.get("channel","WhatsApp"), data["scheduled_for"], data.get("created_by","")))
        rid = cur.lastrowid
    conn.close()
    return rid

def list_wa_templates() -> list:
    conn = get_db()
    rows = conn.execute("SELECT * FROM whatsapp_templates WHERE is_active=1 ORDER BY scenario, name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ── VACCINATIONS ───────────────────────────────────────────────
def list_vaccinations(pet_id: Optional[int] = None, limit: int = 100) -> list:
    conn = get_db()
    if pet_id:
        rows = conn.execute(
            "SELECT v.*, p.pet_name FROM vaccinations v JOIN pets p ON p.id=v.pet_id WHERE v.pet_id=? ORDER BY v.administered_at DESC",
            (pet_id,)).fetchall()
    else:
        rows = conn.execute(
            "SELECT v.*, p.pet_name, o.full_name owner_name FROM vaccinations v"
            " JOIN pets p ON p.id=v.pet_id JOIN owners o ON o.id=p.owner_id"
            " ORDER BY v.administered_at DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_upcoming_vaccines(days: int = 30) -> list:
    conn = get_db()
    threshold = (date.today() + timedelta(days=days)).isoformat()
    rows = conn.execute(
        "SELECT v.*, p.pet_name, o.full_name owner_name, o.whatsapp_phone FROM vaccinations v"
        " JOIN pets p ON p.id=v.pet_id JOIN owners o ON o.id=p.owner_id"
        " WHERE v.next_due_at <= ? AND v.next_due_at >= ? ORDER BY v.next_due_at",
        (threshold, date.today().isoformat())).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ── NOTIFICATIONS ─────────────────────────────────────────────

def create_notification(recipient_id: int, title: str, body: str = "",
                         icon: str = "🔔", link: str = "", module: str = "",
                         entity_type: str = "", entity_id: int = None,
                         recipient_role: str = "") -> None:
    try:
        conn = get_db()
        with conn:
            conn.execute(
                """INSERT INTO notifications(recipient_id,recipient_role,title,body,icon,link,module,entity_type,entity_id)
                   VALUES(?,?,?,?,?,?,?,?,?)""",
                (recipient_id, recipient_role, title, body, icon, link, module, entity_type, entity_id))
        conn.close()
    except Exception:
        # Same reasoning as log_audit: a notification failure must not
        # break the operation that triggered it, but it must not vanish
        # either. A notifier everyone trusts that quietly stopped
        # delivering is worse than one that was never wired up.
        logger.exception("could not write a notification")


def notify_role(role: str, title: str, body: str = "", icon: str = "🔔",
                link: str = "", module: str = "") -> None:
    """Send notification to all active users with a given role."""
    try:
        conn = get_db()
        users = conn.execute(
            "SELECT id FROM users WHERE role=? AND is_active=1", (role,)).fetchall()
        with conn:
            for u in users:
                conn.execute(
                    """INSERT INTO notifications(recipient_id,recipient_role,title,body,icon,link,module)
                       VALUES(?,?,?,?,?,?,?)""",
                    (u["id"], role, title, body, icon, link, module))
        conn.close()
    except Exception:
        # Same reasoning as log_audit: a notification failure must not
        # break the operation that triggered it, but it must not vanish
        # either. A notifier everyone trusts that quietly stopped
        # delivering is worse than one that was never wired up.
        logger.exception("could not write a notification")


def notify_managers(title: str, body: str = "", icon: str = "🔔",
                    link: str = "", module: str = "") -> None:
    """Notify all manager-level roles."""
    for role in ("super_admin", "clinic_owner", "branch_manager", "hr"):
        notify_role(role, title, body, icon, link, module)


def get_user_notifications(user_id: int, limit: int = 20) -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM notifications WHERE recipient_id=? ORDER BY created_at DESC LIMIT ?",
        (user_id, limit)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def count_unread_notifications(user_id: int) -> int:
    conn = get_db()
    n = conn.execute(
        "SELECT COUNT(*) FROM notifications WHERE recipient_id=? AND is_read=0",
        (user_id,)).fetchone()[0]
    conn.close()
    return n


def mark_notifications_read(user_id: int, notif_id: int = None) -> None:
    conn = get_db()
    with conn:
        if notif_id:
            conn.execute("UPDATE notifications SET is_read=1 WHERE id=? AND recipient_id=?",
                         (notif_id, user_id))
        else:
            conn.execute("UPDATE notifications SET is_read=1 WHERE recipient_id=?", (user_id,))
    conn.close()


# ── SERVICE CATALOG ────────────────────────────────────────────

def list_services(category: str = "", active_only: bool = True) -> list:
    conn = get_db()
    q = "SELECT * FROM service_catalog WHERE 1=1"
    params: list = []
    if active_only:
        q += " AND is_active=1"
    if category:
        q += " AND category=?"
        params.append(category)
    q += " ORDER BY category, sort_order, name"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_service(svc_id: int) -> Optional[dict]:
    conn = get_db()
    row = conn.execute("SELECT * FROM service_catalog WHERE id=?", (svc_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def upsert_service(data: dict) -> int:
    conn = get_db()
    svc_id = data.get("id")
    with conn:
        if svc_id:
            conn.execute(
                """UPDATE service_catalog SET code=?,name=?,name_ar=?,category=?,description=?,
                   standard_price=?,tax_rate=?,duration_min=?,species=?,is_active=?,
                   sort_order=?,updated_at=datetime('now') WHERE id=?""",
                (data.get("code",""), data["name"], data.get("name_ar",""),
                 data.get("category","Consultation"), data.get("description",""),
                 float(data.get("standard_price",0)), float(data.get("tax_rate",0)),
                 int(data.get("duration_min",0)), data.get("species","All"),
                 int(data.get("is_active",1)), int(data.get("sort_order",0)), svc_id))
        else:
            cur = conn.execute(
                """INSERT INTO service_catalog(code,name,name_ar,category,description,standard_price,
                   tax_rate,duration_min,species,is_active,sort_order) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (data.get("code",""), data["name"], data.get("name_ar",""),
                 data.get("category","Consultation"), data.get("description",""),
                 float(data.get("standard_price",0)), float(data.get("tax_rate",0)),
                 int(data.get("duration_min",0)), data.get("species","All"),
                 int(data.get("is_active",1)), int(data.get("sort_order",0))))
            svc_id = cur.lastrowid
    conn.close()
    return svc_id


def service_categories() -> list:
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT category FROM service_catalog WHERE is_active=1 ORDER BY category"
    ).fetchall()
    conn.close()
    return [r[0] for r in rows]


# ── LEGACY STATS (Excel) ───────────────────────────────────────
def _xlsx_count(path: str) -> int:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        n = max(0, ws.max_row - 1)
        wb.close()
        return n
    except Exception:
        return 0

def get_legacy_stats(legacy_data_dir: str) -> dict:
    stats = {"owners": 0, "pets": 0, "bookings_today": 0,
             "pending_reminders": 0, "total_bookings": 0}
    try:
        stats["owners"]         = _xlsx_count(os.path.join(legacy_data_dir, "owners.xlsx"))
        stats["pets"]           = _xlsx_count(os.path.join(legacy_data_dir, "pets.xlsx"))
        stats["total_bookings"] = _xlsx_count(os.path.join(legacy_data_dir, "bookings.xlsx"))
        today = datetime.now().strftime("%Y-%m-%d")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(os.path.join(legacy_data_dir,"bookings.xlsx"),read_only=True,data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
            try:
                di = headers.index("appointment_start")
                for row in ws.iter_rows(min_row=2,values_only=True):
                    if str(row[di] or "").startswith(today): stats["bookings_today"] += 1
            except (ValueError,TypeError): pass
            wb.close()
        except Exception: pass
    except Exception: pass
    return stats


# ── ROLES & PERMISSIONS ────────────────────────────────────────

ALL_PERMISSIONS = [
    ("patients",     "Manage Patients & Owners"),
    ("appointments", "Manage Appointments"),
    ("visits",       "Medical Visits & SOAP"),
    ("pharmacy",     "Pharmacy & Dispensing"),
    ("invoicing",    "Invoicing & Payments"),
    ("inventory",    "Inventory & Stock"),
    ("procurement",  "Procurement & Purchasing"),
    ("reports",      "Reports & Analytics"),
    ("whatsapp",     "WhatsApp Messaging"),
    ("catalog",      "Service Catalog"),
    ("grooming",     "Grooming"),
    ("boarding",     "Boarding"),
    ("hr",           "HR & Staff"),
    ("attendance",   "Attendance & Leave"),
    ("accounting",   "Accounting"),
    ("ai",           "AI Assistant"),
    ("system",       "System Admin"),
    ("backup",       "Backup & Restore"),
    ("audit",        "Audit Log"),
    ("settings",     "Platform Settings"),
    # Modules that existed with no grantable key, so the Roles screen could not
    # govern them at all — an administrator revoking everything still left them
    # open, with nothing in the UI to say so.
    ("payroll",      "Payroll & Salaries"),
    ("inpatient",    "Inpatient & Hospitalisation"),
    ("telemedicine", "Telemedicine"),
    ("imaging",      "Imaging & Radiology"),
    ("petshop",      "Pet Shop & Retail"),
]

# What each role may reach out of the box.
#
# Every role previously shipped with permissions_json='[]', which the access
# check reads as "no data — fall open". Combined with 271 routes carrying only
# @login_required, that meant a groomer could open the clinic's accounts, its
# purchase orders and its entire client list. Verified, not assumed.
#
# These are a STARTING POINT an administrator edits on the Roles screen, not a
# fixed policy. Each role gets what its job needs and nothing else: the test of
# a line below is "would this person be asked to do this in a real clinic".
#
# super_admin is absent deliberately — it bypasses the check entirely, so
# giving it a list would imply the list could restrict it.
DEFAULT_ROLE_PERMISSIONS = {
    "clinic_owner":   [k for k, _ in ALL_PERMISSIONS],
    "branch_manager": ["patients", "appointments", "visits", "pharmacy",
                       "invoicing", "inventory", "procurement", "reports",
                       "whatsapp", "catalog", "grooming", "boarding",
                       "attendance", "accounting", "inpatient", "telemedicine",
                       "imaging", "petshop"],
    # "attendance" is on every staff role deliberately. It is not an admin
    # module — it is where an employee clocks in and requests leave, and the
    # routes behind it are already scoped to the requesting user. Leaving it off
    # locked a nurse out of her own timesheet, which a test caught.
    #
    # Clinicians: the medical record and what they prescribe from it. No money.
    "doctor":         ["patients", "appointments", "visits", "pharmacy",
                       "reports", "catalog", "inpatient", "telemedicine",
                       "imaging", "ai", "attendance"],
    "nurse":          ["patients", "appointments", "visits", "pharmacy",
                       "inpatient", "imaging", "attendance"],
    # Front desk: books, bills and talks to clients. Not the clinic's accounts.
    "reception":      ["patients", "appointments", "invoicing", "catalog",
                       "whatsapp", "grooming", "boarding", "petshop",
                       "attendance"],
    "pharmacist":     ["pharmacy", "inventory", "patients", "visits", "attendance"],
    "inventory_mgr":  ["inventory", "procurement", "petshop", "reports", "attendance"],
    # Money roles: the books, but not the medical record.
    "finance":        ["invoicing", "accounting", "reports", "payroll"],
    "hr":             ["hr", "attendance", "payroll"],
    # Service staff: their own diary and the animals in front of them.
    "groomer":        ["grooming", "appointments", "patients", "attendance"],
    "boarding_staff": ["boarding", "appointments", "patients", "attendance"],
    "support_admin":  ["system", "backup", "audit", "settings"],
    # Read-only by role; these keys decide only WHICH screens open.
    "auditor":        ["reports", "audit", "accounting"],
}


def seed_default_permissions(conn) -> int:
    """Give every role with no grants its default set. Returns rows changed.

    Only touches roles whose permissions_json is empty, so an administrator's
    own configuration is never overwritten — including a deliberate empty one
    they set after this ran once.
    """
    import json as _json
    changed = 0
    for role, perms in DEFAULT_ROLE_PERMISSIONS.items():
        cur = conn.execute(
            "UPDATE roles SET permissions_json=? "
            "WHERE name=? AND (permissions_json IS NULL OR permissions_json IN ('', '[]'))",
            (_json.dumps(perms), role))
        changed += cur.rowcount or 0
    return changed


def list_roles() -> list:
    import json
    conn = get_db()
    rows = conn.execute("SELECT * FROM roles ORDER BY name").fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        try:
            d["permissions"] = json.loads(d.get("permissions_json") or "[]")
        except Exception:
            d["permissions"] = []
        result.append(d)
    return result


def get_role(role_id: int) -> Optional[dict]:
    import json
    conn = get_db()
    row = conn.execute("SELECT * FROM roles WHERE id=?", (role_id,)).fetchone()
    conn.close()
    if not row:
        return None
    d = dict(row)
    try:
        d["permissions"] = json.loads(d.get("permissions_json") or "[]")
    except Exception:
        d["permissions"] = []
    return d


def create_role(name: str, display_name: str, display_name_ar: str, permissions: list, color: str) -> int:
    import json
    conn = get_db()
    with conn:
        cur = conn.execute(
            "INSERT INTO roles(name,display_name,display_name_ar,permissions_json,color) VALUES(?,?,?,?,?)",
            (name.strip().lower().replace(" ", "_"), display_name, display_name_ar, json.dumps(permissions), color)
        )
        return cur.lastrowid


def update_role(role_id: int, display_name: str, display_name_ar: str, permissions: list, color: str) -> None:
    import json
    conn = get_db()
    with conn:
        conn.execute(
            "UPDATE roles SET display_name=?,display_name_ar=?,permissions_json=?,color=? WHERE id=?",
            (display_name, display_name_ar, json.dumps(permissions), color, role_id)
        )


def role_holders(role_id: int) -> list:
    """Active usernames still assigned to this role."""
    conn = get_db()
    try:
        row = conn.execute("SELECT name FROM roles WHERE id=?", (role_id,)).fetchone()
        if not row:
            return []
        return [r[0] for r in conn.execute(
            "SELECT username FROM users WHERE role=? AND is_active=1"
            " ORDER BY username", (row[0],)).fetchall()]
    finally:
        conn.close()


def delete_role(role_id: int) -> None:
    """Delete a role. REFUSES while staff still hold it.

    It used to be a bare DELETE that never touched `users`, leaving those
    people on a role name that no longer resolved. The permission check then
    fell open on an unknown role, so deleting a role SILENTLY PROMOTED
    everyone who held it — a nurse gained Finance, Accounting and Inventory,
    and the screen said "Role deleted."

    The fall-open is fixed too, so the failure mode today would be the
    opposite: those people locked out of everything. Neither is acceptable
    silently, so this refuses and names who is in the way.
    """
    holders = role_holders(role_id)
    if holders:
        raise ValueError(
            "%d staff member(s) still hold this role: %s. Move them to another "
            "role first." % (len(holders), ", ".join(holders[:10])))
    conn = get_db()
    with conn:
        conn.execute("DELETE FROM roles WHERE id=?", (role_id,))


def assign_user_role(user_id: int, role: str) -> None:
    conn = get_db()
    with conn:
        conn.execute("UPDATE users SET role=?,updated_at=datetime('now') WHERE id=?", (role, user_id))
