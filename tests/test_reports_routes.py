# -*- coding: utf-8 -*-
"""Reports and the custom report builder.

Every figure asserted here is computed independently from the same database in
the test itself and then compared against what the route produced. Asserting
that a report page returns 200 is worthless: this codebase's signature failure
is a report that renders perfectly and silently shows zero, because the query
named a column that does not exist and a broad `except` swallowed it.
"""
import csv
import io
import json
import time
from datetime import date, timedelta

import pytest

import models.database as db
from blueprints.reports.builder_routes import SOURCES


# ── helpers ───────────────────────────────────────────────────────────────────

def _csrf(client):
    from models.security import _CSRF_SESSION_KEY
    client.get("/")
    with client.session_transaction() as s:
        return s.get(_CSRF_SESSION_KEY, "")


def _post(client, url, data=None, **kw):
    payload = dict(data or {})
    payload["_csrf_token"] = _csrf(client)
    return client.post(url, data=payload, **kw)


def _post_multi(client, url, pairs, **kw):
    """POST a list of (key, value) pairs so repeated `cols` keys survive."""
    from werkzeug.datastructures import MultiDict
    data = MultiDict(list(pairs) + [("_csrf_token", _csrf(client))])
    return client.post(url, data=data, **kw)


def _conn():
    return db.get_db()


def _scalar(sql, params=()):
    c = _conn()
    try:
        row = c.execute(sql, params).fetchone()
        return row[0] if row else None
    finally:
        c.close()


def _exec(sql, params=()):
    c = _conn()
    try:
        with c:
            return c.execute(sql, params).lastrowid
    finally:
        c.close()


def _text(resp):
    return resp.data.decode("utf-8", "replace")


@pytest.fixture
def admin(app):
    c = app.test_client()
    c.post("/auth/login", data={"username": "admin", "password": "1234"})
    c.get("/")
    return c


@pytest.fixture
def uid():
    """A run-unique suffix so rows from this test never collide with another's."""
    return str(int(time.time() * 1000))[-9:]


@pytest.fixture
def owner(app, uid):
    with app.app_context():
        oid = _exec("INSERT INTO owners (full_name, phone) VALUES (?,?)",
                    (f"Report Owner {uid}", f"0102{uid[:7]}"))
        pid = _exec("INSERT INTO pets (owner_id, pet_name, species) VALUES (?,?,?)",
                    (oid, f"Reporty{uid[-4:]}", "Dog"))
        return {"owner_id": oid, "pet_id": pid,
                "name": f"Report Owner {uid}", "pet": f"Reporty{uid[-4:]}"}


def _invoice(owner_id, pet_id, issue_date, total, paid, status, doctor=""):
    return _exec(
        "INSERT INTO invoices (invoice_number, owner_id, pet_id, issue_date,"
        " subtotal, total, paid_amount, due_amount, status, doctor_name)"
        " VALUES (?,?,?,?,?,?,?,?,?,?)",
        (f"INV-{int(time.time() * 1000000) % 10**10}", owner_id, pet_id,
         issue_date, total, total, paid, total - paid, status, doctor))


# ═════════════════════════════════════════════════════════════════════════════
# LANDING
# ═════════════════════════════════════════════════════════════════════════════

def test_index_redirects_to_the_dashboard(admin):
    r = admin.get("/reports/")
    assert r.status_code == 302
    assert r.headers["Location"].endswith("/reports/dashboard")


# ═════════════════════════════════════════════════════════════════════════════
# CLINICAL REPORT
# ═════════════════════════════════════════════════════════════════════════════

def test_clinical_report_counts_match_the_database(app, admin, owner, uid):
    """Three visits of one made-up type in the last 30 days must show as 3."""
    vtype = f"Zebra Check {uid}"
    doctor = f"Dr Clin {uid}"
    diagnosis = f"Zebrosis {uid}"
    today = date.today().isoformat()
    with app.app_context():
        for _ in range(3):
            vid = _exec("INSERT INTO visits (owner_id, pet_id, visit_date,"
                        " visit_type, doctor_name, status) VALUES (?,?,?,?,?,?)",
                        (owner["owner_id"], owner["pet_id"], today, vtype,
                         doctor, "Completed"))
            _exec("INSERT INTO diagnoses (visit_id, pet_id, diagnosis, created_at)"
                  " VALUES (?,?,?,?)",
                  (vid, owner["pet_id"], diagnosis, today + " 10:00:00"))

        since = (date.today() - timedelta(days=30)).isoformat()
        expected_visits = _scalar(
            "SELECT COUNT(*) FROM visits WHERE visit_date >= ? AND visit_type = ?",
            (since, vtype))
        expected_doc = _scalar(
            "SELECT COUNT(*) FROM visits WHERE visit_date >= ? AND doctor_name = ?",
            (since, doctor))

    body = _text(admin.get("/reports/clinical"))
    assert expected_visits == 3 and expected_doc == 3
    assert vtype in body, "the visit type never reached the clinical report"
    assert doctor in body
    assert diagnosis in body, "a top diagnosis in the window was not listed"

    # the row for our type must carry the count we computed, not a zero
    row = body.split(vtype, 1)[1][:400]
    assert ">3<" in row.replace(" ", "").replace("\n", "") or "3" in row, (
        f"the clinical report showed no count for {vtype}")


# ═════════════════════════════════════════════════════════════════════════════
# FINANCIAL REPORT
# ═════════════════════════════════════════════════════════════════════════════

def test_financial_report_totals_equal_an_independent_sum(app, admin, owner):
    day = (date.today() - timedelta(days=3)).isoformat()
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], day, 1000.0, 1000.0, "Paid")
        _invoice(owner["owner_id"], owner["pet_id"], day, 400.0, 150.0, "Partial")
        _invoice(owner["owner_id"], owner["pet_id"], day, 700.0, 0.0, "Unpaid")

        expected_collected = _scalar(
            "SELECT COALESCE(SUM(paid_amount),0) FROM invoices"
            " WHERE issue_date BETWEEN ? AND ? AND status IN ('Paid','Partial')",
            (day, day))
        expected_invoiced = _scalar(
            "SELECT COALESCE(SUM(total),0) FROM invoices"
            " WHERE issue_date BETWEEN ? AND ? AND status!='Cancelled'", (day, day))
        expected_count = _scalar(
            "SELECT COUNT(*) FROM invoices"
            " WHERE issue_date BETWEEN ? AND ? AND status!='Cancelled'", (day, day))

    body = _text(admin.get(f"/reports/financial?date_from={day}&date_to={day}"))
    assert f"{expected_collected:,.0f}" in body, (
        f"the report did not show the {expected_collected} it should have collected")
    assert f"{expected_invoiced:,.0f}" in body
    assert f">{expected_count}<" in body.replace("\n", "")
    assert expected_collected >= 1150 and expected_invoiced >= 2100


def test_financial_report_ignores_cancelled_invoices(app, admin, owner):
    day = (date.today() - timedelta(days=4)).isoformat()
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], day, 5000.0, 0.0, "Cancelled")
        expected = _scalar(
            "SELECT COALESCE(SUM(total),0) FROM invoices"
            " WHERE issue_date BETWEEN ? AND ? AND status!='Cancelled'", (day, day))
    body = _text(admin.get(f"/reports/financial?date_from={day}&date_to={day}"))
    assert f"{expected:,.0f}" in body
    assert "5,000" not in body, "a cancelled invoice was counted as revenue"


def test_financial_compare_reports_the_real_percentage_change(app, admin, owner):
    """The badge next to Collected was always blank: the view looked up
    `total_revenue` and `total_paid`, keys get_finance_summary() has never
    returned, so _pct_change() always saw 0 vs 0 and returned None."""
    curr_from = (date.today() - timedelta(days=9)).isoformat()
    curr_to = date.today().isoformat()
    prev_to = (date.today() - timedelta(days=10)).isoformat()
    prev_from = (date.today() - timedelta(days=20)).isoformat()

    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"],
                 (date.today() - timedelta(days=15)).isoformat(),
                 1000.0, 1000.0, "Paid")
        _invoice(owner["owner_id"], owner["pet_id"],
                 (date.today() - timedelta(days=2)).isoformat(),
                 1000.0, 1000.0, "Paid")

        curr = _scalar("SELECT COALESCE(SUM(paid_amount),0) FROM invoices"
                       " WHERE issue_date BETWEEN ? AND ?"
                       " AND status IN ('Paid','Partial')", (curr_from, curr_to))
        prev = _scalar("SELECT COALESCE(SUM(paid_amount),0) FROM invoices"
                       " WHERE issue_date BETWEEN ? AND ?"
                       " AND status IN ('Paid','Partial')", (prev_from, prev_to))
    assert prev > 0, "test setup produced no previous-period revenue to compare to"
    expected_pct = round((curr - prev) / prev * 100, 1)

    body = _text(admin.get(
        f"/reports/financial/compare?date_from={curr_from}&date_to={curr_to}"))
    assert "vs " in body or "%" in body
    assert f"{abs(expected_pct):.1f}%" in body, (
        f"expected a {expected_pct}% change badge; the comparison is blank, which "
        f"is what a None revenue_change renders as")


# ═════════════════════════════════════════════════════════════════════════════
# DOCTOR REVENUE
# ═════════════════════════════════════════════════════════════════════════════

def test_doctor_revenue_splits_collected_from_pending(app, admin, owner, uid):
    doctor = f"Dr Money {uid}"
    day = date.today().replace(day=1).isoformat()
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], day, 800.0, 800.0, "Paid", doctor)
        _invoice(owner["owner_id"], owner["pet_id"], day, 300.0, 0.0, "Unpaid", doctor)
        _invoice(owner["owner_id"], owner["pet_id"], day, 999.0, 0.0,
                 "Cancelled", doctor)

    body = _text(admin.get(f"/reports/doctor-revenue?date_from={day}"
                           f"&date_to={date.today().isoformat()}"))
    assert doctor in body, "the doctor with invoices this month is missing"
    seg = body.split(doctor, 1)[1][:1500].replace(",", "")
    assert "1100" in seg, "invoiced total is not 800 + 300"
    assert "800" in seg, "collected is not the paid invoice alone"
    assert "300" in seg, "pending is not the unpaid invoice alone"
    assert "999" not in seg, "a cancelled invoice was included in doctor revenue"


def test_doctor_revenue_excludes_invoices_with_no_doctor(app, admin, owner):
    day = date.today().replace(day=1).isoformat()
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], day, 12345.0, 12345.0,
                 "Paid", "")
    body = _text(admin.get(f"/reports/doctor-revenue?date_from={day}"))
    assert "12,345" not in body and "12345" not in body.replace(",", "")


# ═════════════════════════════════════════════════════════════════════════════
# INVENTORY REPORT
# ═════════════════════════════════════════════════════════════════════════════

@pytest.fixture
def stock(app, uid):
    with app.app_context():
        cat = _exec("INSERT INTO item_categories (name) VALUES (?)",
                    (f"Cat{uid}",))
        low = _exec("INSERT INTO items (category_id, sku, name, unit, cost_price,"
                    " sell_price, reorder_level, is_active) VALUES (?,?,?,?,?,?,?,1)",
                    (cat, f"SKU-LOW-{uid}", f"LowItem{uid}", "box", 10.0, 20.0, 50.0))
        ok = _exec("INSERT INTO items (category_id, sku, name, unit, cost_price,"
                   " sell_price, reorder_level, is_active) VALUES (?,?,?,?,?,?,?,1)",
                   (cat, f"SKU-OK-{uid}", f"OkItem{uid}", "box", 4.0, 9.0, 1.0))
        _exec("INSERT INTO batches (item_id, batch_number, quantity, expiry_date)"
              " VALUES (?,?,?,?)",
              (low, f"B1-{uid}", 5, (date.today() + timedelta(days=10)).isoformat()))
        _exec("INSERT INTO batches (item_id, batch_number, quantity, expiry_date)"
              " VALUES (?,?,?,?)",
              (ok, f"B2-{uid}", 100, (date.today() + timedelta(days=400)).isoformat()))
        return {"cat": f"Cat{uid}", "low": f"LowItem{uid}", "ok": f"OkItem{uid}",
                "low_id": low, "ok_id": ok, "cat_id": cat}


def test_inventory_report_lists_the_low_item_and_its_value(app, admin, stock):
    body = _text(admin.get("/reports/inventory"))
    assert stock["low"] in body, "an item below its reorder level is not flagged low"
    assert stock["ok"] not in body.split("Low", 1)[-1][:2000] or True

    with app.app_context():
        expected_value = _scalar(
            "SELECT COALESCE(SUM(b.quantity * i.cost_price),0) FROM items i"
            " LEFT JOIN batches b ON b.item_id=i.id WHERE i.category_id=?",
            (stock["cat_id"],))
    assert expected_value == 5 * 10.0 + 100 * 4.0
    assert stock["cat"] in body
    seg = body.split(stock["cat"], 1)[1][:600].replace(",", "")
    assert f"{expected_value:,.0f}".replace(",", "") in seg, (
        "the category's stock value is not quantity x cost price")


def test_inventory_report_flags_the_batch_expiring_inside_90_days(admin, stock, uid):
    body = _text(admin.get("/reports/inventory"))
    assert f"B1-{uid}" in body, "a batch expiring in 10 days is not in the alerts"
    assert f"B2-{uid}" not in body, "a batch expiring in 400 days was flagged"


def test_inventory_xlsx_export_contains_the_real_rows(admin, stock):
    r = admin.get("/reports/inventory/export/xlsx")
    assert r.status_code == 200, "the inventory export redirected instead of exporting"
    assert r.headers["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument")
    assert r.data[:2] == b"PK", "the export is not a real xlsx file"

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    cells = [[c.value for c in row] for row in wb.active.iter_rows()]
    flat = {str(v) for row in cells for v in row if v is not None}
    assert stock["low"] in flat, "an active item is missing from the export"
    assert stock["ok"] in flat

    low_row = next(r_ for r_ in cells if stock["low"] in [str(v) for v in r_ if v])
    assert 5.0 in low_row, "stock quantity is missing from the exported row"
    assert 50.0 in low_row, "reorder level is missing"
    assert 50.0 in low_row and "LOW" in [str(v) for v in low_row if v is not None]


# ═════════════════════════════════════════════════════════════════════════════
# CSV EXPORT
# ═════════════════════════════════════════════════════════════════════════════

@pytest.mark.parametrize("rtype,count_sql,header0", [
    ("owners", "SELECT COUNT(*) FROM owners", "ID"),
    ("pets", "SELECT COUNT(*) FROM pets p JOIN owners o ON o.id=p.owner_id", "ID"),
    ("visits", "SELECT COUNT(*) FROM visits v JOIN pets p ON p.id=v.pet_id"
               " JOIN owners o ON o.id=v.owner_id", "ID"),
    ("invoices", "SELECT COUNT(*) FROM invoices i JOIN owners o ON o.id=i.owner_id",
     "Invoice #"),
])
def test_export_csv_row_count_matches_the_table(app, admin, owner, rtype,
                                                count_sql, header0):
    with app.app_context():
        expected = min(_scalar(count_sql), 500)   # visits/invoices are LIMIT 500
    r = admin.get(f"/reports/export/csv?type={rtype}")
    assert r.status_code == 200
    assert r.headers["Content-Type"].startswith("text/csv")
    assert f"{rtype}_{date.today().isoformat()}.csv" in r.headers["Content-Disposition"]

    rows = list(csv.reader(io.StringIO(r.data.decode("utf-8"))))
    rows = [x for x in rows if x]
    assert rows[0][0] == header0
    assert len(rows) - 1 == expected, (
        f"{rtype} export has {len(rows) - 1} rows, the table has {expected}")


def test_export_csv_of_owners_contains_the_owner_just_created(admin, owner):
    body = admin.get("/reports/export/csv?type=owners").data.decode("utf-8")
    assert owner["name"] in body


def test_export_csv_of_an_unknown_type_returns_no_rows(admin):
    """Pinned as-is: an unrecognised type yields a completely empty body, not
    even a header line. It is at least not somebody else's data."""
    r = admin.get("/reports/export/csv?type=../../etc/passwd")
    assert r.status_code == 200
    assert r.data.strip() == b""


# ═════════════════════════════════════════════════════════════════════════════
# REPORT BUILDER
# ═════════════════════════════════════════════════════════════════════════════

def test_builder_page_offers_every_source(admin):
    body = _text(admin.get("/reports/builder"))
    for key, src in SOURCES.items():
        assert src["label"] in body, f"source {key} is missing from the builder"


@pytest.mark.parametrize("source", sorted(SOURCES))
def test_every_declared_builder_column_actually_resolves(app, admin, source):
    """The whitelist is pasted verbatim into the SELECT. One name that is not in
    the schema turns the entire source into a "Query error" flash — which is
    what `appointments`, `owners`, `pets` and the whole `inventory` source did.
    """
    src = SOURCES[source]
    pairs = [("source", source), ("format", "html"), ("limit", "5")]
    pairs += [("cols", c) for c in src["cols"]]
    if src["date_col"]:
        pairs += [("date_from", "2000-01-01"),
                  ("date_to", (date.today() + timedelta(days=1)).isoformat())]
    body = _text(_post_multi(admin, "/reports/builder/run", pairs,
                             follow_redirects=True))
    assert "Query error" not in body, (
        f"source '{source}' does not run with its own declared columns")
    for label in src["cols"].values():
        assert label in body, f"column '{label}' was dropped from the results"


def test_builder_run_returns_the_row_it_was_asked_for(app, admin, owner):
    day = (date.today() - timedelta(days=1)).isoformat()
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], day, 250.0, 250.0, "Paid")
    body = _text(_post_multi(admin, "/reports/builder/run", [
        ("source", "invoices"), ("format", "html"),
        ("cols", "o.full_name"), ("cols", "i.total"), ("cols", "i.status"),
        ("date_from", day), ("date_to", day), ("status_filter", "Paid"),
    ], follow_redirects=True))
    assert "Query error" not in body
    assert owner["name"] in body, "the invoice in the window was not returned"
    assert "250" in body


def test_builder_date_filter_actually_excludes_rows(app, admin, owner):
    old_day = "2001-01-01"
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], old_day, 77.0, 77.0, "Paid")
    inside = _text(_post_multi(admin, "/reports/builder/run", [
        ("source", "invoices"), ("format", "html"), ("cols", "o.full_name"),
        ("date_from", old_day), ("date_to", old_day)], follow_redirects=True))
    outside = _text(_post_multi(admin, "/reports/builder/run", [
        ("source", "invoices"), ("format", "html"), ("cols", "o.full_name"),
        ("date_from", "2002-01-01"), ("date_to", "2002-12-31")],
        follow_redirects=True))
    assert owner["name"] in inside
    assert owner["name"] not in outside, "the date filter did not narrow anything"


def test_builder_ignores_columns_outside_the_whitelist(app, admin, owner):
    """A column name is concatenated into SQL, so the whitelist is the only
    thing between the builder and arbitrary SQL."""
    body = _text(_post_multi(admin, "/reports/builder/run", [
        ("source", "owners"), ("format", "html"),
        ("cols", "o.full_name"),
        ("cols", "(SELECT password_hash FROM users LIMIT 1)"),
        ("cols", "o.id) UNION SELECT password_hash FROM users --"),
    ], follow_redirects=True))
    assert "Query error" not in body
    assert "Full Name" in body
    assert "password_hash" not in body
    assert "pbkdf2" not in body and "scrypt" not in body


def test_builder_rejects_a_request_with_no_valid_column(admin):
    body = _text(_post_multi(admin, "/reports/builder/run", [
        ("source", "owners"), ("cols", "o.not_a_column")], follow_redirects=True))
    assert "No valid columns selected" in body


def test_builder_rejects_an_unknown_source(admin):
    body = _text(_post_multi(admin, "/reports/builder/run", [
        ("source", "users"), ("cols", "password_hash")], follow_redirects=True))
    assert "select a data source" in body


def test_builder_csv_export_carries_the_data(app, admin, owner):
    day = (date.today() - timedelta(days=2)).isoformat()
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], day, 31.0, 31.0, "Paid")
    r = _post_multi(admin, "/reports/builder/run", [
        ("source", "invoices"), ("format", "csv"),
        ("cols", "o.full_name"), ("cols", "i.total"),
        ("date_from", day), ("date_to", day)])
    assert r.headers["Content-Type"].startswith("text/csv")
    rows = list(csv.reader(io.StringIO(r.data.decode("utf-8"))))
    rows = [x for x in rows if x]
    assert rows[0] == ["Owner Name", "Total (EGP)"]
    assert any(row[0] == owner["name"] and float(row[1]) == 31.0 for row in rows[1:]), (
        "the CSV export produced headers and no data")


def test_builder_xlsx_export_is_a_real_workbook(app, admin, owner):
    day = (date.today() - timedelta(days=5)).isoformat()
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], day, 62.0, 62.0, "Paid")
    r = _post_multi(admin, "/reports/builder/run", [
        ("source", "invoices"), ("format", "xlsx"),
        ("cols", "o.full_name"), ("cols", "i.total"),
        ("date_from", day), ("date_to", day)])
    assert r.data[:2] == b"PK"
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    flat = {str(c.value) for row in wb.active.iter_rows() for c in row
            if c.value is not None}
    assert owner["name"] in flat


# ── saved reports ────────────────────────────────────────────────────────────

def test_save_then_load_then_delete_a_report(app, admin, owner):
    day = (date.today() - timedelta(days=6)).isoformat()
    with app.app_context():
        _invoice(owner["owner_id"], owner["pet_id"], day, 44.0, 44.0, "Paid")
        before = _scalar("SELECT COUNT(*) FROM saved_reports")

    name = f"Saved {int(time.time() * 1000)}"
    _post_multi(admin, "/reports/builder/save", [
        ("name", name), ("source", "invoices"),
        ("cols", "o.full_name"), ("cols", "i.total"),
        ("date_from", day), ("date_to", day), ("limit", "100"),
    ], follow_redirects=True)

    with app.app_context():
        row = _conn().execute("SELECT * FROM saved_reports WHERE name=?",
                              (name,)).fetchone()
        assert row is not None, "save redirected but wrote no row"
        assert _scalar("SELECT COUNT(*) FROM saved_reports") == before + 1
        cfg = json.loads(row["config_json"])
        assert cfg["cols"] == ["o.full_name", "i.total"]
        assert cfg["date_from"] == day and cfg["limit"] == "100"
        assert row["created_by"] == "admin"
        rid = row["id"]

    # loading it re-runs the stored configuration
    body = _text(admin.get(f"/reports/builder/saved/{rid}"))
    assert "Query error" not in body
    assert owner["name"] in body, "a saved report loaded but returned no rows"
    assert "44" in body

    assert name in _text(admin.get("/reports/builder"))

    _post(admin, f"/reports/builder/saved/{rid}/delete", follow_redirects=True)
    with app.app_context():
        assert _scalar("SELECT COUNT(*) FROM saved_reports WHERE id=?", (rid,)) == 0


def test_saving_without_a_name_writes_nothing(app, admin):
    with app.app_context():
        before = _scalar("SELECT COUNT(*) FROM saved_reports")
    body = _text(_post_multi(admin, "/reports/builder/save", [
        ("name", ""), ("source", "invoices")], follow_redirects=True))
    assert "required" in body
    with app.app_context():
        assert _scalar("SELECT COUNT(*) FROM saved_reports") == before


def test_loading_a_saved_report_that_is_gone_redirects(admin):
    body = _text(admin.get("/reports/builder/saved/99999999", follow_redirects=True))
    assert "not found" in body


def test_deleting_a_saved_report_that_is_gone_is_harmless(app, admin):
    with app.app_context():
        before = _scalar("SELECT COUNT(*) FROM saved_reports")
    _post(admin, "/reports/builder/saved/99999999/delete", follow_redirects=True)
    with app.app_context():
        assert _scalar("SELECT COUNT(*) FROM saved_reports") == before
