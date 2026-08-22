# -*- coding: utf-8 -*-
"""Build docs/market/COMPETITORS.xlsx from docs/market/01_COMPETITORS.md.

The research already existed as 287 lines of prose with sourced URLs and
confidence tags. This turns it into something sortable, filterable, and
usable in a meeting - without losing the two things that make the research
worth trusting: where every number came from, and how confident it is.

Every price converts through ONE exchange-rate cell, and the clinic size is an
input, so changing either re-prices all thirty-one products at once.

    python scripts/build_competitor_workbook.py
    python scripts/recalc.py docs/market/COMPETITORS.xlsx
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = os.path.join("docs", "market", "COMPETITORS.xlsx")

# ── house style ──────────────────────────────────────────────────────────────
FONT = "Arial"
INK = "1A1A1A"
BRAND = "1B6B5C"          # the app's own teal
BRAND_DK = "0D2B24"

H1 = Font(name=FONT, size=15, bold=True, color=BRAND_DK)
H2 = Font(name=FONT, size=11, bold=True, color=BRAND_DK)
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10, color=INK)
BODY_B = Font(name=FONT, size=10, bold=True, color=INK)
NOTE = Font(name=FONT, size=9, color="5B7169", italic=True)
INPUT = Font(name=FONT, size=10, bold=True, color="0000FF")   # blue = type here
LINKF = Font(name=FONT, size=10, color="0000FF", underline="single")

HDR_FILL = PatternFill("solid", fgColor=BRAND)
IN_FILL = PatternFill("solid", fgColor="FFFF00")              # yellow = input
US_FILL = PatternFill("solid", fgColor="E8F3EF")              # Aleefy's own row
WARN_FILL = PatternFill("solid", fgColor="FDECEA")

THIN = Side(style="thin", color="D9E2DF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")


def _hdr(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill, c.border = HDR, HDR_FILL, BOX
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


# ═════════════════════════════════════════════════════════════════════════════
# The data. Every row carries its source URL and a confidence tag, because a
# competitor table without those is a rumour with borders.
# ═════════════════════════════════════════════════════════════════════════════
#  name, tier, owner/origin, price text, USD/mo, model, confidence,
#  arabic, on-prem, MENA, threat, note, url
C = [
    # ── The one that matters most ───────────────────────────────────────────
    ("VetICare", "Regional (Arabic)", "Not disclosed", "$52/mo for 5 users",
     52.0, "Flat (5 users)", "Vendor-published", "Yes - explicit RTL", "No",
     "OPERATING IN EGYPT", "CRITICAL",
     "Names Egyptian clinics as customers: Pets Zone, Dr Men3am Pet Hospital, "
     "Almotawakkel Pet Center, Mojo Veterinary. Since 2020, claims 500+ clients. "
     "Ships pharmacy, lab with Exigo/Edan analyser integration, inventory, POS, "
     "boarding, HR, WhatsApp (+$20/mo), ZATCA e-invoicing, and a pet-owner mobile "
     "app with AI triage. No grooming, no telemedicine.",
     "https://veticareapp.com/price"),
    ("bAItari.vet", "Regional (Arabic)", "Oman", "Not disclosed",
     None, "Not disclosed", "Not disclosed", "Arabic-primary", "No", "Gulf", "HIGH",
     "Arabic-FIRST vet platform with AI documentation and annotated imaging. "
     "Tagline: 'the digital future of veterinary medicine in the Arab world' - "
     "the exact positioning Aleefy wants. Live signup.",
     "https://baitari.vet/"),
    ("Odoo vet modules", "Substitute", "BrowseInfo (India) et al.",
     "$272.20 one-time", None, "One-time", "Vendor-published",
     "Yes - Odoo core is Arabic + RTL", "Yes", "Egyptian partners exist", "HIGH",
     "A $272 module on a platform that already has Arabic, RTL, accounting, "
     "inventory, POS and HR, with Egyptian implementation partners (Macrofix, "
     "OEC-EG). A clinic can get ~70% of Aleefy tomorrow with an ERP vendor's "
     "support contract behind it.",
     "https://apps.odoo.com/apps/modules/13.0/bi_veterinary_management"),
    ("Yolo Clinic", "Regional (Arabic)", "UAE", "Not disclosed",
     None, "Not disclosed", "Not disclosed", "Arabic + English", "No",
     "UAE, markets into Egypt", "MEDIUM",
     "Vet line plus a human-medical line. Android/iOS/Huawei apps. ZATCA tax "
     "reporting. Appears in Egypt-facing Arabic search results.",
     "https://yolo.clinic/ar/"),
    ("Al-Mukhtabarat", "Regional (Arabic)", "EGYPT - Giza", "Not disclosed",
     None, "Not disclosed", "Not disclosed", "Arabic + English", "Not stated",
     "EGYPT", "MEDIUM",
     "The one genuinely Egyptian vet-adjacent vendor found. Vet lab + clinic "
     "system with barcode and medical-device result capture, WhatsApp result "
     "delivery, multi-branch stock transfer. Lab-led, but overlaps Aleefy's lab "
     "and clinical modules directly.",
     "https://almukhtabarat.com/"),
    ("Daftra", "Substitute", "Regional, Egypt phone line",
     "489.50-1,960 EGP/mo", None, "Tiered EGP", "Vendor-published",
     "Yes", "No", "Egypt, KSA, UAE, Jordan, Oman, Qatar", "MEDIUM",
     "PRICES IN EGP, publicly. Basic 489.50, Advanced 977.58, Comprehensive "
     "1,960 EGP/mo. No veterinary module - a vet buying this misuses a human "
     "clinic system. Sets the price expectation an Egyptian clinic arrives with.",
     "https://www.daftra.com/plans/"),
    ("Kawakeb Al-Teknologia", "Regional (Arabic)", "Saudi Arabia",
     "Not disclosed", None, "Not disclosed", "Not disclosed", "Arabic",
     "Yes - web or LAN", "Gulf", "MEDIUM",
     "Vet clinic module: records, appointments, radiology, lab, pharmacy, "
     "accounting, RBAC, vaccination tracking, WhatsApp. Runs on an internal "
     "network, so the on-premise story is not unique to Aleefy.",
     "https://www.const-tech.org/public/products/5"),
    ("Holool Alghad", "Regional (Arabic)", "Riyadh, Saudi Arabia",
     "Not disclosed", None, "Not disclosed", "Not disclosed", "Arabic + English",
     "Not stated", "Gulf", "LOW",
     "Vet clinic system: appointments, records, inventory, purchasing, SMS, "
     "financial analysis, device integration.",
     "https://holoolalghad.com/veterinary-clinic"),
    ("Petsphere (2t Interactive)", "Regional (Arabic)", "Lebanon", "Not found",
     None, "Not disclosed", "UNVERIFIED", "Reported Arabic-Lebanese RTL", "No",
     "Levant", "LOW",
     "UNVERIFIED - the site returned 404 when fetched; details come from the "
     "search index only. Dual currency USD/LBP.",
     "https://2tinteractive.com/solutions/petsphere/"),
    ("Veterical", "Regional (Arabic)", "Gamma Investments LLC",
     "App free, no pricing", None, "Not disclosed", "Third-party",
     "Listed among 33 languages", "No", "Unclear", "LOW",
     "Thin web presence. Arabic appears in the App Store language list; RTL "
     "quality unverified.",
     "https://apps.apple.com/us/app/id6736607434"),
    ("GVET", "Regional (Arabic)", "Latin America",
     "Not disclosed (3-month trial)", None, "Not disclosed", "Not disclosed",
     "Lists Arabic", "No", "None", "LOW",
     "Arabic is a translation checkbox, not a market focus. Low threat today - "
     "but proof that adding Arabic is cheap for an existing product.",
     "https://www.gvetsoft.com/en/"),
    ("vetPMS Cloud (Advitech)", "Regional (Arabic)", "Reseller: Dubai",
     "Not disclosed", None, "Not disclosed", "Not disclosed",
     "Arabic not mentioned", "No", "UAE/Qatar/Kuwait/Saudi/Oman", "LOW",
     "Shows the actual MENA channel model: a local medical-equipment "
     "distributor reselling a foreign PIMS. This is how software reaches Gulf "
     "clinics - through distributors, not direct sales.",
     "https://medicalplus.ae/product/vetpms-cloud-veterinary-practice-management-software/"),

    # ── T1 global incumbents ────────────────────────────────────────────────
    ("ezyVet", "Global", "IDEXX (acq. Jun 2021)", "$260.50/mo",
     260.50, "Flat or per-user (disputed)", "Vendor-published", "No", "No",
     "None found", "LOW",
     "The only major publishing a price on its own site. Vendor says per month; "
     "Capterra and GetApp both render it as per user per month. Unresolved - if "
     "the aggregators are right, a 6-seat Egyptian clinic pays ~79,000 EGP/mo.",
     "https://www.ezyvet.com/pricing/us"),
    ("Provet Cloud", "Global", "Nordhealth (NO/FI)", "$99/vet/mo (Core)",
     99.0, "Per vet", "Third-party", "No - but ships HEBREW", "No",
     "None confirmed", "MEDIUM",
     "THE MOST IMPORTANT LINE IN THE GLOBAL TIER. Provet already ships Hebrew "
     "among 16 locales, which means RTL is already solved in its codebase. "
     "Arabic would be a translation project, not an engineering one. RTL is not "
     "the moat it looks like. Its per-vet, free-other-seats model is also the "
     "most Egypt-compatible of the Western majors.",
     "https://www.capterra.co.uk/software/137569/provet-cloud"),
    ("IDEXX Cornerstone", "Global", "IDEXX", "$549/mo (also cited $420, $399)",
     549.0, "Per practice", "Third-party, wide spread", "No",
     "Yes - officially documented", "None found", "LOW",
     "Genuinely on-premise, and IDEXX is actively migrating customers off it.",
     "https://www.idexx.com/"),
    ("IDEXX Neo", "Global", "IDEXX",
     "$290/mo + $2,375 setup + $590 conversion", 290.0, "Flat + setup",
     "Third-party", "No", "No", "None found", "LOW",
     "Setup and data conversion nearly triple the first-year cost.",
     "https://www.capterra.com/"),
    ("IDEXX Animana", "Global", "IDEXX (acq. 2014)", "Not disclosed",
     None, "Per user", "Not disclosed", "No - Dutch/English/German", "Doubtful",
     "NL/DE/UK only", "LOW", "Europe-only footprint.", "https://www.idexx.com/"),
    ("Covetrus Pulse", "Global", "Covetrus", "Not disclosed",
     None, "Not disclosed", "Not disclosed", "No evidence", "No",
     "EMEA unit exists (product is Ascend)", "LOW",
     "Covetrus is the closest any global vendor gets to MENA - it runs a 'UK, "
     "Europe, Middle East and Africa' unit. But no Arabic and no confirmed "
     "Arab-country deployment.",
     "https://software.covetrus.com/emea/"),
    ("Covetrus AVImark", "Global", "Covetrus", "Not disclosed",
     None, "Not disclosed", "Not disclosed", "No evidence",
     "Yes - real Windows server", "None", "LOW",
     "Documented Windows server install, peer-to-peer capped at nine thick "
     "clients. A legacy product its owner is migrating customers off.",
     "https://covetrus.com/"),
    ("Covetrus Impromed", "Global", "Covetrus", "$5,000 one-time",
     None, "One-time", "Third-party", "No evidence", "Legacy on-prem",
     "None", "LOW", "Legacy perpetual licence.", "https://covetrus.com/"),
    ("Shepherd", "Global", "Independent (US)",
     "$299/mo first vet + $99/additional", 299.0, "Per vet", "Third-party (4 sources)",
     "No evidence", "No (anti-server positioning)", "None", "LOW",
     "Explicitly positions against on-premise servers.", "https://www.capterra.com/"),
    ("Digitail", "Global", "Independent (RO/US)", "$289/user/mo",
     289.0, "Per user", "Third-party + vendor blog", "No (UNVERIFIED)", "No",
     "None", "LOW", "Pricing page displays no numbers.", "https://www.capterra.com/"),
    ("VETport", "Global", "US (Ohio), est. 2002", "$229/mo usage-based",
     229.0, "Usage-based", "Third-party", "No - EN/FR/DE", "No",
     "India office; UAE in country list", "MEDIUM",
     "Second-closest global vendor to MENA: 37 countries claimed, a real India "
     "office, UAE in its dropdown. Its pricing page is a calculator that renders "
     "$0 and funnels to 'Get a Quote'.",
     "https://www.vetport.com/pricing"),
    ("Vetspire", "Global", "US (Palo Alto)", "Not disclosed",
     None, "Not disclosed", "Hostile source only", "No evidence", "No",
     "None", "LOW",
     "The '$349/DVM/mo' figure circulating comes from two COMPETITORS' blogs - "
     "treat as hostile and unusable.", "https://www.vetspire.com/"),
    ("Hippo Manager", "Global", "US (Lexington KY)",
     "$119/mo per FT vet, unlimited other seats", 119.0, "Per vet",
     "Third-party (most transparent)", "No evidence", "No",
     "The ONLY priced product in Capterra's UAE directory", "MEDIUM",
     "Capterra's UAE veterinary directory lists 25 products and shows a price "
     "for exactly one - this. Migration $1,750 one-time.",
     "https://www.capterra.ae/directory/30617/veterinary/software"),
    ("ezVetPro", "Global", "ezofficesystems Ltd (UK)", "Not disclosed",
     None, "Not disclosed", "Unusable", "No - English only",
     "Yes - on-prem or hosted", "None", "LOW",
     "GetApp shows '$600' with no unit and no period. Unusable as a data point.",
     "https://www.getapp.com/"),
    ("OpenVPMS", "Global", "Not-for-profit (AU)",
     "A$450 + GST per FTE vet/year", 24.0, "Per vet per year",
     "Vendor-published", "Not documented (UNVERIFIED)",
     "Yes - Tomcat + MySQL self-hosted", "None", "MEDIUM",
     "OPEN SOURCE. Self-install for testing costs nothing. A$450/FTE-vet/year "
     "is roughly $24/mo equivalent - by far the cheapest credible option, and "
     "genuinely self-hostable.",
     "https://openvpms.org/"),

    # ── Substitutes ─────────────────────────────────────────────────────────
    ("Medicakare", "Substitute", "Riyadh, CAIRO, Dubai, London",
     "Free tier for one clinic", 0.0, "Freemium", "Vendor-published", "Yes",
     "No", "Egypt office", "MEDIUM",
     "A free tier, with a Cairo office. No veterinary module - human "
     "specialties only - but free is a hard price to argue against.",
     "https://www.medicakare.com/ar/"),
    ("Fekra IT", "Substitute", "Al-Madinah, Saudi Arabia", "Not disclosed",
     None, "Not disclosed", "Not disclosed", "Arabic", "Yes - Windows desktop",
     "Gulf", "LOW",
     "Windows desktop, on-premise, works with NO INTERNET. The offline model "
     "Aleefy assumes is what regional buyers already recognise and expect.",
     "https://fekrait.com/"),
    ("ClinicGateway", "Substitute", "Saudi/UAE", "From 2,500 SAR/mo",
     666.0, "Tiered", "Vendor-published", "Full Arabic + RTL + Hijri", "No",
     "Gulf", "LOW",
     "No veterinary anywhere on the site. Useful as a Gulf price anchor: it "
     "benchmarks Kareo at 2,800+ and enterprise systems at 8,000-15,000 SAR/mo.",
     "https://clinicgateway.ae/"),
    ("Nitco", "Substitute", "EGYPT", "Publishes no prices",
     None, "Not disclosed", "Not disclosed", "Arabic", "Not stated", "Egypt",
     "LOW",
     "Publishes an article titled 'clinic management software prices in Egypt' "
     "that contains no prices. Emblematic of the market's opacity.",
     "https://nitcotek.com/"),
    ("Paper, Excel and WhatsApp", "Substitute", "-", "Free",
     0.0, "Free", "Observed", "Yes", "Yes", "Everywhere", "CRITICAL",
     "THE REAL INCUMBENT. No study of software adoption in Egyptian veterinary "
     "practice exists. What does exist: a clinic costs ~60,000 EGP to open and a "
     "consultation grosses ~150 EGP. A Gulf buyer specified a system "
     "feature-for-feature like Aleefy on Mostaql and expected to pay "
     "$1,000-$2,500 ONCE, forever. Thirty developers competed for it.",
     "-"),
]

# Aleefy's own row, kept separate so it is never mistaken for research.
ALEEFY = ("Aleefy (you)", "-", "Egypt", "7,000 EGP once, or 600 EGP/mo",
          None, "Licence or subscription", "Own product",
          "Yes - complete, incl. PDFs", "Yes", "Egypt", "-",
          "413 screens, 34 modules, 2,218 automated tests. Arabic complete "
          "including inside generated PDFs, which is where competing systems "
          "break. Covers grooming, boarding, payroll and pet-shop retail, which "
          "no competitor in this sheet covers together.", "-")


def build():
    wb = Workbook()

    # ═══ 1. Read me ═════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Read me"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Aleefy - Competitive Landscape"
    ws["A1"].font = H1
    ws["A2"] = ("Built from docs/market/01_COMPETITORS.md. Research date "
                "2026-07-28, by web search in English and Arabic.")
    ws["A2"].font = BODY
    ws["A4"] = "How to read the confidence column"
    ws["A4"].font = H2

    for i, (tag, meaning) in enumerate([
        ("Vendor-published", "The number appears on the vendor's own site. Highest confidence."),
        ("Third-party", "From Capterra / GetApp / press. Aggregator pricing is often stale or mis-labelled."),
        ("Not disclosed", "The vendor's pricing page contains no numbers. This is a FINDING, not a gap - "
                          "of 15 global products, only 3 publish a real price."),
        ("UNVERIFIED", "Inferred, or from a snippet that could not be opened. Treat as a lead, not a fact."),
        ("Hostile source", "The only figure available came from a competitor's marketing. Unusable."),
    ], start=5):
        ws.cell(row=i, column=1, value=tag).font = BODY_B
        ws.cell(row=i, column=2, value=meaning).font = BODY
        ws.cell(row=i, column=2).alignment = WRAP

    ws["A11"] = "What the research actually concluded"
    ws["A11"].font = H2
    findings = [
        ("The Arabic vet niche is NOT empty.",
         "VetICare ships Arabic with RTL, names four Egyptian clinics as customers, has run since "
         "2020, and publishes $52/month. Four of the six things listed as Aleefy's advantages are "
         "already shipped by an incumbent - plus two things Aleefy lacks: lab-machine integration "
         "and a pet-owner mobile app."),
        ("RTL is not a moat.",
         "Provet Cloud already ships Hebrew among 16 locales, so its RTL plumbing exists. Adding "
         "Arabic would be a translation project, not an engineering one."),
        ("On-premise is not a wedge.",
         "Four other products self-host: Cornerstone, AVImark, ezVetPro and OpenVPMS. It is a "
         "feature Aleefy shares, not one that distinguishes it."),
        ("Nobody supports Arabic among the global fifteen.",
         "Arabic support was found zero times across all T1 products. That part of the thesis holds."),
        ("The real competitor is paper.",
         "A clinic costs ~60,000 EGP to open; a consultation grosses ~150 EGP. ezyVet's entry price "
         "is 88 consultations a month before the vet earns anything. Even VetICare is 17.5."),
        ("The price ceiling is set by the market, not the software.",
         "Egyptian pharmacy software: 5,000-12,000 EGP perpetual, or 500-1,500 EGP/month. Daftra: "
         "489-1,960 EGP/month. Aleefy at 7,000 EGP once or 600 EGP/month sits inside both bands."),
    ]
    r = 12
    for title, body in findings:
        ws.cell(row=r, column=1, value=title).font = BODY_B
        ws.cell(row=r, column=1).alignment = WRAP
        ws.cell(row=r, column=2, value=body).font = BODY
        ws.cell(row=r, column=2).alignment = WRAP
        ws.row_dimensions[r].height = 46
        r += 1

    ws.cell(row=r + 1, column=1,
            value="Every row in 'All competitors' carries the URL its claims came from. "
                  "Where a cell says Not disclosed, that is what the vendor publishes - "
                  "it is not a gap in the research.").font = NOTE
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 104

    # ═══ 2. All competitors ═════════════════════════════════════════════════
    ws = wb.create_sheet("All competitors")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "All competitors and substitutes"
    ws["A1"].font = H1
    ws["A2"] = ("Sorted by threat. Filter on any column. Aleefy's own row is shaded "
                "and is not research - it is what you have.")
    ws["A2"].font = NOTE

    heads = ["Product", "Tier", "Owner / origin", "Published price", "Pricing model",
             "Confidence", "Arabic / RTL", "On-premise", "MENA presence",
             "Threat", "What matters about it", "Source"]
    widths = [24, 18, 26, 30, 22, 22, 26, 22, 30, 10, 78, 44]
    _hdr(ws, 4, heads, widths)

    order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "-": 4}
    rows = sorted(C, key=lambda x: (order.get(x[10], 5), x[0]))
    rows = [ALEEFY] + rows

    r = 5
    for (name, tier, owner, price, _usd, model, conf, ar, prem, mena,
         threat, note, url) in rows:
        vals = [name, tier, owner, price, model, conf, ar, prem, mena,
                threat, note, url]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY_B if i == 1 else BODY
            c.alignment = WRAP
            c.border = BOX
            if name.startswith("Aleefy"):
                c.fill = US_FILL
            elif threat == "CRITICAL":
                c.fill = WARN_FILL
        if url and url != "-":
            lc = ws.cell(row=r, column=12)
            lc.hyperlink = url
            lc.font = LINKF
        ws.row_dimensions[r].height = 58
        r += 1

    ws.auto_filter.ref = "A4:L%d" % (r - 1)
    ws.freeze_panes = "B5"

    # ═══ 3. Pricing in EGP ══════════════════════════════════════════════════
    ws = wb.create_sheet("Pricing in EGP")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "What each system costs an Egyptian clinic"
    ws["A1"].font = H1
    ws["A2"] = ("Change the three yellow cells and every figure below re-prices. "
                "Only products with a published price appear here.")
    ws["A2"].font = NOTE

    ws["A4"] = "EGP per USD"
    ws["A4"].font = BODY_B
    ws["B4"] = 50.5
    ws["B4"].font, ws["B4"].fill, ws["B4"].border = INPUT, IN_FILL, BOX
    ws["C4"] = "2026 average. Source: exchange-rates.org"
    ws["C4"].font = NOTE

    ws["A5"] = "Vets in the clinic"
    ws["A5"].font = BODY_B
    ws["B5"] = 3
    ws["B5"].font, ws["B5"].fill, ws["B5"].border = INPUT, IN_FILL, BOX
    ws["C5"] = "Drives the per-vet products"
    ws["C5"].font = NOTE

    ws["A6"] = "Consultation fee (EGP)"
    ws["A6"].font = BODY_B
    ws["B6"] = 150
    ws["B6"].font, ws["B6"].fill, ws["B6"].border = INPUT, IN_FILL, BOX
    ws["C6"] = "Used for the last column. Source: small-projects.org [UNVERIFIED]"
    ws["C6"].font = NOTE

    heads2 = ["Product", "Published price", "Pricing model", "USD / month",
              "EGP / month", "Consultations to cover it", "Confidence"]
    _hdr(ws, 8, heads2, [26, 34, 24, 14, 14, 24, 24])

    priced = [(x[0], x[3], x[4], x[5], x[6]) for x in C if x[4] is not None]
    priced.sort(key=lambda x: x[2])

    r = 9
    for name, price, usd, model, conf in priced:
        ws.cell(row=r, column=1, value=name).font = BODY_B
        ws.cell(row=r, column=2, value=price).font = BODY
        ws.cell(row=r, column=3, value=model).font = BODY
        # Per-vet products multiply by the vet count; everything else does not.
        if "Per vet" in model:
            ws.cell(row=r, column=4, value="=%s*$B$5" % usd)
        else:
            ws.cell(row=r, column=4, value=usd)
        ws.cell(row=r, column=5, value="=D%d*$B$4" % r)
        ws.cell(row=r, column=6, value="=IF($B$6=0,\"\",E%d/$B$6)" % r)
        ws.cell(row=r, column=7, value=conf).font = BODY
        for i in range(1, 8):
            c = ws.cell(row=r, column=i)
            c.border = BOX
            c.alignment = WRAP if i in (2, 7) else TOP
            if i in (4, 5):
                c.number_format = '#,##0;(#,##0);-'
                c.font = BODY
            if i == 6:
                c.number_format = '#,##0.0;(#,##0.0);-'
                c.font = BODY
        r += 1

    # Aleefy's own two options, as formulas off the same inputs.
    r += 1
    ws.cell(row=r, column=1, value="Aleefy - subscription").font = BODY_B
    ws.cell(row=r, column=2, value="600 EGP/mo, unlimited users").font = BODY
    ws.cell(row=r, column=3, value="Flat").font = BODY
    ws.cell(row=r, column=4, value="=E%d/$B$4" % r).number_format = '#,##0;(#,##0);-'
    ws.cell(row=r, column=5, value=600).number_format = '#,##0;(#,##0);-'
    ws.cell(row=r, column=6, value="=IF($B$6=0,\"\",E%d/$B$6)" % r).number_format = '#,##0.0'
    ws.cell(row=r, column=7, value="Own product").font = BODY
    for i in range(1, 8):
        ws.cell(row=r, column=i).fill = US_FILL
        ws.cell(row=r, column=i).border = BOX
    r += 1
    ws.cell(row=r, column=1, value="Aleefy - licence").font = BODY_B
    ws.cell(row=r, column=2, value="7,000 EGP once + 1,800/yr from month 7").font = BODY
    ws.cell(row=r, column=3, value="One-time + renewal").font = BODY
    ws.cell(row=r, column=4, value="=E%d/$B$4" % r).number_format = '#,##0;(#,##0);-'
    ws.cell(row=r, column=5, value="=1800/12").number_format = '#,##0;(#,##0);-'
    ws.cell(row=r, column=6, value="=IF($B$6=0,\"\",E%d/$B$6)" % r).number_format = '#,##0.0'
    ws.cell(row=r, column=7, value="Own product").font = BODY
    for i in range(1, 8):
        ws.cell(row=r, column=i).fill = US_FILL
        ws.cell(row=r, column=i).border = BOX

    r += 2
    ws.cell(row=r, column=1,
            value="The last column is the point. At a 150 EGP consultation, "
                  "ezyVet costs a small clinic dozens of consultations a month "
                  "before the vet earns anything. That is the ceiling on what "
                  "any vet software can charge in Egypt - it is set by the "
                  "market, not by what the software is worth.").font = NOTE
    ws.cell(row=r, column=1).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=7)

    # ═══ 4. Feature matrix ══════════════════════════════════════════════════
    ws = wb.create_sheet("Feature matrix")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Aleefy against the systems that actually compete in Egypt"
    ws["A1"].font = H1
    ws["A2"] = ("Only products with enough published detail to judge. A blank "
                "means not documented either way - not proof of absence.")
    ws["A2"].font = NOTE

    cols = ["Capability", "Aleefy", "VetICare", "bAItari.vet", "Odoo vet",
            "Daftra", "ezyVet", "Provet Cloud"]
    _hdr(ws, 4, cols, [40, 14, 14, 14, 14, 14, 14, 16])

    Y, N, B = "Yes", "No", ""
    matrix = [
        ("Arabic interface", Y, Y, Y, Y, Y, N, N),
        ("Right-to-left layout", Y, Y, Y, Y, Y, N, "Hebrew only"),
        ("Arabic inside generated PDFs", Y, B, B, B, B, N, N),
        ("Runs without internet / on-premise", Y, N, N, Y, N, N, N),
        ("Clinical records", Y, Y, Y, Y, N, Y, Y),
        ("Pharmacy and dispensing", Y, Y, B, Y, N, Y, Y),
        ("Laboratory", Y, Y, B, Y, N, Y, Y),
        ("Lab-machine integration", N, Y, B, B, N, Y, Y),
        ("Inventory with batch and expiry", Y, Y, B, Y, Y, Y, Y),
        ("Procurement and suppliers", Y, B, B, Y, Y, B, B),
        ("Invoicing", Y, Y, Y, Y, Y, Y, Y),
        ("Accounting (P&L, cash flow)", Y, B, B, Y, Y, B, B),
        ("HR and attendance", Y, Y, B, Y, B, N, N),
        ("Payroll", Y, B, B, Y, B, N, N),
        ("Pet shop / retail POS", Y, Y, B, Y, Y, N, N),
        ("Grooming", Y, N, B, B, N, B, B),
        ("Boarding", Y, Y, B, Y, N, B, B),
        ("Telemedicine", Y, N, B, B, N, B, B),
        ("WhatsApp messaging", Y, "Yes (+$20/mo)", B, Y, B, N, N),
        ("Pet-owner mobile app", N, Y, B, B, N, Y, Y),
        ("AI triage / documentation", B, Y, Y, B, N, B, B),
        ("Multi-branch", Y, Y, B, Y, Y, Y, Y),
        ("Published price", Y, Y, N, Y, Y, Y, N),
        ("Paying customers in Egypt", N, Y, N, Y, Y, N, N),
    ]
    r = 5
    for row in matrix:
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY_B if i == 1 else BODY
            c.alignment = WRAP if i == 1 else Alignment(horizontal="center", vertical="top")
            c.border = BOX
            if i == 2:
                c.fill = US_FILL
        ws.row_dimensions[r].height = 17
        r += 1

    r += 1
    ws.cell(row=r, column=1,
            value="The last two rows are the ones to sit with. Aleefy covers more "
                  "than anything else here - and is the only column with 'No' "
                  "against paying customers in Egypt. Breadth is not the "
                  "bottleneck; proof is.").font = NOTE
    ws.cell(row=r, column=1).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=8)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("wrote %s" % OUT)
    print("  sheets    : %s" % ", ".join(wb.sheetnames))
    print("  products  : %d (plus Aleefy)" % len(C))
    print("  priced    : %d" % len(priced))


if __name__ == "__main__":
    build()
