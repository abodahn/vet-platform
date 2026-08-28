# -*- coding: utf-8 -*-
"""Build and work the target-market database.

    python scripts/market_db.py import clinics.csv
    python scripts/market_db.py score
    python scripts/market_db.py cohorts
    python scripts/market_db.py status
    python scripts/market_db.py export market.xlsx
    python scripts/market_db.py update market.xlsx --apply
    python scripts/market_db.py funnel
    python scripts/market_db.py why "Cairo Vet Hospital"

WHY A DATABASE AND NOT AN AD AUDIENCE

There is a countable number of veterinary clinics in Cairo and Giza. Each one
either becomes a customer or does not. That is a list to be worked, not an
audience to be targeted, and the difference decides where the money goes: a
list costs time, an audience costs money you do not have.

THE CSV IT IMPORTS

One row per clinic. Only `name` is required; everything else improves the score
or makes the clinic reachable. Unknown columns are ignored rather than
refused, so a scrape can hand over whatever it found.

    name, name_ar, governorate, district, address, phone, whatsapp, email,
    website, facebook, instagram, maps_url, contact_name, contact_role,
    branches, vets, is_hospital, has_grooming, has_boarding, has_pharmacy,
    has_petshop, has_lab, current_software, notes, source, source_url

Booleans accept 1/0, yes/no, true/false, and Arabic نعم/لا.

Re-importing is safe: a clinic already present is updated, not duplicated,
matched on (name, district).
"""
import argparse
import csv
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import logging
logging.disable(logging.CRITICAL)

from app import create_app          # noqa: E402
from config import Config           # noqa: E402
from models import database as db   # noqa: E402
from models import prospects as P   # noqa: E402

_TRUE = {"1", "y", "yes", "true", "t", "نعم", "ايوه", "أيوه"}
_INT_COLS = ("branches", "vets")
_BOOL_COLS = ("is_hospital", "has_grooming", "has_boarding",
              "has_pharmacy", "has_petshop", "has_lab")
_TEXT_COLS = ("name", "name_ar", "governorate", "district", "address", "phone",
              "whatsapp", "email", "website", "facebook", "instagram",
              "maps_url", "contact_name", "contact_role", "current_software",
              "notes", "source", "source_url", "status")


def _clean(raw: dict) -> dict:
    out = {}
    for k, v in raw.items():
        key = (k or "").strip().lower().replace(" ", "_")
        val = (v or "").strip() if isinstance(v, str) else v
        if key in _TEXT_COLS:
            out[key] = val
        elif key in _BOOL_COLS:
            out[key] = 1 if str(val).strip().lower() in _TRUE else 0
        elif key in _INT_COLS:
            try:
                out[key] = int(str(val).strip())
            except (TypeError, ValueError):
                pass
    return out


def cmd_import(conn, path: str) -> int:
    if not os.path.isfile(path):
        print("No such file: %s" % path)
        return 2
    new = updated = skipped = 0
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        for raw in csv.DictReader(fh):
            row = _clean(raw)
            if not row.get("name"):
                skipped += 1
                continue
            row.setdefault("source", os.path.basename(path))
            if P.upsert(conn, row) == "new":
                new += 1
            else:
                updated += 1
    conn.commit()
    print("  new      : %d" % new)
    print("  updated  : %d" % updated)
    if skipped:
        print("  skipped  : %d (no name)" % skipped)
    return 0


def cmd_score(conn) -> int:
    n = P.rescore_all(conn)
    print("  rescored : %d clinic(s)" % n)
    return 0


def cmd_cohorts(conn, strategy: str) -> int:
    out = P.assign_cohorts(conn, strategy=strategy)
    if not out:
        print("  Nothing to assign - no clinics with status 'new'.")
        return 0
    for c in sorted(out):
        print("  cohort %d : %d clinic(s)" % (c, out[c]))
    print("")
    if strategy == "spread":
        print("  Spread deliberately: cohort 1 gets a MIX of scores, so the")
        print("  best accounts are not spent on the least practised pitch.")
        print("  Use --strategy top to work best-first instead.")
    return 0


def cmd_status(conn) -> int:
    s = P.summary(conn)
    print("  clinics mapped     : %d" % s["total"])
    print("  with a score > 0   : %d" % s["scored"])
    print("  reachable by phone : %d" % s["contactable"])
    if s["total"] and s["contactable"] < s["total"]:
        print("                       (%d have no number - they are on the map"
              % (s["total"] - s["contactable"]))
        print("                        but cannot yet be worked)")
    print("")
    if s["by_governorate"]:
        print("  by territory:")
        for g, n in list(s["by_governorate"].items())[:8]:
            print("    %-22s %d" % (g, n))
    if s["by_cohort"]:
        print("")
        print("  by cohort: " + ", ".join("%d=%d" % (k, v)
                                          for k, v in sorted(s["by_cohort"].items())))
    if s["by_status"]:
        print("  by stage : " + ", ".join(
            "%s=%d" % (k, s["by_status"][k])
            for k in P.STAGES if k in s["by_status"]))
    return 0


def cmd_why(conn, name: str) -> int:
    P.ensure_tables(conn)
    row = conn.execute("SELECT * FROM prospects WHERE name LIKE ? LIMIT 1",
                       ("%" + name + "%",)).fetchone()
    if not row:
        print("No clinic matching %r." % name)
        return 1
    row = dict(row)
    print("  %s  (%s, %s)" % (row["name"], row.get("district") or "?",
                              row.get("governorate") or "?"))
    print("  score %d" % (row.get("score") or 0))
    print("")
    for line in P.explain_score(row):
        print("    %s" % line)
    if row.get("current_software"):
        print("")
        print("  Already using: %s" % row["current_software"])
        print("  Not scored, deliberately - it points both ways. They have")
        print("  proven they will pay for software, and they are mid-contract.")
    return 0


def cmd_funnel(conn) -> int:
    """Pillar 6: coverage, conversion, efficiency.

    The number that matters is not how many clinics are mapped - it is how many
    calls buy one demo. Without that you cannot tell a bad pitch from a bad
    list, and you will keep doing whichever one is broken.

    Deliberately says NOT ENOUGH DATA rather than printing a ratio off three
    rows. A conversion rate computed from two conversations is a number that
    will be quoted back later as though it meant something.
    """
    P.ensure_tables(conn)
    counts = {r[0]: r[1] for r in conn.execute(
        "SELECT status, COUNT(*) FROM prospects GROUP BY status").fetchall()}
    total = sum(counts.values())
    if not total:
        print("  Nothing mapped yet.")
        return 0

    # Everything from this stage onward counts as having reached it.
    ORDER = ["new", "researching", "contacted", "conversation",
             "demo", "proposal", "won"]
    reached, seen = {}, 0
    for stage in reversed(ORDER):
        seen += counts.get(stage, 0)
        reached[stage] = seen
    lost = counts.get("lost", 0)
    unreachable = counts.get("unreachable", 0)

    print("  COVERAGE")
    print("    clinics mapped        : %d" % total)
    callable_n = conn.execute(
        "SELECT COUNT(*) FROM prospects WHERE COALESCE(phone,'') <> ''"
        " OR COALESCE(whatsapp,'') <> ''").fetchone()[0]
    print("    with a way to reach   : %d" % callable_n)
    if callable_n < total:
        print("    still need a number   : %d" % (total - callable_n))

    print("")
    print("  PIPELINE")
    labels = {"contacted": "contacted", "conversation": "had a conversation",
              "demo": "saw a demo", "proposal": "got a proposal", "won": "signed"}
    prev = None
    for stage in ("contacted", "conversation", "demo", "proposal", "won"):
        n = reached.get(stage, 0)
        line = "    %-20s %4d" % (labels[stage], n)
        if prev is not None and prev >= 10:
            line += "   %.0f%% of the previous step" % (100.0 * n / prev)
        elif prev is not None and prev > 0:
            line += "   (too few to rate)"
        print(line)
        prev = n
    if lost or unreachable:
        print("")
        print("    lost                 %4d" % lost)
        print("    unreachable          %4d" % unreachable)

    print("")
    print("  EFFICIENCY")
    contacted = reached.get("contacted", 0)
    demos = reached.get("demo", 0)
    won = reached.get("won", 0)
    if contacted < 10:
        print("    NOT ENOUGH DATA. %d clinic(s) contacted." % contacted)
        print("    A conversion rate off a handful of calls is a number that")
        print("    gets quoted back later as though it meant something.")
        print("    Come back at 10 and it will start being worth reading.")
    else:
        if demos:
            print("    calls per demo        : %.1f" % (contacted / float(demos)))
        else:
            print("    calls per demo        : no demo yet after %d calls" % contacted)
        if won:
            print("    calls per signature   : %.1f" % (contacted / float(won)))
            print("    demos per signature   : %.1f"
                  % (demos / float(won) if won else 0))
        else:
            print("    calls per signature   : nothing signed yet")

    print("")
    print("  WHAT TO READ THIS FOR")
    if contacted >= 10 and demos == 0:
        print("    Calls are happening and no demo has come out of them. That")
        print("    is a pitch problem or a list problem, and the two need")
        print("    different fixes - check whether the calls that went nowhere")
        print("    were the low scores or the high ones.")
    elif demos >= 3 and won == 0:
        print("    Demos are happening and nothing has closed. The product is")
        print("    getting seen; the problem is after the demo - price, trust,")
        print("    or no reason to decide today.")
    elif contacted == 0:
        print("    Nothing has been contacted. Every other number here is")
        print("    waiting on that one.")
    else:
        print("    Keep going. The first ratio worth trusting arrives at about")
        print("    ten conversations.")
    return 0


def cmd_update(conn, path: str, apply_it: bool) -> int:
    """Read a worked-on workbook back into the database.

    Without this the export is a printout: twenty calls get typed into Excel
    and the next export overwrites them. A market database that loses the
    afternoon's work is not a CRM foundation, it is a list.

    Only the columns a PERSON fills in during a call are read back - status,
    last contact, next action, notes, and any phone number found on the way.
    Score, cohort and the service signals are left alone: those are researched
    or computed, and a stale copy sitting in a spreadsheet must not overwrite
    a fresher one in the database.

    A row typed into the workbook that is not in the database is treated as a
    clinic discovered during a call and offered as new, because that is exactly
    when new clinics get discovered.
    """
    from openpyxl import load_workbook

    if not os.path.isfile(path):
        print("No such file: %s" % path)
        return 2

    P.ensure_tables(conn)
    wb = load_workbook(path, data_only=True)
    if "Call list" not in wb.sheetnames:
        print("  %s has no 'Call list' sheet - is it the exported workbook?" % path)
        return 2
    ws = wb["Call list"]

    # Find the header row rather than assuming it, so an inserted row at the
    # top - which a person WILL do - does not silently shift every column.
    head_row = None
    for r in range(1, 12):
        vals = [str(ws.cell(row=r, column=c).value or "").strip()
                for c in range(1, 20)]
        if "Clinic" in vals and "Score" in vals:
            head_row = r
            break
    if not head_row:
        print("  Could not find the header row in 'Call list'.")
        return 2
    cols = {str(ws.cell(row=head_row, column=c).value or "").strip(): c
            for c in range(1, 20)}

    # What a person edits during a call. Everything else is left alone.
    EDITABLE = {
        "Status": "status",
        "Last contact": "last_contact",
        "Next action": "next_action",
        "Notes": "notes",
        "Phone": "phone",
        "WhatsApp": "whatsapp",
    }

    changed, added, unmatched, untouched = [], [], [], 0
    for r in range(head_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=cols.get("Clinic", 3)).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        district = str(ws.cell(row=r, column=cols.get("District", 5)).value
                       or "").strip()

        row = conn.execute(
            "SELECT * FROM prospects WHERE name=? AND COALESCE(district,'')=?",
            (name, district)).fetchone()
        if not row:
            added.append((name, district))
            continue
        row = dict(row)

        diffs = {}
        for header, field in EDITABLE.items():
            if header not in cols:
                continue
            new = ws.cell(row=r, column=cols[header]).value
            new = "" if new is None else str(new).strip()
            if field in ("phone", "whatsapp") and new:
                new = P.clean_phone(new)
            old = str(row.get(field) or "").strip()
            # An emptied cell is not treated as "delete this". People clear
            # cells by accident far more often than they mean to erase a
            # phone number that took twenty minutes to find.
            if new and new != old:
                diffs[field] = (old, new)
        if diffs:
            changed.append((row["id"], name, district, diffs))
        else:
            untouched += 1

    print("  rows in workbook : %d" % (ws.max_row - head_row))
    print("  unchanged        : %d" % untouched)
    print("  changed          : %d" % len(changed))
    if added:
        print("  not in database  : %d  (typed in during a call?)" % len(added))

    for _pid, name, district, diffs in changed[:25]:
        print("")
        print("    %s%s" % (name, (" - " + district) if district else ""))
        for field, (old, new) in diffs.items():
            print("      %-13s %s -> %s" % (field, old or "(blank)", new))
    if len(changed) > 25:
        print("")
        print("    ... and %d more" % (len(changed) - 25))

    if added:
        print("")
        print("  These are in the workbook but not the database:")
        for name, district in added[:12]:
            print("    %s%s" % (name, (" - " + district) if district else ""))
        print("")
        print("  Add them with a CSV import - this command only updates rows")
        print("  it can match, so a typo in a clinic name cannot quietly")
        print("  create a duplicate.")

    if not changed:
        print("")
        print("  Nothing to write back.")
        return 0

    if not apply_it:
        print("")
        print("  Re-run with --apply to save these. Nothing has been changed.")
        return 0

    for pid, _name, _district, diffs in changed:
        sets = ", ".join("%s=?" % f for f in diffs)
        conn.execute(
            "UPDATE prospects SET %s, updated_at=datetime('now') WHERE id=?"
            % sets, tuple(v[1] for v in diffs.values()) + (pid,))
    conn.commit()
    print("")
    print("  Wrote %d clinic(s) back." % len(changed))
    print("  Re-export whenever you want a fresh workbook - your notes are now")
    print("  in the database and will come back out with it.")
    return 0


def cmd_export(conn, path: str) -> int:
    """A workbook to actually work from.

    A CLI is the wrong tool for somebody with a phone in one hand. Three
    sheets: what to do today, the market from above, and where every fact came
    from.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    P.ensure_tables(conn)
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM prospects ORDER BY governorate, district,"
        " score DESC, name").fetchall()]
    if not rows:
        print("  Nothing to export yet - import a CSV first.")
        return 1

    F = "Arial"
    INK, BRAND = "1A1A1A", "1B6B5C"
    HDR = Font(name=F, size=10, bold=True, color="FFFFFF")
    BODY = Font(name=F, size=10, color=INK)
    BOLD = Font(name=F, size=10, bold=True, color=INK)
    LINK = Font(name=F, size=10, color="0563C1", underline="single")
    SMALL = Font(name=F, size=9, color="5B7169")
    H1 = Font(name=F, size=14, bold=True, color="0D2B24")
    FILL_HDR = PatternFill("solid", fgColor=BRAND)
    # Score bands. Green is not decoration - it means "worth the drive".
    BAND = {"hot": PatternFill("solid", fgColor="D5F0E4"),
            "warm": PatternFill("solid", fgColor="FDF6E3"),
            "cold": PatternFill("solid", fgColor="FFFFFF")}
    NOCALL = PatternFill("solid", fgColor="FDECEA")
    THIN = Side(style="thin", color="E2E8F0")
    BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TOP = Alignment(vertical="top")
    WRAP = Alignment(vertical="top", wrap_text=True)
    CENTRE = Alignment(horizontal="center", vertical="top")

    def band(score):
        return "hot" if score >= 8 else ("warm" if score >= 3 else "cold")

    def signals(r):
        got = []
        if int(r.get("branches") or 1) > 1:
            got.append("%s branches" % r["branches"])
        for key, label in (("is_hospital", "hospital"),
                           ("has_grooming", "grooming"),
                           ("has_boarding", "boarding"),
                           ("has_pharmacy", "pharmacy"),
                           ("has_petshop", "shop"), ("has_lab", "lab")):
            if int(r.get(key) or 0):
                got.append(label)
        if r.get("vets"):
            got.append("%s vets" % r["vets"])
        return " · ".join(got)

    wb = Workbook()

    # 1. Call list
    ws = wb.active
    ws.title = "Call list"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Who to call, in the order to call them"
    ws["A1"].font = H1
    ws["A2"] = ("Grouped by district first, then by score. Five clinics in one "
                "district in an afternoon beats five scattered across Cairo - "
                "vets in a district talk to each other, and that is the whole "
                "mechanism. Green = 8+, worth the drive. Red = no number yet.")
    ws["A2"].font = SMALL
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 30

    heads = ["Cohort", "Score", "Clinic", "الاسم",
             "District", "Phone", "WhatsApp", "What they run", "Already using",
             "Reachable", "Status", "Last contact", "Next action", "Notes"]
    widths = [7, 6, 32, 26, 15, 15, 15, 32, 14, 13, 13, 13, 28, 46]
    HROW = 4
    for i, (h, w) in enumerate(zip(heads, widths), start=1):
        c = ws.cell(row=HROW, column=i, value=h)
        c.font, c.fill, c.border = HDR, FILL_HDR, BOX
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HROW].height = 28
    ws.freeze_panes = "C%d" % (HROW + 1)

    r = HROW + 1
    for row in rows:
        score = row.get("score") or 0
        level, _missing = P.completeness(row)
        vals = [row.get("cohort"), score, row.get("name"), row.get("name_ar"),
                row.get("district"), row.get("phone"), row.get("whatsapp"),
                signals(row), row.get("current_software"), level,
                row.get("status"), row.get("last_contact"),
                row.get("next_action"), row.get("notes")]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BOLD if i == 3 else BODY
            c.border = BOX
            c.alignment = (WRAP if i in (8, 14)
                           else (CENTRE if i in (1, 2, 10) else TOP))
            c.fill = NOCALL if level != "callable" else BAND[band(score)]
        # Tapping the number should start the call, not select the text.
        if row.get("phone"):
            cell = ws.cell(row=r, column=6)
            cell.hyperlink = "tel:%s" % row["phone"]
            cell.font = LINK
        if row.get("whatsapp"):
            wa = str(row["whatsapp"]).lstrip("0")
            cell = ws.cell(row=r, column=7)
            cell.hyperlink = "https://wa.me/20%s" % wa
            cell.font = LINK
        ws.row_dimensions[r].height = 30
        r += 1
    ws.auto_filter.ref = "A%d:%s%d" % (HROW, get_column_letter(len(heads)), r - 1)

    # 2. Overview
    ov = wb.create_sheet("Overview")
    ov.sheet_view.showGridLines = False
    ov["A1"] = "The market, from above"
    ov["A1"].font = H1
    ov.column_dimensions["A"].width = 30
    ov.column_dimensions["B"].width = 12
    ov.column_dimensions["C"].width = 64

    callable_n = sum(1 for x in rows if P.completeness(x)[0] == "callable")
    research_n = sum(1 for x in rows if P.completeness(x)[0] == "researchable")
    hot = sum(1 for x in rows if (x.get("score") or 0) >= 8)
    on_comp = sum(1 for x in rows if (x.get("current_software") or "").strip())

    facts = [
        ("Clinics mapped", len(rows),
         "every one carries the URL its facts came from"),
        ("Callable today", callable_n, "a phone or WhatsApp number is on file"),
        ("Need a number", research_n,
         "an address or Facebook page but no phone - minutes each to find"),
        ("Score 8 or more", hot,
         "multi-branch or full-service: the clinics Aleefy's breadth is FOR"),
        ("Already on a competitor", on_comp,
         "willingness to pay is proven, and a contract is in the way"),
    ]
    rr = 3
    for label, value, note in facts:
        ov.cell(row=rr, column=1, value=label).font = BOLD
        c = ov.cell(row=rr, column=2, value=value)
        c.font = Font(name=F, size=12, bold=True, color=BRAND)
        c.alignment = CENTRE
        ov.cell(row=rr, column=3, value=note).font = SMALL
        rr += 1

    rr += 1
    ov.cell(row=rr, column=1, value="By district").font = H1
    rr += 1
    for h, i in (("District", 1), ("Clinics", 2), ("Callable", 3)):
        c = ov.cell(row=rr, column=i, value=h)
        c.font, c.fill = HDR, FILL_HDR
    rr += 1
    by_d = {}
    for x in rows:
        d = x.get("district") or "(not recorded)"
        got = by_d.setdefault(d, [0, 0])
        got[0] += 1
        got[1] += 1 if P.completeness(x)[0] == "callable" else 0
    for d, (n, ca) in sorted(by_d.items(), key=lambda kv: -kv[1][0]):
        ov.cell(row=rr, column=1, value=d).font = BODY
        ov.cell(row=rr, column=2, value=n).font = BODY
        ov.cell(row=rr, column=3, value=ca).font = BODY
        rr += 1
    ov.cell(row=rr + 1, column=1,
            value="Work one district at a time. Density is the mechanism - the "
                  "goal is a vet saying 'I know several clinics using it', "
                  "which never happens from one clinic per area.").font = SMALL
    ov.cell(row=rr + 1, column=1).alignment = WRAP
    ov.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 2, end_column=3)

    # 3. Sources
    src = wb.create_sheet("Sources")
    src.sheet_view.showGridLines = False
    src["A1"] = "Where every record came from"
    src["A1"].font = H1
    src["A2"] = ("A market database nobody can audit is a list of rumours, and "
                 "the first wrong phone number destroys trust in all of it.")
    src["A2"].font = SMALL
    for i, h in enumerate(["Clinic", "Source", "URL"], start=1):
        c = src.cell(row=4, column=i, value=h)
        c.font, c.fill = HDR, FILL_HDR
    for w, col in zip((32, 22, 80), "ABC"):
        src.column_dimensions[col].width = w
    for n, row in enumerate(rows, start=5):
        src.cell(row=n, column=1, value=row.get("name")).font = BODY
        src.cell(row=n, column=2, value=row.get("source")).font = BODY
        c = src.cell(row=n, column=3, value=row.get("source_url"))
        if row.get("source_url"):
            c.hyperlink = row["source_url"]
            c.font = LINK
        else:
            c.font = BODY
    src.freeze_panes = "A5"

    wb.save(path)
    print("  wrote %s" % path)
    print("  %d clinic(s)  |  %d callable today  |  %d need a number"
          % (len(rows), callable_n, len(rows) - callable_n))
    print("  %d score 8+ (multi-branch or full service)" % hot)
    return 0


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("command",
                    choices=["import", "score", "cohorts", "status", "export", "update", "why", "funnel"])
    ap.add_argument("arg", nargs="?", default="")
    ap.add_argument("--strategy", default="spread", choices=["spread", "top"],
                    help="cohort fill order (default: spread)")
    ap.add_argument("--apply", action="store_true",
                    help="actually write (update): default is to show only")
    a = ap.parse_args()

    app = create_app(Config)
    with app.app_context():
        conn = db.get_db()
        try:
            if a.command == "import":
                return cmd_import(conn, a.arg)
            if a.command == "score":
                return cmd_score(conn)
            if a.command == "cohorts":
                return cmd_cohorts(conn, a.strategy)
            if a.command == "status":
                return cmd_status(conn)
            if a.command == "update":
                return cmd_update(conn, a.arg or "market.xlsx", a.apply)
            if a.command == "export":
                return cmd_export(conn, a.arg or "market.xlsx")
            if a.command == "funnel":
                return cmd_funnel(conn)
            if a.command == "why":
                if not a.arg:
                    print("  why needs a clinic name")
                    return 2
                return cmd_why(conn, a.arg)
        finally:
            conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
